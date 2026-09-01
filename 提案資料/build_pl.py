# -*- coding: utf-8 -*-
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

F="Arial"
BLUE=Font(name=F,color="0000FF",size=10)          # 入力（変更可）
BLACK=Font(name=F,size=10)                         # 数式
GREEN=Font(name=F,color="008000",size=10)          # 他シート参照
H1=Font(name=F,bold=True,size=14)
H2=Font(name=F,bold=True,size=11,color="7B1E3A")
TH=Font(name=F,bold=True,size=10,color="FFFFFF")
BOLD=Font(name=F,bold=True,size=10)
YEL=PatternFill("solid",fgColor="FFFF00")
HDR=PatternFill("solid",fgColor="7B1E3A")
GREY=PatternFill("solid",fgColor="F2F2F2")
THIN=Border(bottom=Side(style="thin",color="BFBFBF"))
TOP=Border(top=Side(style="thin",color="000000"))
DBL=Border(top=Side(style="thin",color="000000"),bottom=Side(style="double",color="000000"))
MON='#,##0;(#,##0);-'
PCT='0.0%'
R=Alignment(horizontal="right"); C=Alignment(horizontal="center")

wb=Workbook()

# ============ Sheet1 前提 ============
s=wb.active; s.title="前提"
s.column_dimensions['A'].width=26
for c in "BCDEF": s.column_dimensions[c].width=14
s.column_dimensions['G'].width=42
s['A1']="WineBank ワインマイル｜前提条件"; s['A1'].font=H1
s['A2']="青字＝入力値（変更可）／黒字＝数式。単位：万円"; s['A2'].font=Font(name=F,size=9,color="808080")

s['A4']="■ 販売ミックス"; s['A4'].font=H2
hdr=["ランク","単価","人数","販売額","マイル還元率","マイル発行額"]
for i,h in enumerate(hdr):
    c=s.cell(4+1,1+i,h); c.font=TH; c.fill=HDR; c.alignment=C
rows=[("PRESTIGE",100,52,0.03),("GOLD",400,7,0.04),("SIGNATURE",1000,2,0.05)]
for i,(n,u,cnt,r) in enumerate(rows):
    rr=6+i
    s.cell(rr,1,n).font=BLACK
    s.cell(rr,2,u).font=BLUE;  s.cell(rr,2).number_format=MON
    s.cell(rr,3,cnt).font=BLUE; s.cell(rr,3).number_format=MON
    s.cell(rr,4,f"=B{rr}*C{rr}").font=BLACK; s.cell(rr,4).number_format=MON
    s.cell(rr,5,r).font=BLUE;  s.cell(rr,5).number_format=PCT
    s.cell(rr,6,f"=D{rr}*E{rr}").font=BLACK; s.cell(rr,6).number_format=MON
    for cc in range(1,7): s.cell(rr,cc).border=THIN
s.cell(9,1,"合計").font=BOLD
s.cell(9,3,"=SUM(C6:C8)").font=BLACK; s.cell(9,3).number_format=MON
s.cell(9,4,"=SUM(D6:D8)").font=BOLD;  s.cell(9,4).number_format=MON
s.cell(9,5,"=IF(D9=0,0,F9/D9)").font=BLACK; s.cell(9,5).number_format=PCT
s.cell(9,6,"=SUM(F6:F8)").font=BOLD;  s.cell(9,6).number_format=MON
for cc in range(1,7): s.cell(9,cc).border=TOP
s['G9']="← 還元率3〜5%の加重平均＝マイル発行の額面率"; s['G9'].font=Font(name=F,size=9,color="808080")

s['A11']="■ 事業前提"; s['A11'].font=H2
ass=[("ワイン値上がり率（年）",0.06,PCT,"P1ファンド実績・経費控除後"),
     ("借入金利（年）",0.03,PCT,"自社在庫の調達コスト"),
     ("保管料・保険料率（年）",0.015,PCT,"時価ベース。管理手数料2.5%の内数"),
     ("販売粗利率",0.30,PCT,"売価ベース。原価率は70%"),
     ("管理手数料率（年）",0.025,PCT,"預かり資産に対して"),
     ("成功報酬率",0.25,PCT,"値上がり益に対して。契約終了時に精算"),
     ("期首ワイン在庫（簿価）",10000,MON,"自社保有分"),
     ("新規買付額",12000,MON,"販売代金を原資に買い増す額")]
for i,(n,v,fmt,note) in enumerate(ass):
    rr=12+i
    s.cell(rr,1,n).font=BLACK
    c=s.cell(rr,2,v); c.font=BLUE; c.number_format=fmt; c.fill=YEL; c.alignment=R
    s.cell(rr,7,note).font=Font(name=F,size=9,color="808080")
    s.cell(rr,1).border=THIN; s.cell(rr,2).border=THIN

s['A21']="■ マイル交換ミックス（どこで使われるかの想定）"; s['A21'].font=H2
for i,h in enumerate(["交換先","構成比","交換レート","当社原価率"]):
    c=s.cell(22,1+i,h); c.font=TH; c.fill=HDR; c.alignment=C
ch=[("系列レストラン",0.35,1.0,0.90,"額面×1.0×90%保証"),
    ("グランメゾン",0.20,0.5,0.45,"額面×0.5×90%保証"),
    ("WineBank CLUB 会費充当",0.15,1.0,0.00,"内部振替。現金流出なし"),
    ("オークション成約手数料",0.10,1.0,0.00,"手数料の免除。現金流出なし"),
    ("ワイン追加購入",0.05,1.0,0.80,"ワインの仕入原価"),
    ("失効",0.15,0.0,0.00,"未使用のまま期限到来")]
for i,(n,w,rate,cost,note) in enumerate(ch):
    rr=23+i
    s.cell(rr,1,n).font=BLACK
    s.cell(rr,2,w).font=BLUE; s.cell(rr,2).number_format=PCT
    s.cell(rr,3,rate).font=BLUE; s.cell(rr,3).number_format='0.00"円"'
    s.cell(rr,4,cost).font=BLUE; s.cell(rr,4).number_format=PCT
    s.cell(rr,7,note).font=Font(name=F,size=9,color="808080")
    for cc in range(1,5): s.cell(rr,cc).border=THIN
s.cell(29,1,"合計 / 加重平均").font=BOLD
s.cell(29,2,"=SUM(B23:B28)").font=BOLD; s.cell(29,2).number_format=PCT
s.cell(29,4,"=SUMPRODUCT(B23:B28,D23:D28)").font=BOLD; s.cell(29,4).number_format=PCT
for cc in range(1,5): s.cell(29,cc).border=TOP
s['G29']="← これが「マイル原価率」。構成比の合計は100%にすること"; s['G29'].font=Font(name=F,size=9,color="808080")

# ============ Sheet2 マイル計算 ============
m=wb.create_sheet("マイル計算")
m.column_dimensions['A'].width=28
for c in "BCDE": m.column_dimensions[c].width=15
m.column_dimensions['F'].width=44
m['A1']="マイル計算｜額面（3〜5%）→ PL上の実費用（1.6%）の橋渡し"; m['A1'].font=H1
m['A2']="会員から見た還元率と、当社のPLに乗る費用は別物です。差は「どこで使われるか」で決まります。"
m['A2'].font=Font(name=F,size=9,color="808080")

m['A4']="STEP 1｜ランク別のマイル発行額（額面）"; m['A4'].font=H2
for i,h in enumerate(["ランク","販売額","還元率","マイル発行額"]):
    c=m.cell(5,1+i,h); c.font=TH; c.fill=HDR; c.alignment=C
for i,n in enumerate(["PRESTIGE","GOLD","SIGNATURE"]):
    rr=6+i; sr=6+i
    m.cell(rr,1,f"='前提'!A{sr}").font=GREEN
    m.cell(rr,2,f"='前提'!D{sr}").font=GREEN; m.cell(rr,2).number_format=MON
    m.cell(rr,3,f"='前提'!E{sr}").font=GREEN; m.cell(rr,3).number_format=PCT
    m.cell(rr,4,f"=B{rr}*C{rr}").font=BLACK;  m.cell(rr,4).number_format=MON
    for cc in range(1,5): m.cell(rr,cc).border=THIN
m.cell(9,1,"合計").font=BOLD
m.cell(9,2,"=SUM(B6:B8)").font=BOLD; m.cell(9,2).number_format=MON
m.cell(9,3,"=IF(B9=0,0,D9/B9)").font=BOLD; m.cell(9,3).number_format=PCT
m.cell(9,4,"=SUM(D6:D8)").font=BOLD; m.cell(9,4).number_format=MON
for cc in range(1,5): m.cell(9,cc).border=TOP
m['F9']="← 額面の発行率。会員が「還元率」として認識する数字"; m['F9'].font=Font(name=F,size=9,color="808080")

m['A11']="STEP 2｜交換先ごとに、実際に出ていく金額"; m['A11'].font=H2
for i,h in enumerate(["交換先","構成比","額面","原価率","実費用"]):
    c=m.cell(12,1+i,h); c.font=TH; c.fill=HDR; c.alignment=C
for i in range(6):
    rr=13+i; sr=23+i
    m.cell(rr,1,f"='前提'!A{sr}").font=GREEN
    m.cell(rr,2,f"='前提'!B{sr}").font=GREEN; m.cell(rr,2).number_format=PCT
    m.cell(rr,3,f"=$D$9*B{rr}").font=BLACK;   m.cell(rr,3).number_format=MON
    m.cell(rr,4,f"='前提'!D{sr}").font=GREEN; m.cell(rr,4).number_format=PCT
    m.cell(rr,5,f"=C{rr}*D{rr}").font=BLACK;  m.cell(rr,5).number_format=MON
    for cc in range(1,6): m.cell(rr,cc).border=THIN
m.cell(19,1,"合計 / 加重平均").font=BOLD
m.cell(19,2,"=SUM(B13:B18)").font=BOLD; m.cell(19,2).number_format=PCT
m.cell(19,3,"=SUM(C13:C18)").font=BOLD; m.cell(19,3).number_format=MON
m.cell(19,4,"=IF(C19=0,0,E19/C19)").font=BOLD; m.cell(19,4).number_format=PCT
m.cell(19,5,"=SUM(E13:E18)").font=BOLD; m.cell(19,5).number_format=MON
for cc in range(1,6): m.cell(19,cc).border=TOP
m['F19']="← マイル原価率（加重平均）と、PLに乗る実費用"; m['F19'].font=Font(name=F,size=9,color="808080")

m['A21']="STEP 3｜まとめ"; m['A21'].font=H2
sm=[("会員が受け取る額面（対 販売額）","=IF(B9=0,0,D9/B9)",PCT,"会員から見た還元率。3〜5%の加重平均"),
    ("マイル原価率（加重平均）","=D19",PCT,"額面100に対して実際に出ていく割合"),
    ("PL上の実費用（対 販売額）","=IF(B9=0,0,E19/B9)",PCT,"＝ 額面率 × 原価率。これがPLに乗る"),
    ("PL上の実費用（金額）","=E19",MON,"")]
for i,(n,f,fmt,note) in enumerate(sm):
    rr=22+i
    m.cell(rr,1,n).font=BOLD if i>=2 else BLACK
    c=m.cell(rr,2,f); c.font=BOLD if i>=2 else BLACK; c.number_format=fmt; c.alignment=R
    if i>=2: c.fill=YEL
    m.cell(rr,6,note).font=Font(name=F,size=9,color="808080")
    m.cell(rr,1).border=THIN; m.cell(rr,2).border=THIN

# ============ Sheet3 損益計算書 ============
p=wb.create_sheet("損益計算書")
p.column_dimensions['A'].width=30
for c in "BCD": p.column_dimensions[c].width=15
p.column_dimensions['E'].width=46
p['A1']="損益計算書｜WineBank単体"; p['A1'].font=H1
p['A2']="初年度＝1億販売＋1.2億買付／2年目以降＝追加販売なし（ストックのみ）。単位：万円"
p['A2'].font=Font(name=F,size=9,color="808080")
for i,h in enumerate(["","初年度","2年目以降"]):
    c=p.cell(4,1+i,h); c.font=TH; c.fill=HDR; c.alignment=C

def row(r,label,f1,f2,fmt=MON,bold=False,note="",border=THIN,fill=None):
    p.cell(r,1,label).font=BOLD if bold else BLACK
    for col,f in ((2,f1),(3,f2)):
        c=p.cell(r,col,f); c.font=BOLD if bold else BLACK
        c.number_format=fmt; c.alignment=R
        if fill: c.fill=fill
    p.cell(r,5,note).font=Font(name=F,size=9,color="808080")
    for cc in range(1,4): p.cell(r,cc).border=border

row(5,"【売上高】",None,None,MON,True,"",None)
row(6,"　ワイン販売売上","='前提'!D9",0,MON,False,"1億円分を販売")
row(7,"　管理手数料収入","='前提'!D9*'前提'!B16","='前提'!D9*'前提'!B16",MON,False,"預かり資産 × 2.5%")
row(8,"　売上高 計","=SUM(B6:B7)","=SUM(C6:C7)",MON,True,"",TOP)
row(9,"【売上原価】",None,None,MON,True,"",None)
row(10,"　ワイン売上原価","=-B6*(1-'前提'!B15)",0,MON,False,"粗利率30% → 原価率70%")
row(11,"　売上総利益","=B8+B10","=C8+C10",MON,True,"",TOP)
row(12,"【販売費及び一般管理費】",None,None,MON,True,"",None)
row(13,"　マイル費用","=-マイル計算!E19","=-マイル計算!E19",MON,False,"額面368 × 原価率44.5%")
row(14,"　保管料・保険料（預かり資産）","=-'前提'!D9*'前提'!B14","=-'前提'!D9*'前提'!B14",MON,False,"預かり資産 × 1.5%")
row(15,"　保管料・保険料（自社在庫）","=-('前提'!B18+'前提'!B19-B6*(1-'前提'!B15))*'前提'!B14",0,MON,False,"期末自社在庫 × 1.5%")
row(16,"　販管費 計","=SUM(B13:B15)","=SUM(C13:C15)",MON,True,"",TOP)
row(17,"営業利益","=B11+B16","=C11+C16",MON,True,"",TOP,GREY)
row(18,"【営業外費用】",None,None,MON,True,"",None)
row(19,"　支払利息","=-('前提'!B18+MAX(0,'前提'!B19-B6))*'前提'!B13",0,MON,False,"期末借入 × 3%")
row(20,"経常利益","=B17+B19","=C17+C19",MON,True,"",DBL,GREY)
row(22,"（参考）成功報酬の年平均発生額","='前提'!D9*(1+'前提'!B12)^5*0.25/5-'前提'!D9*0.25/5","='前提'!D9*((1+'前提'!B12)^5-1)*'前提'!B17/5",MON,False,"5年保有・値上がり益25%を年割り",None)
row(23,"経常利益＋成功報酬発生","=B20+B22","=C20+C22",MON,True,"",DBL,GREY)

p['A25']="■ 損益分岐：マイル原価率が何%を超えるとストックだけで赤字になるか"; p['A25'].font=H2
p['A26']="（管理手数料 − 保管料）÷ マイル発行額"; p['A26'].font=BLACK
p['B26']="=('前提'!D9*'前提'!B16-'前提'!D9*'前提'!B14)/マイル計算!D9"
p['B26'].font=BOLD; p['B26'].number_format=PCT; p['B26'].fill=YEL; p['B26'].alignment=R
p['E26']="← 現在の想定原価率と比較してください（前提シート D29）"; p['E26'].font=Font(name=F,size=9,color="808080")
p['A27']="現在の想定マイル原価率"; p['A27'].font=BLACK
p['B27']="=マイル計算!D19"; p['B27'].font=BOLD; p['B27'].number_format=PCT; p['B27'].alignment=R
p['A28']="判定"; p['A28'].font=BLACK
p['B28']='=IF(B27<=B26,"OK：ストックのみで黒字","要改善：ストックのみでは赤字")'
p['B28'].font=BOLD; p['B28'].alignment=R

wb.save("WineBank_ワインマイル_PL.xlsx")
print("saved")
