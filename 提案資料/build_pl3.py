# -*- coding: utf-8 -*-
"""WineBank ワインマイル 3か年PL（中谷氏版の構造を踏襲し、合意済み前提へ更新した統合版）"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUT = "WineBank_ワインマイル_3か年PL.xlsx"

NAVY  = "1A0E14"; BURG = "7B1E3A"; GOLD = "C9A227"
HDRF  = PatternFill("solid", fgColor="7B1E3A")
SECF  = PatternFill("solid", fgColor="EFE7EA")
INF   = PatternFill("solid", fgColor="FFF6D5")   # 入力セル
KPIF  = PatternFill("solid", fgColor="F4EFE6")
WHT   = Font(color="FFFFFF", bold=True, size=10)
B     = Font(bold=True)
SM    = Font(size=9, color="666666")
THIN  = Side(style="thin", color="CCCCCC")
BOX   = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

PCT="0.00%"; PCT1="0.0%"; MAN='#,##0.0'; MAN0='#,##0'; YEN='#,##0'; NUM='0.000'

wb = openpyxl.Workbook()

def sheet(name):
    ws = wb.create_sheet(name); return ws

def put(ws, cell, v, *, font=None, fill=None, fmt=None, align=None, border=False):
    c = ws[cell]; c.value = v
    if font: c.font = font
    if fill: c.fill = fill
    if fmt: c.number_format = fmt
    if align: c.alignment = Alignment(horizontal=align, vertical="center")
    if border: c.border = BOX
    return c

def title(ws, t, sub):
    put(ws, "A1", t, font=Font(bold=True, size=14, color=BURG))
    put(ws, "A2", sub, font=SM)

def sec(ws, row, t):
    put(ws, f"A{row}", t, font=Font(bold=True, size=11, color=BURG), fill=SECF)

def hdrow(ws, row, labels, start=1):
    for i, l in enumerate(labels):
        put(ws, f"{get_column_letter(start+i)}{row}", l, font=WHT, fill=HDRF,
            align="center", border=True)

def widths(ws, spec):
    for col, w in spec.items(): ws.column_dimensions[col].width = w

# ══════════════════════════════════════════ 前提
P = sheet("前提")
title(P, "WineBank ワインマイル｜3か年PL 前提条件",
      "黄色セル＝入力値（変更可）。金額単位：万円。中谷氏作成版の構造を踏襲し、8/31以降に合意した前提へ更新。")

sec(P, 4, "■ ランク別プラン・3か年販売計画")
hdrow(P, 5, ["ランク","1人あたり預かり額","管理・保管手数料率","マイル還元率",
             "1年目 新規人数","2年目 新規人数","3年目 新規人数","出典・備考"])
tiers = [("PRESTIGE",100,0.025,0.035,30,52,70),
         ("GOLD",     400,0.025,0.045, 5,10,15),
         ("SIGNATURE",1000,0.025,0.055,2, 3, 5)]
for i,(n,amt,fee,rate,y1,y2,y3) in enumerate(tiers):
    r = 6+i
    put(P,f"A{r}",n,font=B,border=True)
    put(P,f"B{r}",amt,fill=INF,fmt=MAN0,border=True)
    put(P,f"C{r}",fee,fill=INF,fmt=PCT,border=True)
    put(P,f"D{r}",rate,fill=INF,fmt=PCT,border=True)
    for j,v in enumerate((y1,y2,y3)):
        put(P,f"{get_column_letter(5+j)}{r}",v,fill=INF,fmt=MAN0,border=True)
    put(P,f"H{r}","還元率は 3.5/4.5/5.5%（100万会員の額面メリットを1.0%確保するため引上げ済）",font=SM)
put(P,"A9","合計新規会員数",font=B)
for j,c in enumerate("EFG"):
    put(P,f"{c}9",f"=SUM({c}6:{c}8)",fmt=MAN0,font=B)
    put(P,f"{c}10",f"=$B$6*{c}6+$B$7*{c}7+$B$8*{c}8",fmt=MAN0,font=B)
    put(P,f"{c}11",f"=IF({c}10=0,0,($B$6*$D$6*{c}6+$B$7*$D$7*{c}7+$B$8*$D$8*{c}8)/{c}10)",fmt=PCT)
    put(P,f"{c}12",1,fill=INF,fmt=PCT1)
    put(P,f"{c}13",f"={c}10*{c}12",fmt=MAN0)
put(P,"A10","新規預かり資産／販売対象額",font=B)
put(P,"A11","加重平均マイル還元率")
put(P,"A12","ワイン販売を伴う比率"); put(P,"H12","100%＝全件WineBank販売。既存ワイン持込型は比率を下げる",font=SM)
put(P,"A13","ワイン販売売上",font=B)

sec(P, 15, "■ 事業・獲得前提")
biz = [(16,"販売粗利率",0.30,PCT,"売価ベース30%"),
       (17,"ワイン値上がり率（年）",0.06,PCT,"実績6%"),
       (18,"成功報酬率",0.25,PCT,"値上がり益に対する成功報酬。合意値は25%"),
       (19,"預かり資産の保管・保険原価率（年）",0.0105,PCT,"鈴与実額 210円/本/年 ÷ 平均単価20,000円＝1.05%（倉持案の1億＝約7,619本より）"),
       (20,"年間預かり継続率",1.00,PCT1,"基本100%。解約を織り込む場合はここを下げる"),
       (21,"行動マイル追加発行率（対 預かり資産）",0.00,PCT,"落札・会食等の限定マイルは数量前提がないため0%"),
       (22,"CPL（1リード獲得費・万円）",0.8,'0.0',"8,000円"),
       (23,"CPA（1会員獲得費・営業込み・万円）",5.0,'0.0',"50,000円。CPLを内包し二重計上しない")]
for r,l,v,f,note in biz:
    put(P,f"A{r}",l); put(P,f"B{r}",v,fill=INF,fmt=f,border=True); put(P,f"H{r}",note,font=SM)
put(P,"A24","想定リード→会員CVR"); put(P,"B24","=IF(B23=0,0,B22/B23)",fmt=PCT1)
put(P,"A25","1会員獲得に必要なリード数"); put(P,"B25","=IF(B22=0,0,B23/B22)",fmt='0.00')

sec(P, 27, "■ オークション（売り手手数料0・買い手課金モデル）")
auc = [(28,"年間在庫回転率（対 預かり資産）",0.10,PCT1,"預かり資産のうち年間で売却に回る比率"),
       (29,"当社ルート通過率",0.50,PCT1,"他ルート売却の可能性を織り込み保守的に50%"),
       (30,"買い手手数料率",0.10,PCT1,"落札者から10%"),
       (31,"売り手手数料率",0.00,PCT1,"会員（売り手）からは取らない")]
for r,l,v,f,note in auc:
    put(P,f"A{r}",l); put(P,f"B{r}",v,fill=INF,fmt=f,border=True); put(P,f"H{r}",note,font=SM)
put(P,"A32","オークション収益率（対 預かり資産）",font=B)
put(P,"B32","=B28*B29*(B30+B31)",fmt=PCT,font=B)
put(P,"H32","回転率 × 当社ルート率 ×（買い手＋売り手手数料率）",font=SM)

sec(P, 34, "■ マイル消化前提（1年以内にほぼ利用）")
mil = [(35,"最終利用率（1年以内）",0.98,PCT1,"中谷案を踏襲。失効益をあてにしない保守設計"),
       (36,"付与年度内の利用率",0.50,PCT1,"年間で均等獲得する前提"),
       (37,"翌年度の利用率",0.48,PCT1,"残りを翌年度中に利用"),
       (38,"失効率",0.02,PCT1,"1年経過時点で未利用分が失効")]
for r,l,v,f,note in mil:
    put(P,f"A{r}",l); put(P,f"B{r}",v,fill=INF,fmt=f,border=True); put(P,f"H{r}",note,font=SM)
put(P,"A39","合計チェック（100%）",font=B); put(P,"B39","=SUM(B36:B38)",fmt=PCT1,font=B)

sec(P, 41, "■ 利用先ミックス（利用マイル100%に対する構成）")
hdrow(P, 42, ["交換先","構成比（入力）","利用マイル内構成比","交換レート","当社原価率",
              "利用1マイル当たり原価","利用1マイル当たり会員価値","備考"])
mix = [("系列レストラン",       27.8, 1.0, 0.90, "額面×1.0。原価率はオフピーク枠運用で引下げ余地あり（感度分析B参照）"),
       ("グランメゾン（1P=0.5円）",20.0, 0.5, 0.90, "権威性チャネル。レート0.5で当社負担を抑制"),
       ("WineBank CLUB 会費充当", 13.3, 1.0, 0.00, "内部振替。原価ほぼ0でレート1.0"),
       ("オークション参加",       13.3, 1.0, 0.00, "手数料充当。原価0かつ買い手手数料を誘発"),
       ("ワインスクール",         11.1, 1.0, 0.50, "自社開催。原価率50%"),
       ("会員交流イベント",        8.9, 1.0, 0.50, "自社開催。原価率50%"),
       ("ワイン追加購入",          5.6, 1.0, 0.80, "仕入原価80%。上限管理あり")]
R0 = 43; R1 = R0+len(mix)-1
for i,(n,share,rate,cost,note) in enumerate(mix):
    r = R0+i
    put(P,f"A{r}",n,border=True)
    put(P,f"B{r}",share,fill=INF,fmt='0.0',border=True)
    put(P,f"C{r}",f"=IF(SUM($B${R0}:$B${R1})=0,0,B{r}/SUM($B${R0}:$B${R1}))",fmt=PCT1,border=True)
    put(P,f"D{r}",rate,fill=INF,fmt='0.00',border=True)
    put(P,f"E{r}",cost,fill=INF,fmt=PCT1,border=True)
    put(P,f"F{r}",f"=D{r}*E{r}",fmt=NUM,border=True)
    put(P,f"G{r}",f"=D{r}",fmt=NUM,border=True)
    put(P,f"H{r}",note,font=SM)
RT = R1+1
put(P,f"A{RT}","合計 / 加重平均",font=B)
put(P,f"B{RT}",f"=SUM(B{R0}:B{R1})",fmt='0.0',font=B)
put(P,f"C{RT}",f"=SUM(C{R0}:C{R1})",fmt=PCT1,font=B)
put(P,f"F{RT}",f"=SUMPRODUCT(C{R0}:C{R1},F{R0}:F{R1})",fmt=NUM,font=B)
put(P,f"G{RT}",f"=SUMPRODUCT(C{R0}:C{R1},G{R0}:G{R1})",fmt=NUM,font=B)
EC = RT+1; EV = RT+2; MR = RT+3
put(P,f"A{EC}","発行1マイル当たり 当社期待原価",font=B)
put(P,f"B{EC}",f"=B35*F{RT}",fmt=NUM,font=B,fill=KPIF)
put(P,f"H{EC}","最終利用率 × 利用時の加重平均原価",font=SM)
put(P,f"A{EV}","発行1マイル当たり 会員期待価値",font=B)
put(P,f"B{EV}",f"=B35*G{RT}",fmt=NUM,font=B,fill=KPIF)
put(P,f"H{EV}","最終利用率 × 利用時の加重平均交換レート",font=SM)
put(P,f"A{MR}","マイル実質原価率（対 預かり資産・1年目）")
put(P,f"B{MR}",f"=E11*B{EC}",fmt=PCT)

FX = MR+2
sec(P, FX, "■ 年次固定費・任意費用（単位：万円）")
hdrow(P, FX+1, ["項目","1年目","2年目","3年目","","","","出典・備考"])
fixed = [("人件費",0,0,0,"金額未提示のため0。入力するとPLへ反映"),
         ("システム開発・保守費",0,0,0,"金額未提示のため0"),
         ("交流イベント・コミュニティ費",0,0,0,"マイル交換分は原価率50%で別途計上済。追加の持出があれば入力"),
         ("その他固定販管費",0,0,0,"金額未提示のため0"),
         ("自社在庫の保管・保険料",158,158,158,"自社在庫1.5億×1.05%＝158万。在庫を持ち続ける前提で3年計上"),
         ("支払利息",360,360,360,"買付借入1.2億×3%。借換継続の前提で3年計上")]
for i,(l,y1,y2,y3,note) in enumerate(fixed):
    r = FX+2+i
    put(P,f"A{r}",l,border=True)
    for j,v in enumerate((y1,y2,y3)):
        put(P,f"{get_column_letter(2+j)}{r}",v,fill=INF,fmt=MAN0,border=True)
    put(P,f"H{r}",note,font=SM)
IV = FX+2+len(fixed)+1
sec(P, IV, "■ 在庫投資（キャッシュフロー用・単位：万円）")
put(P,f"A{IV+1}","新規ワイン買付額",border=True)
for j,v in enumerate((12000,0,0)):
    put(P,f"{get_column_letter(2+j)}{IV+1}",v,fill=INF,fmt=MAN0,border=True)
put(P,f"H{IV+1}","1年目に1.2億相当を新規買付する前提。2〜3年目は方針未定のため0",font=SM)

widths(P, {"A":34,"B":16,"C":18,"D":14,"E":14,"F":20,"G":20,"H":74})

# 参照定数
FEE="'前提'!$C$6"; STO="'前提'!$B$19"; AUC="'前提'!$B$32"
ECC=f"'前提'!$B${EC}"; EVV=f"'前提'!$B${EV}"; UTL="'前提'!$B$35"
CMIX=f"'前提'!$C${R0}:$C${R1}"; DMIX=f"'前提'!$D${R0}:$D${R1}"; EMIX=f"'前提'!$E${R0}:$E${R1}"
FROW=FX+2   # 人件費行
IVR=IV+1

# ══════════════════════════════════════════ 会員・預かり資産
M = sheet("会員・預かり資産")
title(M,"会員数・預かり資産｜3か年推移","前提シートの年度別会員数を反映。獲得年から手数料・マイルを満額計上。単位：万円")
sec(M,4,"■ 新規獲得・ファネル")
hdrow(M,5,["項目","1年目","2年目","3年目","備考"])
YC = ["B","C","D"]; SRC = ["E","F","G"]
for i in range(3):
    put(M,f"{YC[i]}6",f"=前提!{SRC[i]}6",fmt=MAN0); put(M,f"{YC[i]}7",f"=前提!{SRC[i]}7",fmt=MAN0)
    put(M,f"{YC[i]}8",f"=前提!{SRC[i]}8",fmt=MAN0)
    put(M,f"{YC[i]}9",f"=SUM({YC[i]}6:{YC[i]}8)",fmt=MAN0,font=B)
    put(M,f"{YC[i]}10",f"=IF(前提!$B$22=0,0,{YC[i]}9*前提!$B$23/前提!$B$22)",fmt=MAN0)
    put(M,f"{YC[i]}11",f"=IF({YC[i]}10=0,0,{YC[i]}9/{YC[i]}10)",fmt=PCT1)
    put(M,f"{YC[i]}12",f"={YC[i]}9*前提!$B$23",fmt=MAN0)
    put(M,f"{YC[i]}13",f"=前提!{SRC[i]}10",fmt=MAN0,font=B)
    put(M,f"{YC[i]}14",f"=前提!{SRC[i]}12",fmt=PCT1)
    put(M,f"{YC[i]}15",f"={YC[i]}13*{YC[i]}14",fmt=MAN0,font=B)
    put(M,f"{YC[i]}16",f"=IF({YC[i]}9=0,0,{YC[i]}13/{YC[i]}9)",fmt=MAN)
for r,l,note in [(6,"PRESTIGE 新規会員数",""),(7,"GOLD 新規会員数",""),(8,"SIGNATURE 新規会員数",""),
                 (9,"新規会員数 合計",""),(10,"必要リード数","CPL 0.8万／CPA 5万から逆算"),
                 (11,"リード→会員CVR",""),(12,"ユーザー獲得費","新規会員数 × CPA 5万円"),
                 (13,"新規預かり資産",""),(14,"ワイン販売を伴う比率",""),
                 (15,"ワイン販売売上",""),(16,"新規会員1人あたり預かり資産","")]:
    put(M,f"A{r}",l,font=B if r in (9,13,15) else None); put(M,f"E{r}",note,font=SM)

sec(M,18,"■ 期末会員数")
for i,(r,src) in enumerate([(19,6),(20,7),(21,8)]):
    put(M,f"B{r}",f"=B{src}",fmt=MAN0)
    put(M,f"C{r}",f"=B{r}*前提!$B$20+C{src}",fmt=MAN0)
    put(M,f"D{r}",f"=C{r}*前提!$B$20+D{src}",fmt=MAN0)
for r,l in [(19,"PRESTIGE 期末会員数"),(20,"GOLD 期末会員数"),(21,"SIGNATURE 期末会員数")]:
    put(M,f"A{r}",l); put(M,f"E{r}","継続率反映",font=SM)
put(M,"A22","期末会員数 合計",font=B)
for c in YC: put(M,f"{c}22",f"=SUM({c}19:{c}21)",fmt=MAN0,font=B)

sec(M,24,"■ 預かり資産・ストック採算")
for r,src,tier in [(25,19,"$B$6"),(26,20,"$B$7"),(27,21,"$B$8")]:
    for c in YC: put(M,f"{c}{r}",f"={c}{src}*前提!{tier}",fmt=MAN0)
put(M,"A25","PRESTIGE 期末預かり資産"); put(M,"A26","GOLD 期末預かり資産"); put(M,"A27","SIGNATURE 期末預かり資産")
put(M,"A28","期末預かり資産 合計",font=B)
for c in YC: put(M,f"{c}28",f"=SUM({c}25:{c}27)",fmt=MAN0,font=B,fill=KPIF)
put(M,"A29","期首預かり資産"); put(M,"B29",0,fmt=MAN0); put(M,"C29","=B28",fmt=MAN0); put(M,"D29","=C28",fmt=MAN0)
put(M,"A30","継続後の期首預かり資産"); put(M,"B30",0,fmt=MAN0)
put(M,"C30","=B28*前提!$B$20",fmt=MAN0); put(M,"D30","=C28*前提!$B$20",fmt=MAN0)
put(M,"A31","新規預かり資産")
for c in YC: put(M,f"{c}31",f"={c}13",fmt=MAN0)
put(M,"A32","資産ブリッジ差額（0確認）")
for c in YC: put(M,f"{c}32",f"={c}30+{c}31-{c}28",fmt=MAN0)

rows = [(33,"管理・保管手数料収入", "={c}25*前提!$C$6+{c}26*前提!$C$7+{c}27*前提!$C$8", MAN0,"期末預かり資産×2.5%"),
        (34,"預かり資産の保管・保険原価","={c}28*前提!$B$19",MAN0,"期末預かり資産×1.05%（鈴与実額）"),
        (35,"オークション手数料収入","={c}28*前提!$B$32",MAN0,"買い手10%×当社ルート50%×回転10%"),
        (36,"マイル発行額（額面）","={c}25*前提!$D$6+{c}26*前提!$D$7+{c}27*前提!$D$8+{c}28*前提!$B$21",MAN0,"ランク別還元率＋行動マイル"),
        (37,"加重平均マイル発行率","=IF({c}28=0,0,{c}36/{c}28)",PCT,"実績ミックスで変動"),
        (38,"成功報酬の経済発生額（未実現）","={c}28*前提!$B$17*前提!$B$18",MAN0,"預かり資産×値上がり率6%×成功報酬率25%"),
        (39,"マイル費用（P/L見込計上）",f"={{c}}36*{ECC}",MAN0,"発行額×発行1マイル期待原価"),
        (40,"ストック損益（成功報酬前）","={c}33-{c}34+{c}35-{c}39",MAN0,"手数料－保管原価＋オークション－マイル費用"),
        (41,"ストック経済損益（成功報酬込）","={c}40+{c}38",MAN0,"成功報酬の経済発生を加算"),
        (42,"ストック利幅（成功報酬前）","=IF({c}28=0,0,{c}40/{c}28)",PCT,"現金・会計ベース"),
        (43,"ストック利幅（成功報酬込）","=IF({c}28=0,0,{c}41/{c}28)",PCT,"成功報酬を含む経済利幅")]
for r,l,f,fmt,note in rows:
    put(M,f"A{r}",l,font=B if r in (40,41,42) else None)
    for c in YC: put(M,f"{c}{r}",f.format(c=c),fmt=fmt,font=B if r in (40,41,42) else None)
    put(M,f"E{r}",note,font=SM)
widths(M,{"A":32,"B":14,"C":14,"D":14,"E":52})

# ══════════════════════════════════════════ マイル計算
K = sheet("マイル計算")
title(K,"マイル計算｜発行・利用・P/L費用・現金負担",
      "P/Lは発行時に最終利用率ベースの期待原価を計上。現金負担は実際に利用された年に表示。単位：万円")
sec(K,4,"■ 発行・消化スケジュール")
hdrow(K,5,["項目","1年目","2年目","3年目"]); put(K,"J5","備考",font=WHT,fill=HDRF,align="center")
put(K,"B6",0,fmt=MAN); put(K,"C6","=B12",fmt=MAN); put(K,"D6","=C12",fmt=MAN)
for i,c in enumerate(YC):
    put(K,f"{c}7",f"=会員・預かり資産!{c}36",fmt=MAN,font=B)
    put(K,f"{c}8",f"={c}7*前提!$B$36",fmt=MAN)
    put(K,f"{c}10",f"={c}8+{c}9",fmt=MAN,font=B)
    put(K,f"{c}12",f"={c}6+{c}7-{c}10-{c}11",fmt=MAN)
    put(K,f"{c}13",f"={c}7*{ECC}",fmt=MAN,font=B,fill=KPIF)
    put(K,f"{c}14",f"={c}10*'前提'!$F${RT}",fmt=MAN)
    put(K,f"{c}15",f"={c}13-{c}14",fmt=MAN)
    put(K,f"{c}17",f"={UTL}",fmt=PCT1)
put(K,"B9",0,fmt=MAN); put(K,"C9","=B7*前提!$B$37",fmt=MAN); put(K,"D9","=C7*前提!$B$37",fmt=MAN)
put(K,"B11",0,fmt=MAN); put(K,"C11","=B7*前提!$B$38",fmt=MAN); put(K,"D11","=C7*前提!$B$38",fmt=MAN)
put(K,"B16","=B15",fmt=MAN); put(K,"C16","=B16+C15",fmt=MAN); put(K,"D16","=C16+D15",fmt=MAN)
for r,l,note in [(6,"期首未使用マイル残高（額面）","前年期末残高"),
                 (7,"当年マイル発行額（額面）","期末預かり資産×ランク別還元率"),
                 (8,"当年発行分の当年度利用","年間均等獲得の前提で50%"),
                 (9,"前年度発行分の翌年度利用","前年コホートの48%"),
                 (10,"当年度の利用額 合計（額面）","サービス利用に充当された額面"),
                 (11,"前年度発行分の失効","前年コホートの2%"),
                 (12,"期末未使用マイル残高（額面）","翌年度に利用または失効"),
                 (13,"マイル費用（P/L見込計上）","将来利用分も含め発行年度に期待原価を計上"),
                 (14,"実利用による現金・原価負担","実際にサービス利用された年の原価"),
                 (15,"未払マイル原価の増減","P/L費用－実利用原価"),
                 (16,"期末未払マイル原価（負債）","将来利用に備えた原価見込み"),
                 (17,"コホート最終利用率","")]:
    put(K,f"A{r}",l,font=B if r in (7,10,13) else None); put(K,f"J{r}",note,font=SM)

sec(K,19,"■ 利用先別の実利用額・実コスト")
hdrow(K,20,["交換先","利用マイル内構成比","利用1マイル原価","1年目 額面","1年目 原価",
            "2年目 額面","2年目 原価","3年目 額面","3年目 原価","備考"])
for i in range(len(mix)):
    r = 21+i; sr = R0+i
    put(K,f"A{r}",f"=前提!A{sr}",border=True)
    put(K,f"B{r}",f"=前提!C{sr}",fmt=PCT1,border=True)
    put(K,f"C{r}",f"=前提!F{sr}",fmt=NUM,border=True)
    for j,(dc,ec_,tot) in enumerate([("D","E","$B$10"),("F","G","$C$10"),("H","I","$D$10")]):
        put(K,f"{dc}{r}",f"={tot}*$B{r}",fmt=MAN,border=True)
        put(K,f"{ec_}{r}",f"={dc}{r}*$C{r}",fmt=MAN,border=True)
RE = 21+len(mix)
put(K,f"A{RE}","合計",font=B)
put(K,f"B{RE}",f"=SUM(B21:B{RE-1})",fmt=PCT1,font=B)
for c in "DEFGHI": put(K,f"{c}{RE}",f"=SUM({c}21:{c}{RE-1})",fmt=MAN,font=B)
widths(K,{"A":30,"B":18,"C":16,"D":13,"E":13,"F":13,"G":13,"H":13,"I":13,"J":46})

# ══════════════════════════════════════════ 損益計算書
L = sheet("損益計算書")
title(L,"損益計算書｜WineBank ワインマイル事業 3か年","単位：万円。成功報酬は未実現のため営業利益と分けて表示。")
hdrow(L,4,["","1年目","2年目","3年目","3年累計","備考"])
def lrow(r,label,fml,fmt=MAN0,bold=False,note="",fill=None):
    put(L,f"A{r}",label,font=B if bold else None)
    for c in YC: put(L,f"{c}{r}",fml.format(c=c),fmt=fmt,font=B if bold else None,fill=fill)
    put(L,f"E{r}","=SUM(B{r}:D{r})".format(r=r) if fmt!=PCT else
        f"=IF(E9=0,0,E{r-1}/E9)",fmt=fmt,font=B if bold else None,fill=fill)
    put(L,f"F{r}",note,font=SM)

sec(L,5,"【売上高】")
lrow(6,"ワイン販売売上","=会員・預かり資産!{c}15",note="新規預かり資産×販売を伴う比率")
lrow(7,"管理・保管手数料収入","=会員・預かり資産!{c}33",note="期末預かり資産×2.5%")
lrow(8,"オークション手数料収入","=会員・預かり資産!{c}35",note="買い手10%・売り手0%・当社ルート50%")
lrow(9,"売上高 計","=SUM({c}6:{c}8)",bold=True)
sec(L,10,"【売上原価・売上総利益】")
lrow(11,"ワイン売上原価","={c}6*(1-前提!$B$16)",note="販売粗利率30%")
lrow(12,"売上総利益","={c}9-{c}11",bold=True)
put(L,"A13","売上総利益率")
for c in YC: put(L,f"{c}13",f"=IF({c}9=0,0,{c}12/{c}9)",fmt=PCT1)
put(L,"E13","=IF(E9=0,0,E12/E9)",fmt=PCT1)
sec(L,14,"【販売費及び一般管理費】")
lrow(15,"マイル費用（P/L見込計上）","=マイル計算!{c}13",note="発行額×発行1マイル期待原価")
lrow(16,"預かり資産の保管・保険原価","=会員・預かり資産!{c}34",note="期末預かり資産×1.05%")
lrow(17,"ユーザー獲得費","=会員・預かり資産!{c}12",note="新規会員数×CPA 5万円")
for i,l in enumerate(["人件費","システム開発・保守費","交流イベント・コミュニティ費","その他固定販管費","自社在庫の保管・保険料"]):
    lrow(18+i,l,"=前提!{c}"+str(FROW+i),note="前提シート入力")
lrow(23,"販管費 計","=SUM({c}15:{c}22)",bold=True)
lrow(24,"営業利益","={c}12-{c}23",bold=True,fill=KPIF)
put(L,"A25","営業利益率")
for c in YC: put(L,f"{c}25",f"=IF({c}9=0,0,{c}24/{c}9)",fmt=PCT1)
put(L,"E25","=IF(E9=0,0,E24/E9)",fmt=PCT1)
sec(L,26,"【営業外損益】")
lrow(27,"支払利息","=前提!{c}"+str(FROW+5),note="買付借入1.2億×3%")
lrow(28,"経常利益","={c}24-{c}27",bold=True,fill=KPIF)
put(L,"A29","経常利益率")
for c in YC: put(L,f"{c}29",f"=IF({c}9=0,0,{c}28/{c}9)",fmt=PCT1)
put(L,"E29","=IF(E9=0,0,E28/E9)",fmt=PCT1)
sec(L,31,"【参考：成功報酬の経済発生（契約終了時まで未実現）】")
lrow(32,"成功報酬の経済発生額","=会員・預かり資産!{c}38",note="預かり資産×値上がり率6%×成功報酬率25%")
lrow(33,"経済利益（経常利益＋成功報酬発生）","={c}28+{c}32",bold=True)

sec(L,35,"【★利益の源泉分解｜フロー依存度の確認】")
lrow(36,"ワイン販売粗利（フロー）","={c}6-{c}11",note="新規販売がある限りの利益")
lrow(37,"ストック損益（成功報酬前）","=会員・預かり資産!{c}40",note="手数料－保管－マイル＋オークション")
lrow(38,"ユーザー獲得費","=-会員・預かり資産!{c}12")
lrow(39,"固定費・自社在庫","=-SUM({c}18:{c}22)")
lrow(40,"営業利益（検算）","=SUM({c}36:{c}39)",bold=True)
put(L,"A41","フロー（販売粗利）依存度",font=B)
for c in YC: put(L,f"{c}41",f"=IF({c}24=0,0,{c}36/{c}24)",fmt=PCT1,font=B)
put(L,"E41","=IF(E24=0,0,E36/E24)",fmt=PCT1,font=B)
put(L,"F41","100%超＝新規販売が止まると赤字。ストック単独で黒字化するほど依存度は下がる",font=SM)

sec(L,43,"【ストック部分だけの採算】")
lrow(44,"ストック損益（成功報酬前）","=会員・預かり資産!{c}40",bold=True)
lrow(45,"ストック経済損益（成功報酬込）","=会員・預かり資産!{c}41",bold=True)
put(L,"A46","ストック利幅（成功報酬前）")
for c in YC: put(L,f"{c}46",f"=会員・預かり資産!{c}42",fmt=PCT)
put(L,"A47","ストック利幅（成功報酬込）")
for c in YC: put(L,f"{c}47",f"=会員・預かり資産!{c}43",fmt=PCT)
put(L,"A48","損益分岐となる発行1マイル期待原価",font=B)
for c in YC:
    put(L,f"{c}48",f"=IF(会員・預かり資産!{c}37=0,0,(前提!$C$6-前提!$B$19+前提!$B$32)/会員・預かり資産!{c}37)",fmt=NUM,font=B)
put(L,"A49","現在の発行1マイル期待原価",font=B)
for c in YC: put(L,f"{c}49",f"={ECC}",fmt=NUM,font=B)
put(L,"A50","判定",font=B)
for c in YC:
    put(L,f"{c}50",f'=IF({c}49<={c}48,"OK：ストック単体で黒字","要改善：成功報酬前は赤字")',font=B,fill=KPIF)
widths(L,{"A":32,"B":14,"C":14,"D":14,"E":14,"F":62})

# ══════════════════════════════════════════ キャッシュフロー
CF = sheet("キャッシュフロー")
title(CF,"キャッシュフロー｜3か年","成功報酬は契約終了時まで現金化しないため本表に含めない。単位：万円")
hdrow(CF,4,["項目","1年目","2年目","3年目","3年累計","備考"])
cfr = [(5,"営業利益","=損益計算書!{c}24",False,"P/Lより"),
       (6,"＋マイル費用（P/L計上・非現金分）","=マイル計算!{c}13",False,"発行時に費用計上した額を戻す"),
       (7,"▲実利用マイル原価（現金流出）","=-マイル計算!{c}14",False,"実際に使われた年の原価"),
       (8,"営業キャッシュフロー（在庫投資前）","=SUM({c}5:{c}7)",True,""),
       (9,"▲新規ワイン買付（在庫投資）","=-前提!{c}"+str(IVR),False,"1年目1.2億の買付"),
       (10,"▲支払利息","=-損益計算書!{c}27",False,""),
       (11,"純キャッシュフロー","={c}8+{c}9+{c}10",True,"")]
for r,l,f,bold,note in cfr:
    put(CF,f"A{r}",l,font=B if bold else None)
    for c in YC: put(CF,f"{c}{r}",f.format(c=c),fmt=MAN0,font=B if bold else None,
                     fill=KPIF if bold else None)
    put(CF,f"E{r}",f"=SUM(B{r}:D{r})",fmt=MAN0,font=B if bold else None,fill=KPIF if bold else None)
    put(CF,f"F{r}",note,font=SM)
put(CF,"A12","累計キャッシュフロー",font=B)
put(CF,"B12","=B11",fmt=MAN0,font=B); put(CF,"C12","=B12+C11",fmt=MAN0,font=B)
put(CF,"D12","=C12+D11",fmt=MAN0,font=B)
sec(CF,14,"【参考：未実現の経済価値】")
put(CF,"A15","成功報酬の累計発生額")
put(CF,"B15","=損益計算書!B32",fmt=MAN0); put(CF,"C15","=B15+損益計算書!C32",fmt=MAN0)
put(CF,"D15","=C15+損益計算書!D32",fmt=MAN0)
put(CF,"F15","契約終了時に精算。3年時点では現金化していない",font=SM)
put(CF,"A16","期末未払マイル原価（将来の現金流出）")
for c in YC: put(CF,f"{c}16",f"=マイル計算!{c}16",fmt=MAN0)
put(CF,"F16","翌年度以降に利用される分の原価見込み",font=SM)
widths(CF,{"A":34,"B":14,"C":14,"D":14,"E":14,"F":52})

# ══════════════════════════════════════════ 感度分析
S = sheet("感度分析")
title(S,"感度分析｜交換レート・原価率・利用率・販売比率",
      "中谷氏コメント「レストランは1.0でなく0.8でもよいのでは」への定量回答を含む。")

def mixcost(idx, rate=None, cost=None):
    """系列レストラン(idx=0)の rate / cost を差し替えた加重平均原価"""
    tgt = R0+idx
    r = rate if rate else f"'前提'!D{tgt}"
    e = cost if cost else f"'前提'!E{tgt}"
    others = f"SUMPRODUCT('前提'!C{tgt+1}:C{R1},'前提'!D{tgt+1}:D{R1},'前提'!E{tgt+1}:E{R1})"
    return f"{UTL}*('前提'!C{tgt}*{r}*{e}+{others})"
def mixval(idx, rate=None):
    tgt = R0+idx
    r = rate if rate else f"'前提'!D{tgt}"
    others = f"SUMPRODUCT('前提'!C{tgt+1}:C{R1},'前提'!D{tgt+1}:D{R1})"
    return f"{UTL}*('前提'!C{tgt}*{r}+{others})"

sec(S,4,"■ A. 系列レストランの交換レート別（原価率は90%のまま）")
hdrow(S,5,["交換レート","発行1マイル 当社原価","ストック利幅（成功報酬前）","ストック利幅（成功報酬込）",
           "100万会員の実質メリット率","100万会員の実質メリット額（円）"])
for i,rate in enumerate([1.0,0.9,0.8,0.7]):
    r=6+i
    put(S,f"A{r}",rate,fill=INF,fmt='0.00',border=True)
    put(S,f"B{r}","="+mixcost(0,rate=f"$A{r}"),fmt=NUM,border=True)
    put(S,f"C{r}",f"={FEE}-{STO}+{AUC}-会員・預かり資産!$B$37*B{r}",fmt=PCT,border=True)
    put(S,f"D{r}",f"=C{r}+前提!$B$17*前提!$B$18",fmt=PCT,border=True)
    put(S,f"E{r}",f"=前提!$D$6*("+mixval(0,rate=f"$A{r}")+f")-{FEE}",fmt=PCT,border=True)
    put(S,f"F{r}",f"=E{r}*前提!$B$6*10000",fmt=YEN,border=True)
put(S,"H6","レートを下げると当社利幅は改善するが、会員メリットが同率で削られる。",font=SM)
put(S,"H7","系列レストランは会員が「1マイル＝1円」を実感する中核チャネルであり、",font=SM)
put(S,"H8","ここを削ると訴求力が最も大きく毀損する。次表Bの方が費用対効果が高い。",font=SM)

sec(S,11,"■ B. 系列レストランの当社原価率別（交換レートは1.0のまま）")
hdrow(S,12,["当社原価率","発行1マイル 当社原価","ストック利幅（成功報酬前）","ストック利幅（成功報酬込）",
            "100万会員の実質メリット率","100万会員の実質メリット額（円）"])
for i,(cost,note) in enumerate([(0.90,"満席・機会損失前提（現状の置き方）"),
                                (0.70,"一部を現金併用（マイル充当上限50%）"),
                                (0.55,"平日オフピーク枠に限定（変動費中心）"),
                                (0.40,"F&B原価のみ（空席を埋める前提）")]):
    r=13+i
    put(S,f"A{r}",cost,fill=INF,fmt=PCT1,border=True)
    put(S,f"B{r}","="+mixcost(0,cost=f"$A{r}"),fmt=NUM,border=True)
    put(S,f"C{r}",f"={FEE}-{STO}+{AUC}-会員・預かり資産!$B$37*B{r}",fmt=PCT,border=True)
    put(S,f"D{r}",f"=C{r}+前提!$B$17*前提!$B$18",fmt=PCT,border=True)
    put(S,f"E{r}",f"=前提!$D$6*({mixval(0)})-{FEE}",fmt=PCT,border=True)
    put(S,f"F{r}",f"=E{r}*前提!$B$6*10000",fmt=YEN,border=True)
    put(S,f"H{r}",note,font=SM)
put(S,"H17","原価率を下げても会員メリット率は一切変わらない（E列は全行同値）のが要点。",font=Font(size=9,bold=True,color="7B1E3A"))

sec(S,19,"■ C. 最終利用率別（失効益への依存度）")
hdrow(S,20,["最終利用率","発行1マイル 当社原価","ストック利幅（成功報酬前）","ストック利幅（成功報酬込）",
            "100万会員の実質メリット率",""])
for i,u in enumerate([0.90,0.95,0.98,1.00]):
    r=21+i
    put(S,f"A{r}",u,fill=INF,fmt=PCT1,border=True)
    put(S,f"B{r}",f"=$A{r}*'前提'!$F${RT}",fmt=NUM,border=True)
    put(S,f"C{r}",f"={FEE}-{STO}+{AUC}-会員・預かり資産!$B$37*B{r}",fmt=PCT,border=True)
    put(S,f"D{r}",f"=C{r}+前提!$B$17*前提!$B$18",fmt=PCT,border=True)
    put(S,f"E{r}",f"=前提!$D$6*($A{r}*'前提'!$G${RT})-{FEE}",fmt=PCT,border=True)
put(S,"H21","中谷氏の「3〜6か月失効でコミュニティ通貨として回す」設計は利用率を上げる＝コストも上がる。",font=SM)
put(S,"H22","失効益をあてにしない98%前提でも黒字であることを確認するための表。",font=SM)

sec(S,26,"■ D. ワイン販売を伴う比率別の営業利益（持込型が増えた場合）")
hdrow(S,27,["販売を伴う比率","1年目","2年目","3年目","3年累計","見方"])
for i,(ratio,note) in enumerate([(0.00,"持込・預かりのみ"),(0.25,"販売粗利が一部発生"),
                                 (0.50,"販売と持込が半々"),(0.75,"販売中心"),(1.00,"前提シートの100%ケース")]):
    r=28+i
    put(S,f"A{r}",ratio,fill=INF,fmt=PCT1,border=True)
    for j,c in enumerate(YC):
        put(S,f"{c}{r}",f"=会員・預かり資産!{c}13*$A{r}*前提!$B$16+会員・預かり資産!{c}33"
                        f"+会員・預かり資産!{c}35-損益計算書!{c}23",fmt=MAN0,border=True)
    put(S,f"E{r}",f"=SUM(B{r}:D{r})",fmt=MAN0,border=True)
    put(S,f"F{r}",note,font=SM)
put(S,"A34","営業利益が黒字となる最低の販売伴有率",font=B)
put(S,"B34","=IF(SUM(会員・預かり資産!B13:D13)*前提!$B$16=0,0,"
            "MAX(0,(SUM(損益計算書!B23:D23)-SUM(会員・預かり資産!B33:D33)-SUM(会員・預かり資産!B35:D35))"
            "/(SUM(会員・預かり資産!B13:D13)*前提!$B$16)))",fmt=PCT1,font=B,fill=KPIF)
sec(S,36,"■ E. ランク別のストック採算（★還元率を上げた分、上位ランクほどストック単体は厳しい）")
hdrow(S,37,["ランク","マイル還元率","マイル実費用（対預かり）","ストック利幅（成功報酬前）",
            "ストック利幅（成功報酬込）","判定（成功報酬前）"])
for i in range(3):
    r=38+i; src=6+i
    put(S,f"A{r}",f"=前提!A{src}",border=True)
    put(S,f"B{r}",f"=前提!D{src}",fmt=PCT1,border=True)
    put(S,f"C{r}",f"=前提!D{src}*{ECC}",fmt=PCT,border=True)
    put(S,f"D{r}",f"=前提!C{src}-{STO}+{AUC}-C{r}",fmt=PCT,border=True)
    put(S,f"E{r}",f"=D{r}+前提!$B$17*前提!$B$18",fmt=PCT,border=True)
    put(S,f"F{r}",f'=IF(D{r}>=0,"ストック単体で黒字","成功報酬前は赤字")',border=True)
put(S,"H38","管理料2.5%は全ランク共通なのに還元率は3.5→5.5%と上がるため、",font=SM)
put(S,"H39","上位ランクほどストック単体の利幅は薄い。上位ランクは1人あたりの販売粗利と",font=SM)
put(S,"H40","成功報酬額が大きいことで正当化される構造であり、ストック採算だけで評価してはいけない。",font=SM)
put(S,"H41","SIGNATUREをストック単体で黒字にするには発行1マイル原価をB42まで下げる必要がある（原価率55%単独では届かない）。",
    font=Font(size=9,bold=True,color="7B1E3A"))
put(S,"H42","感度分析Bの原価率引下げと、0円原価チャネル（CLUB充当・オークション参加）の構成比引上げの併用で到達できる。",
    font=Font(size=9,bold=True,color="7B1E3A"))
put(S,"A42","参考：全ランクがストック単体で黒字となる発行1マイル期待原価の上限",font=B)
put(S,"B42",f"=IF(前提!$D$8=0,0,(前提!$C$8-{STO}+{AUC})/前提!$D$8)",fmt=NUM,font=B,fill=KPIF)
put(S,"C42",f"=\"現在は \"&TEXT({ECC},\"0.000\")",font=B)

widths(S,{"A":20,"B":22,"C":24,"D":24,"E":26,"F":28,"G":3,"H":58})

# ══════════════════════════════════════════ 会員メリット
V = sheet("会員メリット")
title(V,"会員メリット｜ランク別 金額シミュレーション","このシートのみ金額単位：円。会員1人あたり。")
sec(V,4,"■ 年間の直接メリット")
hdrow(V,5,["項目","PRESTIGE","GOLD","SIGNATURE","考え方・備考"])
TC=["B","C","D"]
for i,c in enumerate(TC):
    src=6+i
    put(V,f"{c}6",f"=前提!$B${src}*10000",fmt=YEN,font=B)
    put(V,f"{c}7",f"=前提!$C${src}",fmt=PCT1)
    put(V,f"{c}8",f"={c}6*{c}7",fmt=YEN)
    put(V,f"{c}9",f"=前提!$D${src}",fmt=PCT1)
    put(V,f"{c}10",f"={c}6*{c}9",fmt=YEN)
    put(V,f"{c}11",f"={UTL}",fmt=PCT1)
    put(V,f"{c}12",f"='前提'!$G${RT}",fmt=NUM)
    put(V,f"{c}13",f"={EVV}",fmt=NUM)
    put(V,f"{c}14",f"={c}10*{c}13",fmt=YEN,font=B)
    put(V,f"{c}15",f"={c}10-{c}8",fmt=YEN)
    put(V,f"{c}16",f"={c}14-{c}8",fmt=YEN,font=B,fill=KPIF)
    put(V,f"{c}17",f"=IF({c}6=0,0,{c}16/{c}6)",fmt=PCT,font=B,fill=KPIF)
    put(V,f"{c}18",f"={c}10*{c}11-{c}8",fmt=YEN)
    put(V,f"{c}19",f"=IF({c}10*{c}11=0,0,{c}8/({c}10*{c}11))",fmt=NUM)
for r,l,note in [(6,"預かり額",""),(7,"管理・保管手数料率",""),(8,"年間管理・保管手数料",""),
                 (9,"マイル還元率","3.5 / 4.5 / 5.5%"),(10,"年間付与マイル（額面）",""),
                 (11,"最終利用率（1年以内）",""),(12,"利用1マイルの平均交換価値","交換先ミックスの加重平均"),
                 (13,"発行1マイルの期待利用価値","平均交換価値 × 最終利用率"),
                 (14,"実利用ベースの年間マイル価値",""),
                 (15,"額面ベースの年間差引メリット","全マイルを1円相当として見た表示上の差引（＝還元率－手数料率）"),
                 (16,"★実利用ベースの年間差引メリット","交換先ミックスと利用率を反映した実質差引"),
                 (17,"★実質メリット率（預かり額比）",""),
                 (18,"1.0円交換先だけで使った場合の差引","グランメゾンを使わない会員の上限値"),
                 (19,"損益分岐となる平均交換価値","これを上回れば手数料を回収")]:
    put(V,f"A{r}",l,font=B if r in (16,17) else None); put(V,f"E{r}",note,font=SM)

sec(V,21,"■ 3年間保有した場合の会員メリット（経済価値ベース）")
hdrow(V,22,["項目","PRESTIGE","GOLD","SIGNATURE","考え方・備考"])
for c in TC:
    put(V,f"{c}23",f"={c}6*(1+前提!$B$17)^3",fmt=YEN)
    put(V,f"{c}24",f"={c}23-{c}6",fmt=YEN)
    put(V,f"{c}25","=前提!$B$18",fmt=PCT1)
    put(V,f"{c}26",f"={c}24*{c}25",fmt=YEN)
    put(V,f"{c}27",f"={c}24-{c}26",fmt=YEN,font=B)
    put(V,f"{c}28",f"={c}14*3",fmt=YEN)
    put(V,f"{c}29",f"={c}8*3",fmt=YEN)
    put(V,f"{c}30",f"={c}28-{c}29",fmt=YEN)
    put(V,f"{c}31",f"={c}27+{c}30",fmt=YEN,font=B,fill=KPIF)
    put(V,f"{c}32",f"=IF({c}6=0,0,{c}31/{c}6)",fmt=PCT1,font=B,fill=KPIF)
    put(V,f"{c}33",f"={c}6+{c}31",fmt=YEN)
    put(V,f"{c}34",f"=IF({c}6=0,0,({c}33/{c}6)^(1/3)-1)",fmt=PCT,font=B)
for r,l,note in [(23,"3年後ワイン評価額","預かり額×(1+6%)^3"),(24,"3年間の値上がり益",""),
                 (25,"成功報酬率","25%"),(26,"成功報酬（当社取分）",""),
                 (27,"会員に残る値上がり益",""),(28,"3年累計マイル利用価値",""),
                 (29,"3年累計管理・保管手数料",""),(30,"3年累計マイル差引メリット",""),
                 (31,"★3年間の総合メリット","値上がり益（会員取分）＋マイル差引"),
                 (32,"★元本に対する総合メリット率",""),
                 (33,"3年後の会員価値（元本＋総合メリット）",""),
                 (34,"年率換算（CAGR）","")]:
    put(V,f"A{r}",l,font=B if r in (31,32) else None); put(V,f"E{r}",note,font=SM)

sec(V,36,"■ 交換先別の会員価値（利用1万マイルあたり）")
hdrow(V,37,["交換先","利用マイル内構成比","交換レート","1万マイルの価値（円）","構成比加重価値（円）"])
for i in range(len(mix)):
    r=38+i; sr=R0+i
    put(V,f"A{r}",f"=前提!A{sr}",border=True)
    put(V,f"B{r}",f"=前提!C{sr}",fmt=PCT1,border=True)
    put(V,f"C{r}",f"=前提!D{sr}",fmt='0.00',border=True)
    put(V,f"D{r}",f"=10000*C{r}",fmt=YEN,border=True)
    put(V,f"E{r}",f"=10000*B{r}*C{r}",fmt=YEN,border=True)
VE=38+len(mix)
put(V,f"A{VE}","利用時の加重平均",font=B)
put(V,f"B{VE}",f"=SUM(B38:B{VE-1})",fmt=PCT1,font=B)
put(V,f"C{VE}",f"=SUM(E38:E{VE-1})/10000",fmt=NUM,font=B)
put(V,f"D{VE}",f"=10000*C{VE}",fmt=YEN,font=B)
put(V,f"E{VE}",f"=SUM(E38:E{VE-1})",fmt=YEN,font=B)
put(V,f"A{VE+1}","発行1万マイルの期待価値（最終利用率反映）",font=B)
put(V,f"D{VE+1}",f"=10000*{EVV}",fmt=YEN,font=B,fill=KPIF)
NT=VE+3
sec(V,NT,"■ 算定条件・注意点")
for i,txt in enumerate([
    '="・実利用ベースは、最終利用率 "&TEXT(前提!$B$35,"0.0%")&" と交換先別レートの加重平均を反映しています。"',
    '="・3年試算は預かり額を一定とし、ワイン時価のみ年 "&TEXT(前提!$B$17,"0.0%")&" で複利成長。値上がり益の "&TEXT(前提!$B$18,"0.0%")&" を成功報酬として控除しています。"',
    "・マイル価値はサービス利用時の経済価値であり、現金で受け取る利益ではありません。",
    "・行動マイル、未公開ワイン情報、イベント優先案内などは保守的に未算入です。",
    "・市場価格の下落、売却時の手数料、税金等は未考慮です。"]):
    put(V,f"A{NT+1+i}",txt,font=SM)
widths(V,{"A":38,"B":18,"C":18,"D":20,"E":60})

# ══════════════════════════════════════════ ダッシュボード
D = sheet("ダッシュボード")
title(D,"WineBank ワインマイル｜3か年PL ダッシュボード","成功報酬は未実現のため経常利益と分けて表示。単位：万円")
kpi = [("A",4,"3年末 預かり資産","=会員・預かり資産!D28",MAN0),
       ("C",4,"3年末 会員数","=会員・預かり資産!D22",MAN0),
       ("E",4,"3年累計 売上高","=損益計算書!E9",MAN0),
       ("G",4,"3年累計 経常利益","=損益計算書!E28",MAN0),
       ("A",7,"3年累計 経済利益（成功報酬込）","=損益計算書!E33",MAN0),
       ("C",7,"発行1マイル 当社期待原価",f"={ECC}",NUM),
       ("E",7,"ストック利幅（成功報酬前）","=会員・預かり資産!D42",PCT),
       ("G",7,"ストック利幅（成功報酬込）","=会員・預かり資産!D43",PCT),
       ("A",10,"100万会員 実質メリット率","=会員メリット!B17",PCT),
       ("C",10,"フロー（販売粗利）依存度","=損益計算書!E41",PCT1),
       ("E",10,"3年累計 純キャッシュフロー","=キャッシュフロー!E11",MAN0),
       ("G",10,"損益分岐マイル原価 vs 実績",'=TEXT(損益計算書!D48,"0.000")&" / "&TEXT(損益計算書!D49,"0.000")',"@")]
for col,row,lab,fml,fmt in kpi:
    put(D,f"{col}{row}",lab,font=Font(size=9,bold=True,color="666666"))
    put(D,f"{col}{row+1}",fml,font=Font(size=16,bold=True,color=BURG),fmt=fmt,fill=KPIF)

sec(D,14,"■ 年度別サマリ")
hdrow(D,15,["指標","1年目","2年目","3年目","3年累計"])
dash = [(16,"売上高","=損益計算書!{c}9"),(17,"営業利益","=損益計算書!{c}24"),
        (18,"経常利益","=損益計算書!{c}28"),(19,"経済利益（成功報酬発生込）","=損益計算書!{c}33"),
        (20,"ストック損益（成功報酬前）","=会員・預かり資産!{c}40"),
        (21,"ストック経済損益（成功報酬込）","=会員・預かり資産!{c}41"),
        (22,"マイル費用（P/L）","=マイル計算!{c}13"),
        (23,"実利用マイル原価（現金）","=マイル計算!{c}14"),
        (24,"純キャッシュフロー","=キャッシュフロー!{c}11"),
        (25,"期末預かり資産","=会員・預かり資産!{c}28")]
for r,l,f in dash:
    put(D,f"A{r}",l,border=True)
    for c in YC: put(D,f"{c}{r}",f.format(c=c),fmt=MAN0,border=True)
    put(D,f"E{r}",f"=D{r}" if r==25 else f"=SUM(B{r}:D{r})",fmt=MAN0,border=True,font=B)

sec(D,27,"■ 中谷氏コメントへの回答")
notes = [
 '="Q1「レストランは1.0でなく0.8でもよいのでは」→ 感度分析Aの通り、0.8にすると当社利幅は改善するが、"'
 '&"100万会員の実質メリットは "&TEXT(感度分析!F6,"#,##0")&"円 → "&TEXT(感度分析!F8,"#,##0")&"円へ縮小する。"',
 '="   代わりに感度分析Bの通り、レートは1.0のまま原価率を90%→55%（平日オフピーク枠に限定）にすれば、"'
 '&"利幅は "&TEXT(感度分析!C6,"0.00%")&" → "&TEXT(感度分析!C15,"0.00%")&" と、0.8にする以上の改善を会員メリットを一切削らずに得られる。"',
 "Q2「マイル消化98%＋新規流入を織り込んでも回るか」→ 全費目が預かり資産に対して線形のため、ストック利幅は規模によらず一定。上表の通り3年とも同水準。",
 '="Q3「PL上で回収できるか」→ フロー（ワイン販売粗利）依存度は "&TEXT(損益計算書!E41,"0.0%")&"。"'
 '&"100%を下回るほどストック単体で自立していることを意味する。"',
 "前提の更新点：保管料1.5%→1.05%（鈴与実額）／還元率3・4・5%→3.5・4.5・5.5%／成功報酬30%→25%／オークション買い手手数料を収益計上／ワインスクール・交流イベント（原価率50%）を交換先に追加。",
 "本ファイルは事業採算確認用の管理PLです。法定会計ではマイルを履行義務として収益繰延する可能性があるため、税理士・会計士の確認が必要です。"]
for i,t in enumerate(notes):
    put(D,f"A{28+i}",t,font=Font(size=9))
widths(D,{"A":34,"B":16,"C":22,"D":16,"E":22,"F":16,"G":24,"H":16})

del wb["Sheet"]
wb._sheets = [wb["ダッシュボード"],wb["前提"],wb["会員・預かり資産"],wb["マイル計算"],
              wb["損益計算書"],wb["キャッシュフロー"],wb["感度分析"],wb["会員メリット"]]
wb.save(OUT)
print("saved", OUT, "| 前提 mix rows", R0, "-", R1, "| EC row", EC, "| EV row", EV, "| 固定費 row", FROW, "| 在庫投資 row", IVR)
