# -*- coding: utf-8 -*-
"""WineBank 資本政策シミュレーション（戦略パートナー受入れ）
出典: 202602 資本政策案（2026年1月末迄 割当後・潜在株式考慮後）"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

F='メイリオ'
BK=Font(name=F); BL=Font(name=F,color='8A6A24',bold=True); GR=Font(name=F,color='595959')
SB=Font(name=F,bold=True); HD=Font(name=F,bold=True,color='FFFFFF')
SM=Font(name=F,size=9,color='8C8C8C'); RD=Font(name=F,size=9,color='333333')
HF=PatternFill('solid',fgColor='1A1A1A'); SF=PatternFill('solid',fgColor='F2F1EE')
YL=PatternFill('solid',fgColor='FAF3E2'); TF=PatternFill('solid',fgColor='F7F6F3')
OR_=PatternFill('solid',fgColor='F4EDDC'); GRP=PatternFill('solid',fgColor='EFEEEA')
TB=Border(top=Side(style='thin',color='1A1A1A'))
NUM='#,##0;(#,##0);-'; YEN='¥#,##0;(¥#,##0);-'; PCT='0.0%'
wb=Workbook(); wb.remove(wb.active)
def bar(ws,r,t,span=8):
    ws.cell(r,2,t).font=HD
    for c in range(2,2+span): ws.cell(r,c).fill=HF
    ws.cell(r,2).font=HD

# ============ ①現状の資本構成 ============
s=wb.create_sheet('①現状の資本構成')
for col,w in zip('ABCDEFGH',[3,40,13,11,13,11,46,2]): s.column_dimensions[col].width=w
s['B1']='WineBank 資本構成（2026年1月末 割当後）'; s['B1'].font=Font(name=F,bold=True,size=15)
s['B2']='出典：202602 資本政策案。FD＝潜在株式（ストックオプション）考慮後。単位：株'; s['B2'].font=SM
bar(s,4,'【1】株主構成',6)
for j,t in enumerate(['株主','発行済 株数','比率','FD 株数','FD 比率','区分']):
    c=s.cell(5,2+j,t); c.font=HD; c.fill=HF; c.alignment=Alignment(horizontal='center')
rows=[('中野邦人',3413,3413,'中野ブロック'),
      ('中野邦人 SO相当分',0,170,'中野ブロック'),
      ('Crown Jewel box LLC SO相当分',0,287,'中野ブロック'),
      ('株式会社スペースバンク',45,91,'中野ブロック'),
      ('マネーフォワードベンチャーズ（Hirac Fund 2号）',749,749,'外部'),
      ('株式会社Private BANK',375,375,'外部'),
      ('戸高薫 SO相当分',0,200,'外部'),
      ('小田垣栄司',200,200,'外部'),
      ('株式会社ベクトル',157,157,'外部'),
      ('個人投資家様',90,90,'外部'),
      ('寺田倉庫株式会社',52,52,'外部'),
      ('執行役員・従業員・その他関係者等',0,6,'外部')]
R0=6
for i,(n,iss,fdv,kind) in enumerate(rows):
    r=R0+i
    s.cell(r,2,n).font=BK
    c=s.cell(r,3,iss); c.number_format=NUM; c.font=BK
    c=s.cell(r,4,f'=C{r}/$C${R0+len(rows)}'); c.number_format=PCT; c.font=BK
    c=s.cell(r,5,fdv); c.number_format=NUM; c.font=BK
    c=s.cell(r,6,f'=E{r}/$E${R0+len(rows)}'); c.number_format=PCT; c.font=BK
    s.cell(r,7,kind).font=SM
    if kind=='中野ブロック':
        for cc in range(2,8): s.cell(r,cc).fill=TF
R1=R0+len(rows)-1; RT=R1+1
s.cell(RT,2,'合計').font=SB
for col,rng in (('C',f'C{R0}:C{R1}'),('E',f'E{R0}:E{R1}')):
    c=s.cell(RT,3 if col=='C' else 5,f'=SUM({rng})'); c.number_format=NUM; c.font=SB; c.fill=TF; c.border=TB
for cc in (4,6):
    c=s.cell(RT,cc,1.0); c.number_format=PCT; c.font=SB; c.fill=TF; c.border=TB
for cc in (2,7): s.cell(RT,cc).fill=TF; s.cell(RT,cc).border=TB
BL_=RT+2
s.cell(BL_,2,'中野ブロック 小計').font=SB; s.cell(BL_,2).fill=OR_
c=s.cell(BL_,3,f'=SUM(C{R0}:C{R0+3})'); c.number_format=NUM; c.font=SB; c.fill=OR_
c=s.cell(BL_,4,f'=C{BL_}/C{RT}'); c.number_format=PCT; c.font=SB; c.fill=OR_
c=s.cell(BL_,5,f'=SUM(E{R0}:E{R0+3})'); c.number_format=NUM; c.font=SB; c.fill=OR_
c=s.cell(BL_,6,f'=E{BL_}/E{RT}'); c.number_format=PCT; c.font=SB; c.fill=OR_
s.cell(BL_,7,'中野邦人＋100%保有法人（Crown Jewel box・スペースバンク）').font=RD
s.cell(BL_+1,2,'外部株主 小計').font=SB
c=s.cell(BL_+1,3,f'=C{RT}-C{BL_}'); c.number_format=NUM; c.font=SB; c.fill=TF
c=s.cell(BL_+1,4,f'=C{BL_+1}/C{RT}'); c.number_format=PCT; c.font=SB; c.fill=TF
c=s.cell(BL_+1,5,f'=E{RT}-E{BL_}'); c.number_format=NUM; c.font=SB; c.fill=TF
c=s.cell(BL_+1,6,f'=E{BL_+1}/E{RT}'); c.number_format=PCT; c.font=SB; c.fill=TF
CUR_T, CUR_B = RT, BL_
n=BL_+3
bar(s,n,'【2】直近ラウンドの条件',6)
pr=[('直近発行価額（2026年1月ラウンド）',450000,YEN,'第三者割当 357株・調達160,650千円'),
    ('ポストマネー評価額',None,YEN,'FD株数 × 発行価額'),
    ('中野ブロックの持分価値',None,YEN,'FD株数 × 発行価額')]
for i,(t,v,fmt,nt) in enumerate(pr):
    r=n+1+i
    s.cell(r,2,t).font=BK
    c=s.cell(r,3,v); c.number_format=fmt
    if v is not None: c.font=BL; c.fill=YL
    else: c.font=SB; c.fill=TF
    s.cell(r,7,nt).font=SM
PRICE=n+1
s.cell(n+2,3,f'=E{CUR_T}*C{PRICE}').number_format=YEN
s.cell(n+3,3,f'=E{CUR_B}*C{PRICE}').number_format=YEN
s.cell(n+5,2,'※Crown Jewel box は資本政策表上「Clown Jewel box LLC」と表記。上場審査資料では登記名との一致が必要です。').font=RD

# ============ ②希薄化余地 ============
d=wb.create_sheet('②希薄化余地')
for col,w in zip('ABCDEFGH',[3,42,15,15,15,15,44,2]): d.column_dimensions[col].width=w
d['B1']='「中野ブロック50%」をどこで測るか'; d['B1'].font=Font(name=F,bold=True,size=15)
d['B2']='戦略ラウンドの直後で50%にすると、上場時の公募増資でさらに希薄化し50%を割ります。'; d['B2'].font=RD
bar(d,4,'【1】前提',6)
pa=[('現在のFD株数',None,NUM,"①現状の資本構成から参照"),
    ('中野ブロック FD株数',None,NUM,'①現状の資本構成から参照'),
    ('中野ブロック 現在比率',None,PCT,''),
    ('流通株式適格（既存外部・10%未満かつ非役員）',1423,NUM,'MFV749＋Private BANK375＋ベクトル157＋寺田52＋個人90'),
    ('上場時 流通株式比率の要件',0.25,PCT,'東証グロース。10%以上の株主・役員等の保有分は流通株式から除外'),
    ('直近発行価額',None,YEN,'①現状の資本構成から参照')]
for i,(t,v,fmt,nt) in enumerate(pa):
    r=5+i
    d.cell(r,2,t).font=BK
    c=d.cell(r,3,v); c.number_format=fmt
    if v is not None: c.font=BL; c.fill=YL
    else: c.font=GR; c.fill=TF
    d.cell(r,7,nt).font=SM
P_T,P_B,P_R,P_LQ,P_FL,P_PR=5,6,7,8,9,10
d.cell(P_T,3,f"='①現状の資本構成'!E{CUR_T}").number_format=NUM
d.cell(P_B,3,f"='①現状の資本構成'!E{CUR_B}").number_format=NUM
d.cell(P_R,3,f'=C{P_B}/C{P_T}').number_format=PCT
d.cell(P_R,3).fill=OR_; d.cell(P_R,3).font=SB
d.cell(P_PR,3,f"='①現状の資本構成'!C{PRICE}").number_format=YEN
n=P_PR+2
bar(d,n,'【2】戦略ラウンドの規模別シミュレーション　※上場時に流通25%を満たす最小公募を置いた場合',6)
for j,t in enumerate(['戦略ラウンド 発行株数','ラウンド直後の\n新株主比率','ラウンド直後の\n中野ブロック','上場時に必要な\n公募株数','上場時の\n中野ブロック']):
    c=d.cell(n+1,2+j,t); c.font=HD; c.fill=HF
    c.alignment=Alignment(horizontal='center',wrap_text=True,vertical='center')
d.row_dimensions[n+1].height=40
S0=n+2
sizes=[0,600,900,1200,1500,1600,1800,2132]
for i,x in enumerate(sizes):
    r=S0+i
    c=d.cell(r,2,x); c.number_format=NUM; c.font=BL; c.fill=YL
    c=d.cell(r,3,f'=B{r}/($C${P_T}+B{r})'); c.number_format=PCT; c.font=BK
    c=d.cell(r,4,f'=$C${P_B}/($C${P_T}+B{r})'); c.number_format=PCT; c.font=BK
    # 流通25%を満たす最小公募 N: (LQ+N)/(T+X+N)>=0.25  →  N >= (0.25*(T+X)-LQ)/0.75
    c=d.cell(r,5,f'=MAX(0,ROUNDUP((($C${P_FL}*($C${P_T}+B{r}))-$C${P_LQ})/(1-$C${P_FL}),0))')
    c.number_format=NUM; c.font=BK
    c=d.cell(r,6,f'=$C${P_B}/($C${P_T}+B{r}+E{r})'); c.number_format=PCT; c.font=SB
    if x==1500:
        for cc in range(2,7): d.cell(r,cc).fill=OR_
        d.cell(r,7,'★上場時に中野ブロック50%超を維持できる上限').font=RD
S1=S0+len(sizes)-1
d.cell(S1+2,2,'※黄色セル（発行株数）を変えると全列が再計算されます。').font=SM
d.cell(S1+3,2,'※上場時の中野ブロックは、戦略ラウンドと公募増資の二段階の希薄化を反映した値です。').font=RD
d.cell(S1+4,2,'※戦略株主が各10%超で保有する前提。10%未満に抑えれば流通株式に算入され、必要公募株数は減ります。').font=SM

# ============ ③トランシェ比較 ============
t=wb.create_sheet('③トランシェ比較')
for col,w in zip('ABCDEFGH',[3,34,16,16,16,16,40,2]): t.column_dimensions[col].width=w
t['B1']='同じ希薄化で、いくら調達できるか'; t['B1'].font=Font(name=F,bold=True,size=15)
t['B2']='戦略ラウンドを2回に分けると、着地の持分を変えずに調達額を約1.4倍にできます。'; t['B2'].font=RD
bar(t,4,'【1】比較',6)
for j,lab in enumerate(['','案A：一括','案B：2トランシェ','差 B−A']):
    c=t.cell(5,2+j,lab); c.font=HD; c.fill=HF; c.alignment=Alignment(horizontal='center')
tr=[('第1トランシェ 株数',1500,900,NUM),
    ('第1トランシェ 発行価額',450000,450000,YEN),
    ('第1トランシェ 調達額',None,None,YEN),
    ('第2トランシェ 株数',0,600,NUM),
    ('第2トランシェ 発行価額',0,900000,YEN),
    ('第2トランシェ 調達額',None,None,YEN),
    ('調達額 合計',None,None,YEN),
    ('戦略株主 合計株数',None,None,NUM),
    ('上場前 総株数（FD）',None,None,NUM),
    ('上場時に必要な公募株数',None,None,NUM),
    ('上場時 総株数',None,None,NUM),
    ('上場時 中野ブロック比率',None,None,PCT)]
T0=6
for i,(lab,a,b,fmt) in enumerate(tr):
    r=T0+i
    t.cell(r,2,lab).font=SB if '合計' in lab or '上場時 中野' in lab else BK
    for col,v in ((3,a),(4,b)):
        c=t.cell(r,col,v); c.number_format=fmt
        if v is not None: c.font=BL; c.fill=YL
        else: c.font=SB; c.fill=TF
V=lambda r,c: f'{get_column_letter(c)}{r}'
for col in (3,4):
    L=get_column_letter(col)
    t.cell(T0+2,col,f'={L}{T0}*{L}{T0+1}').number_format=YEN
    t.cell(T0+5,col,f'={L}{T0+3}*{L}{T0+4}').number_format=YEN
    t.cell(T0+6,col,f'={L}{T0+2}+{L}{T0+5}').number_format=YEN
    t.cell(T0+7,col,f'={L}{T0}+{L}{T0+3}').number_format=NUM
    t.cell(T0+8,col,f"='②希薄化余地'!$C${P_T}+{L}{T0+7}").number_format=NUM
    t.cell(T0+9,col,f"=MAX(0,ROUNDUP((('②希薄化余地'!$C${P_FL}*{L}{T0+8})-'②希薄化余地'!$C${P_LQ})/(1-'②希薄化余地'!$C${P_FL}),0))").number_format=NUM
    t.cell(T0+10,col,f'={L}{T0+8}+{L}{T0+9}').number_format=NUM
    t.cell(T0+11,col,f"='②希薄化余地'!$C${P_B}/{L}{T0+10}").number_format=PCT
    t.cell(T0+11,col).fill=OR_
    t.cell(T0+6,col).fill=OR_
for r,fmt in ((T0+6,YEN),(T0+11,PCT)):
    c=t.cell(r,5,f'=D{r}-C{r}'); c.number_format=fmt; c.font=SB; c.fill=TF
t.cell(T0+13,2,'※案Bの第2トランシェ価額90万円は、プランD（経常＋120.4百万円）を実現した後のステップアップを想定した仮置きです。').font=RD
t.cell(T0+14,2,'※FY2026の着地（経常▲124.8百万円）で全量を price すると、最も不利な時点で希薄化を確定させることになります。').font=RD

# ============ ④2社への割付 ============
a=wb.create_sheet('④2社への割付')
for col,w in zip('ABCDEFGH',[3,34,14,14,14,14,44,2]): a.column_dimensions[col].width=w
a['B1']='戦略パートナー2社への割付案（案B・合計1,500株）'; a['B1'].font=Font(name=F,bold=True,size=15)
a['B2']='各社を15%未満に抑え、持分法適用（相手側の損益取込み）を回避します。'; a['B2'].font=SM
bar(a,4,'【1】割付',6)
for j,lab in enumerate(['','第1トランシェ','第2トランシェ','合計 株数','上場前 比率','上場時 比率']):
    c=a.cell(5,2+j,lab); c.font=HD; c.fill=HF; c.alignment=Alignment(horizontal='center')
al=[('金融・資本市場パートナー',450,300),('富裕層・二次流通パートナー',450,300)]
A0=6
for i,(nm,t1,t2) in enumerate(al):
    r=A0+i
    a.cell(r,2,nm).font=BK
    for col,v in ((3,t1),(4,t2)):
        c=a.cell(r,col,v); c.number_format=NUM; c.font=BL; c.fill=YL
    a.cell(r,5,f'=C{r}+D{r}').number_format=NUM
    a.cell(r,5).font=SB
    a.cell(r,6,f"=E{r}/'③トランシェ比較'!$D${T0+8}").number_format=PCT
    a.cell(r,7,f"=E{r}/'③トランシェ比較'!$D${T0+10}").number_format=PCT
A1=A0+1; AT=A1+1
a.cell(AT,2,'合計').font=SB
for col in (3,4,5):
    c=a.cell(AT,col,f'=SUM({get_column_letter(col)}{A0}:{get_column_letter(col)}{A1})')
    c.number_format=NUM; c.font=SB; c.fill=TF; c.border=TB
for col in (6,7):
    c=a.cell(AT,col,f'=SUM({get_column_letter(col)}{A0}:{get_column_letter(col)}{A1})')
    c.number_format=PCT; c.font=SB; c.fill=TF; c.border=TB
a.cell(AT,2).fill=TF; a.cell(AT,2).border=TB
n=AT+2
bar(a,n,'【2】守るべき閾値',6)
th=[('各社 15%未満','持分法適用（相手側が当社損益を取り込む）を回避。20%以上は原則適用、15〜20%も役員派遣等で適用され得る'),
    ('2社合計 33.3%以下','特別決議の拒否権を渡さない'),
    ('上場時 中野ブロック 50%超','戦略ラウンドと公募増資の二段階の希薄化を反映した後の水準で測る'),
    ('流通株式比率 25%以上','東証グロース。戦略株主にはロックアップと一部売出しを出資契約の時点で合意しておく')]
for i,(k,v) in enumerate(th):
    r=n+1+i
    a.cell(r,2,k).font=SB; a.cell(r,2).fill=TF
    a.cell(r,3,v).font=BK
    a.merge_cells(start_row=r,start_column=3,end_row=r,end_column=7)
    a.cell(r,3).alignment=Alignment(wrap_text=True,vertical='top')
    a.row_dimensions[r].height=30

# ============ ⑤論点 ============
z=wb.create_sheet('⑤事前に潰す論点')
for col,w in zip('ABCH',[3,26,96,2]): z.column_dimensions[col].width=w
z['B1']='パートナー打診の前に片づけておく論点'; z['B1'].font=Font(name=F,bold=True,size=15)
z['B2']='いずれもデューデリジェンスで必ず出ます。後から出ると交渉力を失います。'; z['B2'].font=RD
bar(z,4,'【1】論点',2)
iss=[('関連当事者　持分',
      'Crown Jewel box LLC SO 287株（FD 5.0%）とスペースバンク 91株（同1.6%）の計378株・6.5%が中野氏100%保有法人。'
      '法人へのストックオプション付与は通常の設計ではなく、主幹事・監査法人から必ず論点化されます。'),
     ('関連当事者　取引',
      'FY2027計画のコンサルティング料92.5百万円は中野出資グループからの収入。'
      '銀行向けに用意した「支払能力と対価の算定根拠」の文書化を先行させ、DDの一次資料として提出できる状態にしておく。'),
     ('選択肢：取り込み',
      'スペースバンク・Crown Jewel box に実体事業があるなら、株式交換または現物出資でWineBankに取り込む案がある。'
      '①関連当事者論点が消える ②中野氏の持分が増えて戦略ラウンドの希薄化を相殺できる ③事業ストーリーが1本化される。'
      '両社の事業内容・売上・純資産の確認が先。'),
     ('プロラタ権',
      'マネーフォワードベンチャーズ（FD 12.9%・最大の外部株主）の引受権・希薄化防止条項の有無を投資契約で確認。'
      'パートナー探索を勧めているのは同社なので、事前に方針をすり合わせておくのが早い。'),
     ('種類株の設計',
      '2026年1月ラウンドは種類株での調達実績あり。議決権のない種類株にすれば、中野持分を薄めずに資金を入れられ、'
      '相手側も持分法適用を避けられる。取締役会オブザーバー席＋事前承諾事項で実質的な関与を担保する。'),
     ('プライシングの時期',
      'FY2026着地（経常▲124.8百万円）でバリュエーション交渉に入るのは最も不利。'
      'デューデリジェンスは今すぐ着手し、価格決定はFY2027 第1〜第2四半期の実績が出た後に置く。'),
     ('ワインファンドの前歴',
      '2016年のVIN-NET破綻（約36億円が償還不能）で日本のワインファンド市場は止まった経緯があり、'
      'どの相手からも必ず問われます。現物保管・所有権の明確化・第三者評価の3点で答えを用意する。'
      '規制業者である証券会社を株主に入れること自体が、この疑念への回答になります。')]
for i,(k,v) in enumerate(iss):
    r=6+i*2
    z.cell(r,2,k).font=SB; z.cell(r,2).fill=OR_
    z.cell(r,3,v).font=BK
    z.cell(r,3).alignment=Alignment(wrap_text=True,vertical='top')
    z.row_dimensions[r].height=46
# ============ ⑥既存株主入替えシナリオ ============
x=wb.create_sheet('⑥既存株主入替え')
for col,w in zip('ABCDEFGH',[3,34,15,15,15,15,40,2]): x.column_dimensions[col].width=w
x['B1']='既存株主を入れ替えて強力なパートナーを迎える場合'; x['B1'].font=Font(name=F,bold=True,size=15)
x['B2']='経営陣以外の株主を買い取り、パートナーが譲受＋増資で持分を取る前提。単位：株／円'; x['B2'].font=SM
bar(x,4,'【1】株主の色分け',6)
for j,t in enumerate(['','FD 株数','FD 比率','','','']):
    c=x.cell(5,2+j,t); c.font=HD; c.fill=HF; c.alignment=Alignment(horizontal='center')
grp=[('経営陣ブロック（残留）',4367,'中野ブロック3,961＋小田垣200＋戸高SO200＋従業員6'),
     ('　うち 中野ブロック',3961,'中野個人3,413＋SO170＋Crown Jewel box 287＋スペースバンク91'),
     ('売却対象（外部株主）',1423,'MFV749＋Private BANK375＋ベクトル157＋個人90＋寺田52')]
G0=6
for i,(n,v,nt) in enumerate(grp):
    r=G0+i
    x.cell(r,2,n).font=SB if i!=1 else BK
    c=x.cell(r,3,v); c.number_format=NUM; c.font=SB if i!=1 else BK
    c=x.cell(r,4,f"=C{r}/'②希薄化余地'!$C${P_T}"); c.number_format=PCT; c.font=SB if i!=1 else BK
    x.cell(r,7,nt).font=SM
    if i==2:
        for cc in range(2,5): x.cell(r,cc).fill=OR_
G_EXT=G0+2
PX=G0+3
x.cell(PX,2,'買取・引受 単価（前提）').font=BK
c=x.cell(PX,3,450000); c.number_format=YEN; c.font=BL; c.fill=YL
x.cell(PX,7,'2026年1月ラウンドと同額を仮置き。ここを変えると【2】【3】が再計算されます。').font=SM
n=G0+5
bar(x,n,'【2】売却対象株主の取得原価　※買取交渉の難易度',6)
for j,t in enumerate(['株主','株数','取得原価 合計','取得単価','@45万での倍率','受取額 @45万']):
    c=x.cell(n+1,2+j,t); c.font=HD; c.fill=HF; c.alignment=Alignment(horizontal='center',wrap_text=True)
sh=[('マネーフォワードベンチャーズ',749,527*285000+222*450000,'527株@28.5万＋222株@45万'),
    ('株式会社Private BANK',375,375*50000,'2021/12ラウンド @5万'),
    ('株式会社ベクトル',157,157*192000,'2022/10ラウンド @19.2万'),
    ('個人投資家様',90,90*450000,'2026/01ラウンド @45万。直近参加のため簿価売却'),
    ('寺田倉庫株式会社',52,52*192000,'2022/10ラウンド @19.2万')]
H0=n+2
for i,(nm,cnt,cst,nt) in enumerate(sh):
    r=H0+i
    x.cell(r,2,nm).font=BK
    c=x.cell(r,3,cnt); c.number_format=NUM; c.font=BK
    c=x.cell(r,4,cst); c.number_format=YEN; c.font=BK
    c=x.cell(r,5,f'=D{r}/C{r}'); c.number_format=YEN; c.font=BK
    c=x.cell(r,6,f"=C{r}*$C${PX}/D{r}"); c.number_format='0.00"x"'; c.font=SB
    c=x.cell(r,7,f"=C{r}*$C${PX}"); c.number_format=YEN; c.font=BK
H1=H0+len(sh)-1; HT=H1+1
x.cell(HT,2,'合計').font=SB
for col in (3,4,7):
    c=x.cell(HT,col,f'=SUM({get_column_letter(col)}{H0}:{get_column_letter(col)}{H1})')
    c.number_format=NUM if col==3 else YEN; c.font=SB; c.fill=TF; c.border=TB
for cc in (2,5,6): x.cell(HT,cc).fill=TF; x.cell(HT,cc).border=TB
x.cell(HT+2,2,'※取得原価の大半が低い株主（Private BANK 9.0倍・ベクトル/寺田 2.3倍）は売却に応じやすい。').font=SM
x.cell(HT+3,2,'※実質的な交渉相手はマネーフォワードベンチャーズ1社。@45万では1.35倍にとどまり、価格を押してくる可能性が高い。').font=RD
x.cell(HT+4,2,'※個人投資家様90株は2026年1月に@45万で参加したばかりで簿価売却となるため、別途の配慮が要る。').font=RD
n=HT+6
bar(x,n,'【3】パートナー持分別の投下額　※外部1,423株の譲受＋増資',6)
for j,t in enumerate(['パートナー 目標持分','必要な増資 株数','譲受＋増資の\n投下総額','上場前 総株数','中野ブロック\n比率','']):
    c=x.cell(n+1,2+j,t); c.font=HD; c.fill=HF; c.alignment=Alignment(horizontal='center',wrap_text=True,vertical='center')
x.row_dimensions[n+1].height=34
K0=n+2
tg=[(0.334,'拒否権ライン。持分法適用で相手のPLに乗る'),
    (0.40,'筆頭株主は中野氏のまま。実質的な共同経営'),
    (0.501,'連結子会社化。上場は親子上場となる')]
for i,(t2,nt) in enumerate(tg):
    r=K0+i
    c=x.cell(r,2,t2); c.number_format=PCT; c.font=BL; c.fill=YL
    c=x.cell(r,3,f"=($B{r}*'②希薄化余地'!$C${P_T}-$C${G_EXT})/(1-$B{r})"); c.number_format=NUM; c.font=BK
    c=x.cell(r,4,f"=($C${G_EXT}+C{r})*$C${PX}"); c.number_format=YEN; c.font=SB
    c=x.cell(r,5,f"='②希薄化余地'!$C${P_T}+C{r}"); c.number_format=NUM; c.font=BK
    c=x.cell(r,6,f"='②希薄化余地'!$C${P_B}/E{r}"); c.number_format=PCT; c.font=SB
    x.cell(r,7,nt).font=SM
    if i==2:
        for cc in range(2,7): x.cell(r,cc).fill=OR_
K1=K0+2
x.cell(K1+2,2,'※投下総額は@45万換算。支配権プレミアムを乗せれば増え、FY2026の着地を理由に値引きされれば減ります。').font=SM
n=K1+4
bar(x,n,'【4】過半数を渡す場合に必ず起きること',6)
w=[('親子上場','上場会社が過半数を持つと、WineBankの上場は親子上場になる。東証は支配株主を有する上場会社への'
    'ガバナンス要求を強めており、上場審査では親会社からの独立性が厳しく問われる。'
    '実務的には、出口はIPOではなく将来の完全子会社化（M&A）に寄る。'),
   ('赤字の連結','過半数を取った相手はWineBankの損益を全部取り込む。FY2026の経常▲124.8百万円を連結する判断は'
    '上場会社にとって重い。プランDの黒字化を確認してからのほうが、相手は動きやすい。'),
   ('IPOを残す道','パートナーを33.4〜49%にとどめれば親子上場にはならず、持分法で相手のPLには乗る。'
    '「本気度」は比率ではなく、独占提携・役員派遣・相手の営業目標への組み込み・未達時のラチェットで担保する。'),
   ('PEという選択肢','既存株主を全部買い取り、経営改革し、数年でIPOに持っていくのはプライベートエクイティの標準的な仕事。'
    'PEが40%前後、事業会社が10〜15%で並走する二階建てなら、株主整理・経営改革・IPOの3つを同時に満たせる。')]
for i,(k,v) in enumerate(w):
    r=n+1+i
    x.cell(r,2,k).font=SB; x.cell(r,2).fill=OR_
    x.cell(r,3,v).font=BK
    x.merge_cells(start_row=r,start_column=3,end_row=r,end_column=7)
    x.cell(r,3).alignment=Alignment(wrap_text=True,vertical='top')
    x.row_dimensions[r].height=46

for ws in wb: ws.sheet_view.showGridLines=False
wb.save('WineBank_資本政策シミュレーション.xlsx'); print('saved')
