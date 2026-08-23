# -*- coding: utf-8 -*-
"""FY2027(2027年9月期)利益予測モデル生成スクリプト
出典: 事業計画202608(銀行様) 全社シート月次 / 決算報告書 第54期"""
import json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

F='メイリオ'
# WineBank ハウススタイル：無彩色モノトーン ＋ メイリオ
BK=Font(name=F); BL=Font(name=F,color='8A6A24',bold=True); GR=Font(name=F,color='595959')
SB=Font(name=F,bold=True); HD=Font(name=F,bold=True,color='FFFFFF')
SM=Font(name=F,size=9,color='8C8C8C'); RD=Font(name=F,size=9,color='333333')
GY=Font(name=F,color='8C8C8C',italic=True)
HF=PatternFill('solid',fgColor='1A1A1A'); SF=PatternFill('solid',fgColor='F2F1EE')
YL=PatternFill('solid',fgColor='FAF3E2'); TF=PatternFill('solid',fgColor='F7F6F3')
OP=PatternFill('solid',fgColor='F7F6F3'); OR_=PatternFill('solid',fgColor='F4EDDC')
GRP=PatternFill('solid',fgColor='EFEEEA')
CREAMFILL=PatternFill('solid',fgColor='FBFAF7')
TB=Border(top=Side(style='thin',color='1A1A1A'))
YEN='¥#,##0;(¥#,##0);-'; PCT='0.0%'
SGA=json.load(open('sga_monthly.json'))
FM=['2025/10','2025/11','2025/12','2026/01','2026/02','2026/03','2026/04','2026/05','2026/06','2026/07','2026/08','2026/09']
SALES=[32789307,96919229,43039844,32477958,41290432,33033585,31863679,33917025,31504488,60489854,37962048,232918371]
COGS =[18502877,69212121,21085289,16521949,21070367,18454664,16184238,15777832,15172640,38396406,23721353,204966678]
GP   =[14286430,27707108,21954555,15956009,20220065,14578921,15679441,18139193,16331848,22093448,14240695,27951692]
ACC=list(SGA.keys())
wb=Workbook(); wb.remove(wb.active)

def bar(ws,r,t,span=6):
    ws.cell(r,2,t).font=HD
    for c in range(2,2+span): ws.cell(r,c).fill=HF
    ws.cell(r,2).font=HD

# ============ ④FY2026月次実績 ============
h=wb.create_sheet('④FY2026月次実績')
h.column_dimensions['A'].width=3; h.column_dimensions['B'].width=26
for i in range(12): h.column_dimensions[get_column_letter(3+i)].width=13
h.column_dimensions['O'].width=15; h.column_dimensions['P'].width=15
h.column_dimensions['Q'].width=15; h.column_dimensions['R'].width=14
h['B1']='FY2026（2025年10月-2026年9月）月次　実績／見込'; h['B1'].font=Font(name=F,bold=True,size=13)
h['B2']='青字＝実績（2026/06まで）　灰字斜体＝見込（2026/07以降・同一値が並ぶプラグ）　出典：事業計画202608（銀行様）全社シート'; h['B2'].font=SM
h['B3']='不採算3店舗の撤退により、2026/04以降が再編後の実力値。WineBankテラスはWineBank CLUBのフラッグシップ店として改善のうえ継続。'; h['B3'].font=RD
r=5
h.cell(r,2,'科目').font=SB; h.cell(r,2).fill=SF
for i,m in enumerate(FM):
    c=h.cell(r,3+i,m); c.font=SB; c.fill=SF; c.alignment=Alignment(horizontal='center')
for j,t in enumerate(['通期','①10-3月\n平均','②4-6月\n平均','差 ②-①']):
    c=h.cell(r,15+j,t); c.font=SB; c.fill=SF; c.alignment=Alignment(horizontal='center',wrap_text=True)
h.row_dimensions[r].height=30
def hrow(rr,label,vals,bold=False,fill=None,avg=True):
    c=h.cell(rr,2,label); c.font=SB if bold else BK
    if fill: c.fill=fill
    for i,v in enumerate(vals):
        cc=h.cell(rr,3+i,v); cc.number_format=YEN; cc.font=(SB if bold else (BL if i<9 else GY))
        if fill: cc.fill=fill
    for j,f in enumerate([f'=SUM(C{rr}:N{rr})',f'=AVERAGE(C{rr}:H{rr})',f'=AVERAGE(I{rr}:K{rr})',f'=Q{rr}-P{rr}']):
        cc=h.cell(rr,15+j,f if (avg or j==0) else None); cc.number_format=YEN; cc.font=SB; cc.fill=fill or TF
SR,CR,GR_=6,7,8
hrow(6,'売上',SALES); hrow(7,'売上原価',COGS); hrow(8,'売上総利益',GP,bold=True,fill=TF)
h.cell(10,2,'【販売費及び一般管理費 内訳】').font=SB; h.cell(10,2).fill=GRP
SG0=11
for k,acc in enumerate(ACC): hrow(SG0+k,acc,SGA[acc])
SG1=SG0+len(ACC)-1
SGT=SG1+1
h.cell(SGT,2,'販管費 合計').font=SB; h.cell(SGT,2).border=TB
for i in range(12):
    col=get_column_letter(3+i)
    c=h.cell(SGT,3+i,f'=SUM({col}{SG0}:{col}{SG1})'); c.number_format=YEN; c.font=SB; c.fill=TF; c.border=TB
for j,f in enumerate([f'=SUM(C{SGT}:N{SGT})',f'=AVERAGE(C{SGT}:H{SGT})',f'=AVERAGE(I{SGT}:K{SGT})',f'=Q{SGT}-P{SGT}']):
    c=h.cell(SGT,15+j,f); c.number_format=YEN; c.font=SB; c.fill=TF; c.border=TB
OPR=SGT+1
h.cell(OPR,2,'営業利益').font=SB; h.cell(OPR,2).fill=OP; h.cell(OPR,2).border=TB
for i in range(12):
    col=get_column_letter(3+i)
    c=h.cell(OPR,3+i,f'={col}{GR_}-{col}{SGT}'); c.number_format=YEN; c.font=SB; c.fill=OP; c.border=TB
c=h.cell(OPR,15,f'=SUM(C{OPR}:N{OPR})'); c.number_format=YEN; c.font=SB; c.fill=OP; c.border=TB
h.cell(OPR+2,2,'※「差 ②-①」がマイナスの科目＝不採算3店舗の撤退および固定費削減により減少した費用。合計で月▲8.58百万円（年▲103百万円）。').font=RD
h.freeze_panes='C6'

# ============ ②前提 ============
p=wb.create_sheet('②前提')
p.column_dimensions['A'].width=3; p.column_dimensions['B'].width=38
for c in 'CDE': p.column_dimensions[c].width=17
p.column_dimensions['F'].width=66
p['B1']='FY2027（2027年9月期）利益予測モデル　－　前提条件'; p['B1'].font=Font(name=F,bold=True,size=14)
p['B2']='青字＝入力セル／黄色＝主要前提／黒字＝計算式　単位：円（税別）　作成日 2026/08/21'; p['B2'].font=SM
bar(p,4,'【1】ベース：再編後の実力値（2026/04-06 実績3か月平均）')
p.cell(5,2,'項目').font=SB; p.cell(5,3,'月次平均').font=SB; p.cell(5,3).alignment=Alignment(horizontal='center')
for c in (2,3): p.cell(5,c).fill=SF
base=[('売上',f"=AVERAGE('④FY2026月次実績'!I{SR}:K{SR})"),
      ('売上総利益',f"=AVERAGE('④FY2026月次実績'!I{GR_}:K{GR_})"),
      ('販管費',f"=AVERAGE('④FY2026月次実績'!I{SGT}:K{SGT})")]
for i,(t,f_) in enumerate(base):
    p.cell(6+i,2,t).font=BK
    c=p.cell(6+i,3,f_); c.number_format=YEN; c.font=GR; c.fill=TF
p.cell(9,2,'売上総利益率').font=BK
c=p.cell(9,3,'=C7/C6'); c.number_format=PCT; c.font=BK
p.cell(6,6,'2026/03以前は撤退した不採算3店舗の売上・費用を含むため基準から除外（貴社ご指摘）。').font=RD
S_M,G_M,P_M,GM=6,7,8,9
bar(p,11,'【2】上乗せ案件（2026/09-2027/08）')
for j,t in enumerate(['区分・案件','金額','取引先','損益計上区分']):
    c=p.cell(12,2+j,t); c.font=SB; c.fill=SF
items=[('経営指導料　間接グループ5社',80000000,'中野出資','営業収益（12か月按分）'),
       ('経営指導料　Thierry Marx（2026年4月開業）',2500000,'Apiciusグループ','営業収益（12か月按分）'),
       ('経営指導料　Apicius',10000000,'中野出資','営業収益（12か月按分）'),
       ('M&A仲介　Apicius既存株主売却仲介',20000000,'外部','営業収益（一時）'),
       ('新規事業利益　新規クルーザー事業 1回目',10000000,'外部','営業収益（一時）'),
       ('新規事業利益　新規クルーザー事業 2回目',10000000,'外部','営業収益（一時）')]
I0=13
for i,(n,v,tori,k) in enumerate(items):
    p.cell(I0+i,2,n).font=BK
    c=p.cell(I0+i,3,v); c.font=BL; c.number_format=YEN; c.fill=YL
    p.cell(I0+i,4,tori).font=BK
    p.cell(I0+i,5,k).font=BK
I1=I0+5; TT=I1+1
MNA=I0+3
CR1,CR2=I0+4,I0+5
for lab,f_,rr in [('合計',f'=SUM(C{I0}:C{I1})',TT),('　うち 営業収益',f'=SUM(C{I0}:C{I1})',TT+1),('　うち 営業外収益',0,TT+2)]:
    p.cell(rr,2,lab).font=SB if rr==TT else BK
    c=p.cell(rr,3,f_); c.number_format=YEN; c.font=SB if rr==TT else BK
    if rr==TT: c.fill=TF
ADV,INV=TT+1,TT+2
p.cell(TT+2,6,'※新規クルーザー事業利益は投資損益ではなく事業利益のため、全額を営業収益に計上。営業外収益は0。').font=RD
p.cell(TT,6,'※期間表記は2026/09-2027/08。決算期(2026/10-2027/09)と1か月ズレるため全額FY2027帰属で試算。').font=RD
p.cell(TT+1,6,'※グループ内 92.5百万円（中野出資90.0＋Apiciusグループ2.5）／外部 40.0百万円（M&A仲介20.0＋新規クルーザー事業20.0）。').font=RD
O0=TT+4
bar(p,O0,'【3】その他前提')
oth=[('営業外費用（支払利息等・年額）',15000000,YEN,'事業計画FY2027計画値'),
     ('M&A仲介 計上月（1〜12）',6,'0','6＝2027年3月'),
     ('新規クルーザー事業 1回目 計上月',6,'0','6＝2027年3月'),
     ('新規クルーザー事業 2回目 計上月',12,'0','12＝2027年9月（期末までに計上）'),
     ('（未使用）',0,YEN,'—'),
     ('新規顧客売上 計画額（FY2027）',1066000000,YEN,'事業計画のワイン投資 新規顧客計画 合計'),
     ('シナリオ選択（1＝B／2＝C／3＝D）',3,'0','③月次推移・⑦資金繰り表に反映'),
     ('ワイン仕入 方針（1＝在庫横ばい／2＝FY2026と同額）',2,'0','⑥仕入・在庫計画に反映'),
     ('FY2026 年間仕入額',419066414,YEN,'売上原価479,066,414＋在庫増減▲60,000,000（480→420百万円）')]
for i,(t,v,fmt,nt) in enumerate(oth):
    p.cell(O0+1+i,2,t).font=BK
    c=p.cell(O0+1+i,3,v); c.font=BL; c.number_format=fmt; c.fill=YL
    p.cell(O0+1+i,6,nt).font=SM
NOE,MI,MC1,MC2,TGT,PLAN=O0+1,O0+2,O0+3,O0+4,O0+5,O0+6
SCN,BUYP,BUY26=O0+7,O0+8,O0+9
rr=O0+10
p.cell(rr,2,'法人税等').font=BK; p.cell(rr,3,'均等割のみ').font=BL
p.cell(rr,6,'繰越欠損金 約252百万円（FY2025▲127.4＋FY2026▲124.8）。資本金1,000万円の中小法人は所得の100%控除可。').font=SM
p.cell(rr,6).alignment=Alignment(wrap_text=True,vertical='top'); p.row_dimensions[rr].height=32

# 【6】消費税の算定（FY2026分＝FY2027の資金繰りに乗る納付額）
CT0=rr+2
bar(p,CT0,'【6】消費税の算定　※FY2027の資金繰りに乗るのはFY2026分の確定納付と当期の中間納付')
ct=[('消費税率',0.10,PCT,True,''),
    ('FY2026 課税売上高',708205820,YEN,True,'事業計画FY2026見込の売上高'),
    ('FY2026 ワイン仕入（課税仕入）',None,YEN,False,'売上原価479,066,414＋在庫増減▲60,000,000'),
    ('FY2026 販管費のうち課税仕入',208594315,YEN,True,'販管費343,932,836−人件費・減価償却・租税公課等135,338,521'),
    ('　仮受消費税',None,YEN,False,''),
    ('　仮払消費税',None,YEN,False,''),
    ('FY2026 確定消費税額',None,YEN,False,'仮受−仮払。赤字でも人件費・減価償却は仕入税額控除できないため納付が生じる'),
    ('　うちFY2026中に中間納付済',4500000,YEN,True,'FY2025確定額ベースの中間納付（年3回）の概算。実績値に差し替えてください'),
    ('2026/11 確定納付額',None,YEN,False,'FY2027の資金繰りに計上'),
    ('FY2027 中間納付（1回あたり）',None,YEN,False,'確定額÷4。2027/02・05・08の年3回'),]
for i,(t,v,fmt,inp,nt) in enumerate(ct):
    r2=CT0+1+i
    p.cell(r2,2,t).font=BK
    c=p.cell(r2,3,v if inp else None); c.number_format=fmt
    if inp: c.font=BL; c.fill=YL
    else: c.font=SB; c.fill=TF
    p.cell(r2,6,nt).font=SM
CT_RATE,CT_SALES,CT_BUY,CT_SGA=CT0+1,CT0+2,CT0+3,CT0+4
CT_UKE,CT_HARAI,CT_FIX,CT_MIDPAID,CT_FIN,CT_MID=CT0+5,CT0+6,CT0+7,CT0+8,CT0+9,CT0+10
p.cell(CT_BUY,3,f'=C{O0+9}').number_format=YEN
p.cell(CT_UKE,3,f'=C{CT_SALES}*C{CT_RATE}').number_format=YEN
p.cell(CT_HARAI,3,f'=(C{CT_BUY}+C{CT_SGA})*C{CT_RATE}').number_format=YEN
p.cell(CT_FIX,3,f'=C{CT_UKE}-C{CT_HARAI}').number_format=YEN
p.cell(CT_FIN,3,f'=C{CT_FIX}-C{CT_MIDPAID}').number_format=YEN
p.cell(CT_MID,3,f'=C{CT_FIX}/4').number_format=YEN
for r2 in (CT_BUY,CT_UKE,CT_HARAI,CT_FIX,CT_FIN,CT_MID):
    p.cell(r2,3).font=SB; p.cell(r2,3).fill=TF
p.cell(CT_FIX,3).fill=OR_
p.cell(CT_FIX+4,6,'※ワイン仕入が課税売上−課税販管費（約499.6百万円）を超えると還付に転じる。FY2026の仕入419.1百万円では納付側。').font=RD

# 【7】プランC・Dの前提（B＋差額売上×粗利率30%）
D0=CT_MID+3
bar(p,D0,'【7】プランC・Dの前提　Bに、前期・今期の売上平均までの差額売上を粗利率30%で加算')
p.cell(D0+1,2,'前期・今期の売上平均').font=BK
c=p.cell(D0+1,3,730535360); c.font=BL; c.number_format=YEN; c.fill=YL
p.cell(D0+1,6,'FY2025実績752,864,901とFY2026見込708,205,820の平均').font=SM
p.cell(D0+2,2,'上乗せ案件（経営指導料等）').font=BK
c=p.cell(D0+2,3,f'=C{ADV}'); c.font=SB; c.number_format=YEN; c.fill=TF
p.cell(D0+3,2,'差額売上の粗利率').font=BK
c=p.cell(D0+3,3,0.30); c.font=BL; c.number_format=PCT; c.fill=YL
p.cell(D0+3,6,'既存事業の実績（FY2025 26.8%・FY2026 32.4%）を踏まえ保守的に30%').font=SM
C_AVG,C_ADVR,C_GM=D0+1,D0+2,D0+3
blocks=[('【プランC】上乗せ案件を売上平均の内数とする',D0+5,f'=C{C_AVG}-C{C_ADVR}'),
        ('【プランD】上乗せ案件を売上平均に外数で加算する',D0+12,f'=C{C_AVG}')]
for title,base,exf in blocks:
    p.cell(base,2,title).font=SB; p.cell(base,2).fill=GRP
    for cc in range(3,7): p.cell(base,cc).fill=GRP
    labs=[('　既存事業 売上',exf),('　差額売上（Bプランからの上積み）',None),
          ('　差額売上の粗利',None),('　売上総利益 合計',None),
          ('　売上高（上乗せ含む）',None),('　総合粗利率',None)]
    for i,(t,f_) in enumerate(labs):
        r3=base+1+i
        p.cell(r3,2,t).font=BK
        c=p.cell(r3,3,f_); c.number_format=PCT if '粗利率' in t else YEN
        c.font=SB; c.fill=TF
    EX,DF,GP,GT,TOT,GMT=base+1,base+2,base+3,base+4,base+5,base+6
    p.cell(DF,3,f'=C{EX}-C{S_M}*12').number_format=YEN
    p.cell(GP,3,f'=C{DF}*C{C_GM}').number_format=YEN
    p.cell(GT,3,f'=C{G_M}*12+C{GP}+C{C_ADVR}').number_format=YEN
    p.cell(TOT,3,f'=C{EX}+C{C_ADVR}').number_format=YEN
    p.cell(GMT,3,f'=C{GT}/C{TOT}').number_format=PCT
    p.cell(GT,3).fill=OR_
C_EX,C_DIFF,C_DGP,C_GT,C_TOT,C_GMT=D0+6,D0+7,D0+8,D0+9,D0+10,D0+11
D_EX,D_DIFF,D_DGP,D_GT,D_TOT,D_GMT=D0+13,D0+14,D0+15,D0+16,D0+17,D0+18
p.cell(D_GMT+2,6,'※Cは上乗せ案件を売上平均の内数、Dは外数として扱う。Dの既存事業売上730.5百万円は前期・今期の平均そのもの。').font=RD
p.cell(D_GMT+3,6,'※差額売上はCが208.9百万円（新規顧客計画の19.6%）、Dが341.4百万円（同32.0%）。').font=RD
Q=lambda r: f"'②前提'!$C${r}"

# ============ ③FY2027月次推移 ============
m=wb.create_sheet('③FY2027月次推移')
m.column_dimensions['A'].width=3; m.column_dimensions['B'].width=30
m.column_dimensions['C'].width=15
for i in range(12): m.column_dimensions[get_column_letter(4+i)].width=13
m.column_dimensions['P'].width=16
MO=['2026/10','2026/11','2026/12','2027/01','2027/02','2027/03','2027/04','2027/05','2027/06','2027/07','2027/08','2027/09']
m['B1']='FY2027（2026年10月-2027年9月）月次推移表　②前提の「シナリオ選択」で B（上乗せのみ）／C（内数）／D（外数）を切替'
m['B1'].font=Font(name=F,bold=True,size=13)
m['B2']='C列＝月次ベース（2026/04-06実績平均、④シートから自動リンク）。D列以降は同額で横ばい。緑字＝他シートからのリンク'; m['B2'].font=SM
m.cell(4,2,'科目').font=SB; m.cell(4,2).fill=SF
c=m.cell(4,3,'月次ベース\n(26/4-6平均)'); c.font=SB; c.fill=SF; c.alignment=Alignment(horizontal='center',wrap_text=True)
for i,mo in enumerate(MO):
    c=m.cell(4,4+i,mo); c.font=SB; c.fill=SF; c.alignment=Alignment(horizontal='center')
c=m.cell(4,16,'通期合計'); c.font=SB; c.fill=SF; c.alignment=Alignment(horizontal='center')
m.row_dimensions[4].height=30
def mrow(rr,label,basef,monf=None,bold=False,fill=None,font=BK,top=False,indent=0):
    c=m.cell(rr,2,('　'*indent)+label); c.font=SB if bold else font
    if fill: c.fill=fill
    if top: c.border=TB
    if basef is not None:
        cb=m.cell(rr,3,basef); cb.number_format=YEN; cb.font=SB if bold else GR
        if fill: cb.fill=fill
        if top: cb.border=TB
    for i in range(12):
        col=get_column_letter(4+i)
        cc=m.cell(rr,4+i, monf(i,col) if monf else f'=$C${rr}')
        cc.number_format=YEN; cc.font=SB if bold else font
        if fill: cc.fill=fill
        if top: cc.border=TB
    t=m.cell(rr,16,f'=SUM(D{rr}:O{rr})'); t.number_format=YEN; t.font=SB; t.fill=fill or TF
    if top: t.border=TB
r=5
m.cell(r,2,'【収益】').font=SB; m.cell(r,2).fill=GRP
for c in range(3,17): m.cell(r,c).fill=GRP
R_S=r+1;  mrow(R_S,'既存事業 売上',f'={Q(S_M)}')
R_AS=r+2; mrow(R_AS,'差額売上（シナリオC）',f"=0")
R_TS=r+3; mrow(R_TS,'売上 合計',f'=C{R_S}+C{R_AS}',lambda i,c:f'={c}{R_S}+{c}{R_AS}',bold=True,fill=TF)
R_C=r+4;  mrow(R_C,'　売上原価',f'=C{R_TS}-C{R_TS+2}-C{R_TS+3}',lambda i,c:f'={c}{R_TS}-{c}{R_TS+2}-{c}{R_TS+3}',font=BK)
R_G=r+5;  mrow(R_G,'既存事業 売上総利益',f'={Q(G_M)}')
R_AG=r+6; mrow(R_AG,'追加売上 総利益',f'=C{R_AS}*{Q(GM)}',lambda i,c:f'={c}{R_AS}*{Q(GM)}')
R_A0=r+7
adv_lbl=['経営指導料 間接グループ5社','経営指導料 Thierry Marx','経営指導料 Apicius']
for k,lab in enumerate(adv_lbl):
    mrow(R_A0+k,'＋'+lab,f'={Q(I0+k)}/12')
R_INC=R_A0+3
mrow(R_INC,'＋M&A仲介（Apicius株主売却）',None,lambda i,c:f'=IF({i+1}={Q(MI)},{Q(MNA)},0)')
R_CR=R_INC+1
mrow(R_CR,'＋新規事業利益（クルーザー）',None,
     lambda i,c:f'=IF({i+1}={Q(MC1)},{Q(CR1)},0)+IF({i+1}={Q(MC2)},{Q(CR2)},0)')
R_GT=R_CR+1
mrow(R_GT,'売上総利益 合計',f'=C{R_G}+C{R_AG}+SUM(C{R_A0}:C{R_A0+2})',
     lambda i,c:f'={c}{R_G}+{c}{R_AG}+SUM({c}{R_A0}:{c}{R_CR})',bold=True,fill=TF,top=True)
r=R_GT+2
m.cell(r,2,'【販売費及び一般管理費】').font=SB; m.cell(r,2).fill=GRP
for c in range(3,17): m.cell(r,c).fill=GRP
G0=r+1
for k,acc in enumerate(ACC):
    mrow(G0+k,acc,f"=AVERAGE('④FY2026月次実績'!I{SG0+k}:K{SG0+k})")
G1=G0+len(ACC)-1
R_SGT=G1+1
mrow(R_SGT,'販管費 合計',f'=SUM(C{G0}:C{G1})',lambda i,c:f'=SUM({c}{G0}:{c}{G1})',bold=True,fill=TF,top=True)
r=R_SGT+2
m.cell(r,2,'【損益】').font=SB; m.cell(r,2).fill=GRP
for c in range(3,17): m.cell(r,c).fill=GRP
R_OP=r+1
mrow(R_OP,'営業利益',f'=C{R_GT}-C{R_SGT}',lambda i,c:f'={c}{R_GT}-{c}{R_SGT}',bold=True,fill=OP,top=True)
R_NOI=R_OP+1; mrow(R_NOI,'営業外収益',None,lambda i,c:'=0')
R_NOE=R_NOI+1; mrow(R_NOE,'営業外費用（支払利息等）',f'={Q(NOE)}/12')
R_ORD=R_NOE+1
mrow(R_ORD,'経常利益',f'=C{R_OP}+C{R_NOI}-C{R_NOE}',lambda i,c:f'={c}{R_OP}+{c}{R_NOI}-{c}{R_NOE}',bold=True,fill=OR_,top=True)
R_CUM=R_ORD+1
m.cell(R_CUM,2,'経常利益 累計').font=SB
for i in range(12):
    col=get_column_letter(4+i)
    f_=f'=D{R_ORD}' if i==0 else f'={get_column_letter(3+i)}{R_CUM}+{col}{R_ORD}'
    c=m.cell(R_CUM,4+i,f_); c.number_format=YEN; c.font=SB; c.fill=OR_
m.cell(R_CUM+2,2,'※既存事業の売上・粗利・販管費は2026/04-06実績の月平均を横ばい。成長・季節変動は織り込んでいない保守前提。').font=SM
m.cell(R_CUM+3,2,'※経営指導料は12か月按分。M&A仲介と新規クルーザー事業は②前提で指定した月に計上（クルーザーは2027/03と2027/09に各10百万円）。').font=SM
m.freeze_panes='D5'

# ============ ①サマリー ============
s=wb.create_sheet('①サマリー',0)
s.column_dimensions['A'].width=3; s.column_dimensions['B'].width=38
for col in 'CDEFG': s.column_dimensions[col].width=18
s.column_dimensions['H'].width=44
s['B1']='FY2027（2027年9月期）利益予測　シナリオ比較'; s['B1'].font=Font(name=F,bold=True,size=15)
s['B2']='ベース：再編後の実力値（2026/04-06 実績3か月平均）を横ばい延伸'; s['B2'].font=Font(name=F,size=10,color='8C8C8C')
s['B3']='単位：円（税別）　出典：事業計画202608（銀行様）月次／決算報告書 第54期'; s['B3'].font=SM
heads=['科目','FY2026 見込','A. 横ばいのみ','B. A＋上乗せ\n132.5百万円','C. 上乗せを\n売上平均の内数','D. 上乗せを\n売上平均に外数','コメント']
for j,t in enumerate(heads):
    c=s.cell(5,2+j,t); c.font=HD; c.fill=HF; c.alignment=Alignment(horizontal='center',wrap_text=True,vertical='center')
s.row_dimensions[5].height=44
b=lambda rr: f"{Q(rr)}*12"
spec=[
 ('既存事業 売上',      708205820, f'={b(S_M)}', f'={b(S_M)}', f'={Q(C_EX)}',f'={Q(D_EX)}','Dの730.5百万円は前期・今期の平均そのもの'),
 ('売上総利益（既存事業）', 229139405, f'={b(G_M)}', f'={b(G_M)}', f'=E7+{Q(C_DGP)}',f'=E7+{Q(D_DGP)}','差額売上の粗利率は30%'),
 ('＋上乗せ案件（営業収益）',0,        0,            f'={Q(ADV)}',  f'={Q(ADV)}',f'={Q(ADV)}','経営指導料92.5＋M&A仲介20＋新規事業利益20'),
 ('売上高（上乗せ含む）',   708205820, '=D6',        '=E6+E8',      f'={Q(C_TOT)}',f'={Q(D_TOT)}','Cは内数のため730.5、Dは外数のため863.0'),
 ('売上総利益 合計',       229139405, '=D7',        '=E7+E8',      f'={Q(C_GT)}',f'={Q(D_GT)}',''),
 ('販管費',              343932836, f'={b(P_M)}', f'={b(P_M)}',  '=E11','=E11','2026/04-06実績の月平均×12'),
 ('営業利益',             -114793431,'=D10-D11',   '=E10-E11',    '=F10-F11','=G10-G11',''),
 ('営業外収益',           0,         0,            0,             0,0,'新規クルーザー事業利益は営業収益に計上'),
 ('営業外費用',           10000000,  f'={Q(NOE)}', f'={Q(NOE)}',  f'={Q(NOE)}',f'={Q(NOE)}',''),
 ('経常利益',             -124793431,'=D12+D13-D14','=E12+E13-E14','=F12+F13-F14','=G12+G13-G14',''),
 ('法人税等',             0,0,0,0,0,'繰越欠損金 約252百万円により実質非課税'),
 ('当期純利益',           -124793431,'=D15-D16',   '=E15-E16',    '=F15-F16','=G15-G16',''),
]
bold={9,10,12,15,17}
for i,(lab,*v) in enumerate(spec):
    rr=6+i; note=v[-1]
    c=s.cell(rr,2,lab); c.font=SB if rr in bold else BK
    for j,val in enumerate(v[:-1]):
        cc=s.cell(rr,3+j,val); cc.number_format=YEN
        cc.font=SB if rr in bold else (BL if j==0 and isinstance(val,int) else BK)
    s.cell(rr,8,note).font=SM
    if rr in bold:
        for c2 in range(2,9): s.cell(rr,c2).fill=TF; s.cell(rr,c2).border=TB
    if rr==12:
        for c2 in range(2,9): s.cell(rr,c2).fill=OP
    if rr==15:
        for c2 in range(2,9): s.cell(rr,c2).fill=OR_

r=21
s.cell(r,2,'【差額売上（Bプランからの上積み）】').font=Font(name=F,bold=True,size=12)
add=[('C：上乗せを内数とした場合の差額売上','='+Q(C_DIFF),'=C23/'+Q(PLAN)),
     ('D：上乗せを外数とした場合の差額売上','='+Q(D_DIFF),'=C24/'+Q(PLAN))]
s.cell(22,3,'追加売上').font=SB; s.cell(22,4,'新規顧客売上計画\n1,066百万円に対する比率').font=SB
s.cell(22,4).alignment=Alignment(wrap_text=True,horizontal='center'); s.row_dimensions[22].height=30
for c2 in (3,4): s.cell(22,c2).fill=SF
for i,(lab,f1,f2) in enumerate(add):
    rr=23+i
    s.cell(rr,2,lab).font=BK
    c=s.cell(rr,3,f1); c.number_format=YEN; c.font=SB; c.fill=YL
    c=s.cell(rr,4,f2); c.number_format=PCT; c.font=SB; c.fill=YL
notes=['','【判定】',
 'A：再編後の実力値を横ばいで延ばすだけでは営業利益▲99.6百万円・経常利益▲114.6百万円。',
 'B：上乗せ案件132.5百万円を計上すると営業利益＋32.9百万円・経常利益＋17.9百万円。契約ベースの下限。',
 'C：上乗せを売上平均の内数とした場合。売上730.5百万円・営業利益＋95.6百万円・経常利益＋80.6百万円。',
 'D：上乗せを外数として加算した場合。売上863.0百万円・営業利益＋135.4百万円・経常利益＋120.4百万円。',
 '',
 '※Bは「何もしなくても」到達する下限。Dが目指すべき最低限の姿。',
 '※既存事業売上はCが598.0百万円（過去2期平均の81.9%）、Dが730.5百万円（同100.0%）。',
 '　　過去2期の売上実績に照らすと、C〜Dのレンジが最も蓋然性が高い。',
 '※差額売上の粗利率30%はFY2025 26.8%・FY2026 32.4%の実績水準。4-6月実績の51.5%は用いない。']
for i,n in enumerate(notes):
    c=s.cell(26+i,2,n); c.font=Font(name=F,bold=True,size=12) if n=='【判定】' else Font(name=F,size=10)
    s.merge_cells(start_row=26+i,start_column=2,end_row=26+i,end_column=8)

# ============ ⑤利益ブリッジ ============
br=wb.create_sheet('⑤利益ブリッジ')
br.column_dimensions['A'].width=3; br.column_dimensions['B'].width=48
br.column_dimensions['C'].width=20; br.column_dimensions['D'].width=20; br.column_dimensions['E'].width=58
br['B1']='FY2026 → FY2027 営業利益ブリッジ（保守シナリオB）'; br['B1'].font=Font(name=F,bold=True,size=14)
br['B2']='単位：円（税別）'; br['B2'].font=SM
for j,t in enumerate(['項目','増減','累計','内容']):
    c=br.cell(4,2+j,t); c.font=HD; c.fill=HF
steps=[('FY2026 営業利益（見込）',-114793431,'現行計画の今期着地見込'),
       ('① 既存事業 粗利の減少',f'={Q(G_M)}*12-229139405','FY2026は2025/11・2026/09の大口を含む。撤退後実力値ベースでは減少'),
       ('② 販管費の削減',f'=343932836-{Q(P_M)}*12','不採算3店舗の撤退＋固定費削減。10-3月平均比で月▲8.58百万円'),
       ('③ 上乗せ案件（営業収益）',f'={Q(ADV)}','経営指導料92.5＋M&A仲介20＋新規事業利益20'),
       ('FY2027 営業利益（保守シナリオB）',None,'①〜③の合計')]
for i,(lab,val,note) in enumerate(steps):
    rr=5+i
    c=br.cell(rr,2,lab); c.font=SB if val is None or i==0 else BK
    if val is not None:
        c=br.cell(rr,3,val); c.number_format=YEN; c.font=SB if i==0 else BK
    br.cell(rr,5,note).font=SM
    if i==0:
        c=br.cell(rr,4,'=C5'); c.number_format=YEN; c.font=SB; c.fill=TF
    elif val is not None:
        c=br.cell(rr,4,f'=D{rr-1}+C{rr}'); c.number_format=YEN; c.font=SB; c.fill=TF
    else:
        c=br.cell(rr,4,f'=D{rr-1}'); c.number_format=YEN; c.font=SB; c.fill=OP
        br.cell(rr,2).fill=OP; br.cell(rr,4).border=TB
br.cell(12,2,'【プランCの差額売上】').font=Font(name=F,bold=True,size=12)
br.cell(13,2,'Bからの差額売上（粗利率30%）').font=BK
br.cell(13,3,"='①サマリー'!C23").number_format=YEN; br.cell(13,3).font=SB; br.cell(13,3).fill=YL
c=br.cell(13,5,"='①サマリー'!D23"); c.number_format=PCT; c.font=SB
br.cell(13,6,'新規顧客売上計画1,066百万円に対する比率').font=SM

# ============ ⑥ワイン仕入・在庫計画 ============
iv=wb.create_sheet('⑥ワイン仕入・在庫計画')
iv.column_dimensions['A'].width=3; iv.column_dimensions['B'].width=30
iv.column_dimensions['C'].width=15
for i in range(12): iv.column_dimensions[get_column_letter(4+i)].width=13
iv.column_dimensions['P'].width=16; iv.column_dimensions['Q'].width=52
MO=['2026/10','2026/11','2026/12','2027/01','2027/02','2027/03','2027/04','2027/05','2027/06','2027/07','2027/08','2027/09']
iv['B1']='ワイン仕入・自社在庫計画（FY2027）'; iv['B1'].font=Font(name=F,bold=True,size=14)
iv['B2']='ワイン仕入は「売上原価（出庫）＋在庫増減」がキャッシュアウト。PLに現れない在庫積み増しが資金繰りの最大論点。'; iv['B2'].font=RD
bar(iv,4,'【1】自社在庫（簿価）の実績推移　単位：百万円',7)
hist=[('2024/12末',560),('2025/07末',470),('2025/09末',480),('2025/12末',530),('2026/03末',620)]
for j,(m,v) in enumerate(hist):
    c=iv.cell(5,3+j,m); c.font=SB; c.fill=SF; c.alignment=Alignment(horizontal='center')
    c=iv.cell(6,3+j,v); c.number_format='#,##0'; c.font=BL; c.alignment=Alignment(horizontal='center')
iv.cell(5,2,'時点').font=SB; iv.cell(5,2).fill=SF
iv.cell(6,2,'自社在庫 簿価').font=BK
iv.cell(7,2,'2025/09末→2026/03末の増加').font=SB
c=iv.cell(7,3,'=C6*0-(E6-G6)'); c.value='=G6-E6'; c.number_format='#,##0'; c.font=SB; c.fill=YL
iv.cell(7,4,'百万円 / 6か月').font=SM
c=iv.cell(7,5,'=C7/6'); c.number_format='#,##0.0'; c.font=SB; c.fill=YL
iv.cell(7,6,'百万円/月のキャッシュアウト').font=RD
iv.cell(8,2,'出典：2026年7月度 取締役会資料「ワイン預かり残高推移KPI」（税抜・簿価）').font=SM

bar(iv,10,'【2】FY2027 仕入・在庫計画（税抜）',16)
iv.cell(11,2,'科目').font=SB; iv.cell(11,2).fill=SF
c=iv.cell(11,3,'月次ベース'); c.font=SB; c.fill=SF; c.alignment=Alignment(horizontal='center')
for i,m in enumerate(MO):
    c=iv.cell(11,4+i,m); c.font=SB; c.fill=SF; c.alignment=Alignment(horizontal='center')
c=iv.cell(11,16,'通期'); c.font=SB; c.fill=SF; c.alignment=Alignment(horizontal='center')
IV_OPEN=12; IV_POL=13; IV_SPOT=14; IV_PRIM=15; IV_NRM=16; IV_BUY=17; IV_OUT=18; IV_END=19
iv.cell(IV_OPEN,2,'期首在庫').font=BK
c=iv.cell(IV_OPEN,3,420000000); c.font=BL; c.number_format=YEN; c.fill=YL
iv.cell(IV_OPEN,17,'2026/03末実績620百万円 −2026/09の在庫販売200百万円＝420百万円。2026/04-08の増減は横ばいと仮定。').font=RD
iv.cell(IV_POL,2,'在庫方針（月次 増減額）').font=BK
c=iv.cell(IV_POL,3,0); c.font=BL; c.number_format=YEN; c.fill=YL
iv.cell(IV_POL,17,'0＝在庫横ばい（仕入＝売上原価）。積み増す場合は月額を入力。').font=SM
iv.cell(IV_SPOT,2,'スポット大口仕入').font=BK
iv.cell(IV_SPOT,17,'野村ユニソン等のロット仕入。発生月に入力。').font=SM
iv.cell(IV_PRIM,2,'プリムール等 前払仕入').font=BK
iv.cell(IV_PRIM,17,'引渡し前の前払。発生月に入力。').font=SM
for i in range(12):
    for r in (IV_SPOT,IV_PRIM):
        c=iv.cell(r,4+i,0); c.font=BL; c.number_format=YEN; c.fill=YL
for r in (IV_SPOT,IV_PRIM):
    c=iv.cell(r,16,f'=SUM(D{r}:O{r})'); c.number_format=YEN; c.font=SB; c.fill=TF
iv.cell(IV_NRM,2,'通常仕入').font=BK
iv.cell(IV_NRM,17,'方針1＝売上原価と同額（在庫横ばい）／方針2＝FY2026の年間仕入額を12等分。').font=SM
iv.cell(IV_BUY,2,'仕入 合計（入庫）').font=SB; iv.cell(IV_BUY,2).border=TB
iv.cell(IV_OUT,2,'売上原価（出庫）').font=BK
iv.cell(IV_END,2,'期末在庫').font=SB; iv.cell(IV_END,2).border=TB
for i in range(12):
    col=get_column_letter(4+i)
    c=iv.cell(IV_NRM,4+i,f'=IF({Q(BUYP)}=2,{Q(BUY26)}/12,{col}{IV_OUT})')
    c.number_format=YEN; c.font=BK
for i in range(12):
    col=get_column_letter(4+i); pv=get_column_letter(3+i)
    c=iv.cell(IV_BUY,4+i,f'={col}{IV_NRM}+$C${IV_POL}+{col}{IV_SPOT}+{col}{IV_PRIM}')
    c.number_format=YEN; c.font=SB; c.fill=TF; c.border=TB
    c=iv.cell(IV_OUT,4+i,f"='③FY2027月次推移'!{col}{R_C}"); c.number_format=YEN; c.font=GR
    prev=f'$C${IV_OPEN}' if i==0 else f'{pv}{IV_END}'
    c=iv.cell(IV_END,4+i,f'={prev}+{col}{IV_BUY}-{col}{IV_OUT}')
    c.number_format=YEN; c.font=SB; c.fill=CREAMFILL; c.border=TB
for r,f_ in [(IV_NRM,f'=SUM(D{IV_NRM}:O{IV_NRM})'),(IV_BUY,f'=SUM(D{IV_BUY}:O{IV_BUY})'),(IV_OUT,f'=SUM(D{IV_OUT}:O{IV_OUT})')]:
    c=iv.cell(r,16,f_); c.number_format=YEN; c.font=SB; c.fill=TF
c=iv.cell(IV_END,16,f'=O{IV_END}'); c.number_format=YEN; c.font=SB; c.fill=CREAMFILL
iv.cell(21,2,'※仕入方針2（FY2026と同額419.1百万円）では、売上原価との差額が在庫増となり、その分がキャッシュアウトになります。').font=RD
iv.cell(22,2,'※スポット大口仕入・プリムール前払は、売上計上より先に支払が発生するため、資金繰り上は最も注意を要する項目。').font=RD
iv.cell(23,2,'※2026/09に在庫200百万円を販売（削減）し、その代金をみずほ銀行への融資返済200百万円に充当する前提。期首在庫はその後の残高。').font=RD

# ============ ⑦FY2027資金繰り表 ============
cf=wb.create_sheet('⑦FY2027資金繰り表')
cf.column_dimensions['A'].width=3; cf.column_dimensions['B'].width=34
cf.column_dimensions['C'].width=14
for i in range(12): cf.column_dimensions[get_column_letter(4+i)].width=12.5
cf.column_dimensions['P'].width=15; cf.column_dimensions['Q'].width=46
cf['B1']='FY2027（2026年10月-2027年9月）月次資金繰り表'; cf['B1'].font=Font(name=F,bold=True,size=14)
cf['B2']='単位：円（税込ベース）　シナリオD（②前提の「シナリオ選択」に連動）／ワイン仕入FY2026同額の前提　青字＝入力セル・緑字＝他シートからのリンク'; cf['B2'].font=SM
bar(cf,4,'【前提】',16)
pre=[('期首現預金（2026/09末）',30000000,'2026/09末にみずほ銀行へ200百万円を返済した後のキャッシュポジション（貴社ご指定）。',True),
     ('消費税率',None,'②前提の【6】消費税の算定から自動参照',False),
     ('売上 回収サイト（か月）',1.0,'売掛金回転32.2日（2025/09末BS）より1.06か月→1.0か月と設定。',False),
     ('仕入 支払サイト（か月）',1.0,'買掛金回転23.5日（2025/09末BS）より0.77か月→1.0か月と設定。',False),
     ('長期借入金 約定返済（月額）',2662179,'★要差替え。長期借入423,623,000円−みずほ一括返済200,000,000円＝223,623,000円を7年均等返済と仮定した暫定値。',True),
     ('短期借入金 純増減（月額）',0,'短期借入162,440,980円は同額借換（継続）を前提。',False)]
for i,(t,v,nt,warn) in enumerate(pre):
    r=5+i
    cf.cell(r,2,t).font=BK
    c=cf.cell(r,3,f'={Q(CT_RATE)}' if t=='消費税率' else v); c.font=GR if t=='消費税率' else BL
    if warn: c.fill=YL
    c.number_format=PCT if t=='消費税率' else ('0.0' if 'サイト' in t else YEN)
    cf.cell(r,4,nt).font=RD if warn else SM
CF_CASH,CF_TAX,CF_AR,CF_AP,CF_LTR,CF_STN=5,6,7,8,9,10
V=lambda r: f"'⑦FY2027資金繰り表'!$C${r}"
Y=lambda r: f"$C${r}"

r0=12
cf.cell(r0,2,'科目').font=SB; cf.cell(r0,2).fill=SF
c=cf.cell(r0,3,'月次ベース'); c.font=SB; c.fill=SF; c.alignment=Alignment(horizontal='center')
for i,m in enumerate(MO):
    c=cf.cell(r0,4+i,m); c.font=SB; c.fill=SF; c.alignment=Alignment(horizontal='center')
c=cf.cell(r0,16,'通期'); c.font=SB; c.fill=SF; c.alignment=Alignment(horizontal='center')

def crow(r,label,monf,basef=None,bold=False,fill=None,font=BK,top=False,tot=True,inp=False):
    c=cf.cell(r,2,label); c.font=SB if bold else font
    if fill: c.fill=fill
    if top: c.border=TB
    if basef is not None:
        cb=cf.cell(r,3,basef); cb.number_format=YEN; cb.font=SB if bold else GR
        if fill: cb.fill=fill
    for i in range(12):
        col=get_column_letter(4+i)
        cc=cf.cell(r,4+i, monf(i,col))
        cc.number_format=YEN; cc.font=SB if bold else (BL if inp else font)
        if inp: cc.fill=YL
        elif fill: cc.fill=fill
        if top: cc.border=TB
    if tot:
        t=cf.cell(r,16,f'=SUM(D{r}:O{r})'); t.number_format=YEN; t.font=SB; t.fill=fill or TF
        if top: t.border=TB

hd=r0+1
cf.cell(hd,2,'【営業収入】').font=SB; cf.cell(hd,2).fill=GRP
for c in range(3,17): cf.cell(hd,c).fill=GRP
IN_S=hd+1; IN_A=hd+2; IN_I=hd+3; IN_T=hd+6
crow(IN_S,'売上入金（税込）',lambda i,c:f"='③FY2027月次推移'!{c}{R_TS}*(1+{Y(CF_TAX)})",font=GR)
crow(IN_A,'経営指導料 入金（税込）',lambda i,c:f"=SUM('③FY2027月次推移'!{c}{R_A0}:{c}{R_A0+2})*(1+{Y(CF_TAX)})",font=GR)
crow(IN_I,'M&A仲介 入金（税込）',lambda i,c:f"='③FY2027月次推移'!{c}{R_INC}*(1+{Y(CF_TAX)})",font=GR)
IN_R=IN_I+1
recv=[0,0,0,0,0,0,0,0,0,0,0,0]
crow(IN_R,'前期末売掛金 回収',lambda i,c:recv[i],inp=True)
IN_CR=IN_R+1
crow(IN_CR,'新規事業利益 入金（税込）',lambda i,c:f"='③FY2027月次推移'!{c}{R_CR}*(1+{Y(CF_TAX)})",font=GR)
cf.cell(IN_CR,17,'新規クルーザー事業。2027/03と2027/09に各10百万円。課税取引として税込入金と仮定。').font=SM
cf.cell(IN_R,17,'2026/09の在庫販売200百万円の代金回収およびみずほ銀行への返済は2026/09末までに完了済み（期首現預金30百万円に反映）。').font=RD
crow(IN_T,'営業収入 計',lambda i,c:f'=SUM({c}{IN_S}:{c}{IN_CR})',bold=True,fill=TF,top=True)

hd2=IN_T+2
cf.cell(hd2,2,'【営業支出】').font=SB; cf.cell(hd2,2).fill=GRP
for c in range(3,17): cf.cell(hd2,c).fill=GRP
OUT_W=hd2+1; OUT_P=hd2+2; OUT_O=hd2+3; OUT_INT=hd2+4; OUT_CT=hd2+5; OUT_TX=hd2+6; OUT_T=hd2+7
crow(OUT_W,'ワイン仕入 支払（税込）',
     lambda i,c:(f"='⑥ワイン仕入・在庫計画'!D{IV_BUY}*(1+{Y(CF_TAX)})" if i==0
                 else f"='⑥ワイン仕入・在庫計画'!{get_column_letter(3+i)}{IV_BUY}*(1+{Y(CF_TAX)})"),font=GR)
HR=[SG0+ACC.index(a) for a in ['役員報酬','従業員給与','雑給','法定福利費']]
hexpr='+'.join([f"AVERAGE('④FY2026月次実績'!I{x}:K{x})" for x in HR])
crow(OUT_P,'人件費（役員報酬・給与・雑給・法定福利）',lambda i,c:f'={hexpr}',font=GR)
NTR=[SG0+ACC.index(a) for a in ['租税公課','保険料']]
ntexpr='+'.join([f"AVERAGE('④FY2026月次実績'!I{x}:K{x})" for x in NTR])
depr=SG0+ACC.index('減価償却費')
crow(OUT_O,'その他販管費（税込・減価償却除く）',
     lambda i,c:(f"=(('③FY2027月次推移'!{c}{R_SGT}-AVERAGE('④FY2026月次実績'!I{depr}:K{depr})-({hexpr})-({ntexpr}))*(1+{Y(CF_TAX)}))+({ntexpr})"),font=GR)
crow(OUT_INT,'支払利息',lambda i,c:f"='③FY2027月次推移'!{c}{R_NOE}",font=GR)
def _ctax(i,c):
    if i==1:  return f'={Q(CT_FIN)}'      # 2026/11 確定納付
    if i in (4,7,10): return f'={Q(CT_MID)}'  # 2027/02・05・08 中間納付
    return 0
crow(OUT_CT,'消費税 納付',_ctax,font=GR)
cf.cell(OUT_CT,17,'②前提【6】の算定結果を参照。2026/11＝FY2026分の確定納付、2027/02・05・08＝当期の中間納付（年3回）。').font=RD
tx=[0,463105,0,0,0,0,0,0,0,0,0,0]
crow(OUT_TX,'法人税等（均等割）',lambda i,c:tx[i],inp=True)
crow(OUT_T,'営業支出 計',lambda i,c:f'=SUM({c}{OUT_W}:{c}{OUT_TX})',bold=True,fill=TF,top=True)
CF_ORD=OUT_T+1
crow(CF_ORD,'経常収支',lambda i,c:f'={c}{IN_T}-{c}{OUT_T}',bold=True,fill=OP,top=True)

hd3=CF_ORD+2
cf.cell(hd3,2,'【財務収支】').font=SB; cf.cell(hd3,2).fill=GRP
for c in range(3,17): cf.cell(hd3,c).fill=GRP
FI_R=hd3+1; FI_S=hd3+2; FI_N=hd3+3; FI_T=hd3+5
crow(FI_R,'長期借入金 約定返済',lambda i,c:f'=-{Y(CF_LTR)}',font=GR)
crow(FI_S,'短期借入金 純増減',lambda i,c:f'={Y(CF_STN)}',font=GR)
crow(FI_N,'新規調達（借入・増資）',lambda i,c:0,inp=True)
FI_M=FI_N+1
miz=[0,0,0,0,0,0,0,0,0,0,0,0]
crow(FI_M,'みずほ銀行 融資返済',lambda i,c:miz[i],inp=True)
cf.cell(FI_M,17,'2026/09末に200百万円を返済済みのためFY2027には計上しない。').font=RD
crow(FI_T,'財務収支 計',lambda i,c:f'=SUM({c}{FI_R}:{c}{FI_M})',bold=True,fill=TF,top=True)

CF_NET=FI_T+2; CF_BEG=CF_NET+1; CF_END=CF_NET+2; CF_SHORT=CF_NET+3
crow(CF_NET,'当月収支',lambda i,c:f'={c}{CF_ORD}+{c}{FI_T}',bold=True,fill=TF,top=True)
crow(CF_BEG,'前月繰越 現預金',lambda i,c:(f'={Y(CF_CASH)}' if i==0 else f'={get_column_letter(3+i)}{CF_END}'),tot=False)
crow(CF_END,'翌月繰越 現預金',lambda i,c:f'={c}{CF_BEG}+{c}{CF_NET}',bold=True,fill=OR_,top=True,tot=False)
crow(CF_SHORT,'資金不足額（マイナス時）',lambda i,c:f'=IF({c}{CF_END}<0,{c}{CF_END},0)',bold=True,fill=PatternFill('solid',fgColor='E0DEDA'),tot=False)
c=cf.cell(CF_END,16,f'=O{CF_END}'); c.number_format=YEN; c.font=SB; c.fill=OR_
c=cf.cell(CF_SHORT,16,f'=MIN(D{CF_SHORT}:O{CF_SHORT})'); c.number_format=YEN; c.font=SB; c.fill=PatternFill('solid',fgColor='E0DEDA')
cf.cell(CF_SHORT,17,'通期列＝最大不足額（必要調達額の下限）').font=RD

n=CF_SHORT+2
notes=['※売上・仕入とも回収／支払サイト1.0か月。既存事業の売上は横ばい前提のため、月ズレの影響は通期では相殺されます。',
 '※ワイン仕入の支払は前月仕入額を当月に支払う前提（⑥シートで仕入月を入力すると、翌月に資金流出として反映されます）。',
 '※減価償却費（月3.6百万円）は資金支出を伴わないため、営業支出から除外しています。',
 '※消費税納付・法人税等は暫定値です。税理士確定額に差し替えてください。',
 '※期首現預金と長期借入返済額（黄色セル）は必ず実際の数値に差し替えてください。この2つで結論が大きく変わります。']
for i,t in enumerate(notes):
    cf.cell(n+i,2,t).font=RD if i>=3 else SM
cf.freeze_panes='D13'

# 感応度分析
sr=CF_SHORT+8
bar(cf,sr,'【感応度】どこまで耐えられるか',16)
cf.cell(sr+1,2,'ケース').font=SB; cf.cell(sr+1,2).fill=SF
for j,t in enumerate(['期末 現預金残高','期中 最低残高','必要調達額']):
    c=cf.cell(sr+1,3+j,t); c.font=SB; c.fill=SF; c.alignment=Alignment(horizontal='center',wrap_text=True)
cf.cell(sr+1,6,'前提').font=SB; cf.cell(sr+1,6).fill=SF
BUF=sr+6
cases=[('A. 基本（シナリオD：仕入FY2026同額・CFが回る売上）',f'=O{CF_END}',f'=MIN(D{CF_END}:O{CF_END})',
        '期首現預金30百万円、ワイン仕入FY2026と同額419.1百万円、売上621.1百万円'),
       ('B. ワイン仕入を在庫横ばいにした場合',f"=C{sr+2}+({Q(BUY26)}-'③FY2027月次推移'!$P${R_C})*(1+{Y(CF_TAX)})",
        f"=D{sr+2}+({Q(BUY26)}-'③FY2027月次推移'!$P${R_C})*(1+{Y(CF_TAX)})",
        '仕入＝売上原価（②前提のワイン仕入方針を1に変更）。在庫は増えない'),
       ('C. 2027/03・09の一時収益が入らない',f'=C{sr+2}-(P{IN_I}+P{IN_CR})',f'=D{sr+2}-(P{IN_I}+P{IN_CR})',
        'M&A仲介22.0百万円＋新規事業利益22.0百万円（いずれも税込）が未入金'),
       ('D. 追加売上が未達（プランBにとどまる）',f"=C{sr+2}-'①サマリー'!$C$24*(1+{Y(CF_TAX)})",
        f"=D{sr+2}-'①サマリー'!$C$24*(1+{Y(CF_TAX)})",
        '追加売上231.9百万円が未達。仕入方針2では仕入額が減らないため全額が資金不足に直結')]
for i,(lab,fe,fm,nt) in enumerate(cases):
    r=sr+2+i
    cf.cell(r,2,lab).font=SB if i==0 else BK
    for j,f_ in enumerate([fe,fm]):
        c=cf.cell(r,3+j,f_); c.number_format=YEN; c.font=SB
        c.fill=TF if i==0 else PatternFill('solid',fgColor='E0DEDA')
    c=cf.cell(r,5,f'=IF(C{r+0}<0,-MIN(C{r},D{r})+$C${BUF},MAX(0,-D{r}+$C${BUF}))')
    c.number_format=YEN; c.font=SB; c.fill=YL
    cf.cell(r,6,nt).font=SM
cf.cell(BUF,2,'運転資金バッファ（月商1か月）').font=BK
c=cf.cell(BUF,3,f"='③FY2027月次推移'!$C${R_TS}*(1+{Y(CF_TAX)})"); c.number_format=YEN; c.font=GR
cf.cell(BUF,6,'最低残高をゼロに戻すだけでは足りないため、月商1か月分を上乗せして必要調達額を算定。').font=SM
cf.cell(BUF+2,2,'※ケースAとBの差が、ワイン仕入方針（在庫を積むか積まないか）による資金インパクトです。').font=RD
cf.cell(BUF+3,2,'※ケースCとDが同時に起きた場合、必要調達額は両者の合計に近い水準となります。').font=RD

# 現預金残高の推移グラフ
from openpyxl.chart import LineChart, Reference
ch=LineChart(); ch.title='月末 現預金残高の推移（ケースA：在庫横ばい・調達なし）'
ch.y_axis.title='円'; ch.x_axis.title=None; ch.height=7.6; ch.width=24
data=Reference(cf,min_col=4,max_col=15,min_row=CF_END,max_row=CF_END)
cats=Reference(cf,min_col=4,max_col=15,min_row=r0,max_row=r0)
ch.add_data(data,from_rows=True,titles_from_data=False)
ch.set_categories(cats)
ch.series[0].graphicalProperties.line.width=28000
ch.series[0].graphicalProperties.line.solidFill='1A1A1A'
ch.legend=None
cf.add_chart(ch,f'B{BUF+5}')

# 【必要売上高の算定】ワイン仕入を固定した場合にCFが回る売上
RS=BUF+24
bar(cf,RS,'【必要売上高の算定】ワイン仕入をFY2026と同水準（419.1百万円）に維持した場合',16)
cf.cell(RS+1,2,'目標 期中最低残高').font=BK
c=cf.cell(RS+1,3,0); c.number_format=YEN; c.font=BL; c.fill=YL
cf.cell(RS+1,6,'0＝期中に一度も現預金がマイナスにならない水準。バッファを確保したい場合は金額を入力。').font=SM
sga_m=Q(P_M)
other_expr=f"(({sga_m}-AVERAGE('④FY2026月次実績'!I{depr}:K{depr})-({hexpr})-({ntexpr}))*(1+{Y(CF_TAX)}))+({ntexpr})"
cf.cell(RS+2,2,'月次ネット収支（追加売上を除く）').font=BK
c=cf.cell(RS+2,3,f"={Q(S_M)}*(1+{Y(CF_TAX)})+SUM('②前提'!$C${I0}:$C${I0+2})/12*(1+{Y(CF_TAX)})"
                 f"-{Q(BUY26)}/12*(1+{Y(CF_TAX)})-({hexpr})-({other_expr})-{Q(NOE)}/12-{Y(CF_LTR)}")
c.number_format=YEN; c.font=SB; c.fill=TF
cf.cell(RS+2,6,'既存事業売上＋経営指導料の入金から、ワイン仕入・人件費・その他販管費・支払利息・長期借入返済を差し引いた月次ネット。').font=SM
RH=RS+4
cf.cell(RH,2,'経過月').font=SB; cf.cell(RH,2).fill=SF
for i in range(12):
    c=cf.cell(RH,4+i,i+1); c.font=SB; c.fill=SF; c.alignment=Alignment(horizontal='center')
    c=cf.cell(RH+1,4+i,MO[i]); c.font=SM; c.alignment=Alignment(horizontal='center')
RB=RH+2; RN=RH+3
cf.cell(RB,2,'基礎 累計現預金（追加売上なし）').font=BK
cf.cell(RN,2,'その月をゼロにするのに必要な追加売上').font=BK
for i in range(12):
    col=get_column_letter(4+i)
    oi=f"SUM($D${IN_I}:{col}{IN_I})+SUM($D${IN_CR}:{col}{IN_CR})"
    oo=f"SUM($D${OUT_CT}:{col}{OUT_CT})+SUM($D${OUT_TX}:{col}{OUT_TX})"
    c=cf.cell(RB,4+i,f'=$C${CF_CASH}+{col}${RH}*$C${RS+2}+({oi})-({oo})')
    c.number_format=YEN; c.font=BK
    c=cf.cell(RN,4+i,f'=MAX(0,($C${RS+1}-{col}{RB})*12/((1+{Y(CF_TAX)})*{col}${RH}))')
    c.number_format=YEN; c.font=BK
RQ=RN+2
cf.cell(RQ,2,'必要な追加売上（年額）').font=SB
c=cf.cell(RQ,3,f'=MAX(D{RN}:O{RN})'); c.number_format=YEN; c.font=SB; c.fill=YL
cf.cell(RQ+1,2,'必要な売上高（既存＋追加）').font=SB
c=cf.cell(RQ+1,3,f'={Q(S_M)}*12+C{RQ}'); c.number_format=YEN; c.font=SB; c.fill=YL
cf.cell(RQ+2,2,'ボトルネックとなる月').font=BK
c=cf.cell(RQ+2,3,f'=INDEX(D{RH+1}:O{RH+1},MATCH(C{RQ},D{RN}:O{RN},0))'); c.font=SB
cf.cell(RQ+2,6,'この月の資金需要が最も大きく、必要売上高を決めています。').font=SM
cf.cell(RQ+4,2,'※追加売上は12か月に均等按分し、当月入金する前提。売上が期の前半に寄れば必要額は下がります。').font=RD
cf.cell(RQ+5,2,'※本算定はワイン仕入方針2（FY2026と同額の固定仕入）を前提としています。方針1では仕入が売上原価に連動するため別途算定が必要です。').font=RD
cf.cell(RQ+6,2,'※追加売上に係る消費税は翌期（2027/11）の納付となるため、FY2027の資金繰りには現れません。').font=RD

# ③の追加売上をシナリオ3対応に書き戻す（RQ確定後）
_as=f"=IF({Q(SCN)}=3,{Q(D_DIFF)}/12,IF({Q(SCN)}=2,{Q(C_DIFF)}/12,0))"
_m3=wb['③FY2027月次推移']
c=_m3.cell(R_AS,3,_as); c.number_format=YEN; c.font=GR
x=_m3.cell(R_AG,3,f'=C{R_AS}*{Q(C_GM)}'); x.number_format=YEN; x.font=GR
for _i in range(12):
    _c=get_column_letter(4+_i)
    c=_m3.cell(R_AS,4+_i,_as); c.number_format=YEN; c.font=GR
    x=_m3.cell(R_AG,4+_i,f'={_c}{R_AS}*{Q(C_GM)}'); x.number_format=YEN; x.font=GR



wb.save('WineBank_FY2027_forecast.xlsx'); print('saved')
