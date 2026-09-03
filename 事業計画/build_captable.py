# -*- coding: utf-8 -*-
"""株式会社Brasserie Thierry Marx 資本政策（修正版）
出典: 合意書(2026年9月)、資本政策202609(投資家様)、事業収支資料20260904、5か年事業計画
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

OUT = "/home/user/wine-auction/事業計画/株式会社Brasserie Thierry Marx_資本政策_修正版_202609.xlsx"
JP = "Meiryo"
BLUE, BLACK, GREEN, RED = "0000FF", "000000", "008000", "C00000"
HDR = PatternFill("solid", fgColor="1F3864")
SUB = PatternFill("solid", fgColor="D9E2F3")
TOT = PatternFill("solid", fgColor="F2F2F2")
KEY = PatternFill("solid", fgColor="FFFF00")
NG  = PatternFill("solid", fgColor="FFC7CE")
OKF = PatternFill("solid", fgColor="C6EFCE")
thin = Side(style="thin", color="BFBFBF")
BOX = Border(left=thin, right=thin, top=thin, bottom=thin)

SH = '#,##0;(#,##0);-'          # 株数
JPY = '#,##0;(#,##0);-'          # 円
TH = '#,##0;(#,##0);-'           # 千円
MM = '#,##0.0;(#,##0.0);-'       # 百万円
PCT = '0.0%;(0.0%);-'
PCT2 = '0.00%;(0.00%);-'

wb = openpyxl.Workbook()

def S(ws, cell, *, bold=False, color=BLACK, fmt=None, fill=None, align=None,
      size=10, border=False, wrap=False):
    c = ws[cell]
    c.font = Font(name=JP, bold=bold, color=color, size=size)
    if fmt: c.number_format = fmt
    if fill: c.fill = fill
    if align or wrap: c.alignment = Alignment(horizontal=align, vertical="center", wrap_text=wrap)
    if border: c.border = BOX
    return c

def band(ws, row, text, last, fill=HDR, color="FFFFFF", size=11):
    ws.cell(row=row, column=1, value=text)
    for col in range(1, openpyxl.utils.column_index_from_string(last) + 1):
        c = ws.cell(row=row, column=col)
        c.fill = fill
        c.font = Font(name=JP, bold=True, color=color, size=size)

def sec(ws, row, text, last):
    band(ws, row, text, last, fill=SUB, color="000000", size=10)

# ============================================================
# 1. 前提（入力）
# ============================================================
ws1 = wb.create_sheet("前提")
ws1.sheet_view.showGridLines = False
ws1.column_dimensions["A"].width = 36
ws1.column_dimensions["B"].width = 16
ws1.column_dimensions["C"].width = 4
ws1.column_dimensions["D"].width = 74

band(ws1, 1, "前提条件（青字＝入力値）", "D")
S(ws1, "A2", size=9, color="7F7F7F")
ws1["A2"] = "出典：合意書（2026年9月）／資本政策202609（投資家様）／事業収支資料20260904／5か年事業計画FY2027-2031"

sec(ws1, 4, "■ 現在の株主構成（2026年9月末）", "D")
cur = [
    ("中野邦人　保有株数", 1300, SH, "資本政策202609より"),
    ("合同会社WineBank P3（WBP3）保有株数", 700, SH, "アピシウスの株主。対象会社の直接株主はこの2者のみ"),
    ("資本金（千円）", 10000, TH, "資本政策202609より"),
    ("資本準備金（千円）", 0, TH, ""),
]
r = 5
for lab, v, f, note in cur:
    S(ws1, f"A{r}"); ws1[f"A{r}"] = lab
    S(ws1, f"B{r}", color=BLUE, fmt=f, border=True, align="center"); ws1[f"B{r}"] = v
    S(ws1, f"D{r}", size=9, color="7F7F7F"); ws1[f"D{r}"] = note
    r += 1
S(ws1, "A9", bold=True); ws1["A9"] = "現在の発行済株式総数"
S(ws1, "B9", bold=True, fmt=SH, border=True, align="center", fill=TOT); ws1["B9"] = "=B5+B6"
# B5 中野 B6 WBP3 B7 資本金 B8 資本準備金 B9 計

sec(ws1, 11, "■ 第三者割当①（2026年10月末＝本件実行日）", "D")
r1 = [
    ("1株あたり発行価額（円）", 150000, JPY, "Preバリュー300,000,000円 ÷ 現在2,000株＝@150,000円"),
    ("投資家A（AnyColorオーナー）割当株数", 133, SH, "合意書は「金20,000,000円」。@150,000では133.33株となり端数が出る（後掲チェック参照）"),
    ("投資家B　割当株数", 267, SH, "4,000万円枠。266.67株を切上げ"),
    ("投資家C　割当株数", 400, SH, "6,000万円枠。端数なし"),
    ("投資家D　割当株数", 200, SH, "3,000万円枠。端数なし"),
    ("資本金への計上割合", 0.5, PCT, "残額は資本準備金へ（会社法445条2項・3項）"),
]
r = 12
for lab, v, f, note in r1:
    S(ws1, f"A{r}"); ws1[f"A{r}"] = lab
    S(ws1, f"B{r}", color=BLUE, fmt=f, border=True, align="center"); ws1[f"B{r}"] = v
    S(ws1, f"D{r}", size=9, color="7F7F7F"); ws1[f"D{r}"] = note
    r += 1
S(ws1, "A18", bold=True); ws1["A18"] = "割当株数　計"
S(ws1, "B18", bold=True, fmt=SH, border=True, align="center", fill=TOT); ws1["B18"] = "=SUM(B13:B16)"
S(ws1, "A19", bold=True); ws1["A19"] = "本件出資総額（円）"
S(ws1, "B19", bold=True, fmt=JPY, border=True, align="center", fill=KEY); ws1["B19"] = "=B18*B12"
S(ws1, "D19", size=9, color=RED, bold=True); ws1["D19"] = "合意書：総額150,000,000円の募集"
# B12 価額 B13-16 各割当 B17 資本金割合 B18 計 B19 総額

sec(ws1, 21, "■ ストックオプション（2028年想定）", "D")
so = [
    ("SO付与株数（潜在株式）", 300, SH, "資本政策202609の備考「潜在株10％以内」を発行済3,000株の10％として設定"),
]
r = 22
for lab, v, f, note in so:
    S(ws1, f"A{r}"); ws1[f"A{r}"] = lab
    S(ws1, f"B{r}", color=BLUE, fmt=f, border=True, align="center"); ws1[f"B{r}"] = v
    S(ws1, f"D{r}", size=9, color="7F7F7F"); ws1[f"D{r}"] = note
    r += 1
# B22 SO

sec(ws1, 24, "■ 第三者割当②（2029年・想定枠）", "D")
r2 = [
    ("割当株数", 0, SH, "事業収支資料p11「調達金額2.0-3.0億円」を実現する場合の枠。バリュエーション未確定のため0で設定"),
    ("1株あたり発行価額（円）", 0, JPY, "旧資本政策の@192,000円は前回@150,000円より高いが、2027年の@50,000円はダウンラウンドで整合しない"),
]
r = 25
for lab, v, f, note in r2:
    S(ws1, f"A{r}"); ws1[f"A{r}"] = lab
    S(ws1, f"B{r}", color=BLUE, fmt=f, border=True, align="center"); ws1[f"B{r}"] = v
    S(ws1, f"D{r}", size=9, color="7F7F7F"); ws1[f"D{r}"] = note
    r += 1
# B25 割当 B26 価額

sec(ws1, 28, "■ 合意書の確約条件", "D")
cov = [
    ("甲側 持株比率の下限", 0.66, PCT, "合意書 第2条3項3号。現行文は「WBP3の持株比率」だがWBP3単独は23.3%。中野邦人との合計と解する"),
    ("外部投資家 比率（合意書表記）", 0.34, PCT, "合意書「本件34％投資家」。実際の外部合計は33.3%"),
    ("投資家A 出資額（合意書記載・円）", 20000000, JPY, "合意書 第1条1項"),
    ("Preバリュー（合意書記載・円）", 300000000, JPY, "合意書 第1条1項"),
    ("本件出資総額（合意書記載・円）", 150000000, JPY, "合意書 第1条1項"),
]
r = 29
for lab, v, f, note in cov:
    S(ws1, f"A{r}"); ws1[f"A{r}"] = lab
    S(ws1, f"B{r}", color=BLUE, fmt=f, border=True, align="center"); ws1[f"B{r}"] = v
    S(ws1, f"D{r}", size=9, color="7F7F7F"); ws1[f"D{r}"] = note
    r += 1
# B29 下限66% B30 34% B31 2000万 B32 Pre3億 B33 総額1.5億

# ============================================================
# 2. 資本政策
# ============================================================
ws2 = wb.create_sheet("資本政策")
ws2.sheet_view.showGridLines = False
ws2.column_dimensions["A"].width = 4
ws2.column_dimensions["B"].width = 30
for c in ["C","D","E","F","G","H","I","J","K","L","M"]:
    ws2.column_dimensions[c].width = 12
ws2.column_dimensions["F"].width = 15

band(ws2, 1, "株式会社Brasserie Thierry Marx　資本政策（修正版）", "M")
S(ws2, "A2", size=9, color="7F7F7F")
ws2["A2"] = "単位：株／円。青字＝「前提」シートの入力値に連動。※旧版の「株式譲渡」列見出しは実態（第三者割当増資）に合わせて修正済み"

# ラウンド見出し
groups = [("C","D","現在\n2026年9月末"),
          ("E","H","第三者割当①\n2026年10月末（本件実行日）"),
          ("I","J","SO付与\n2028年"),
          ("K","M","第三者割当②\n2029年（想定枠）")]
for c1, c2, t in groups:
    ws2.merge_cells(f"{c1}3:{c2}3")
    S(ws2, f"{c1}3", bold=True, fill=SUB, align="center", size=9, wrap=True); ws2[f"{c1}3"] = t
ws2.row_dimensions[3].height = 30
hdrs = [("C","株数"),("D","比率"),
        ("E","割当株数"),("F","払込金額(円)"),("G","発行済株数"),("H","比率"),
        ("I","潜在込株数"),("J","潜在込比率"),
        ("K","割当株数"),("L","潜在込株数"),("M","潜在込比率")]
S(ws2, "A4", bold=True, border=True, fill=TOT, align="center"); ws2["A4"] = "No"
S(ws2, "B4", bold=True, border=True, fill=TOT); ws2["B4"] = "株主"
for c, t in hdrs:
    S(ws2, f"{c}4", bold=True, border=True, fill=TOT, align="center", size=9, wrap=True); ws2[f"{c}4"] = t
ws2.row_dimensions[4].height = 26

# 行定義: (No, 名称, 現在株数式, 割当式, SO付与式, R2割当式)
ROWS = [
    (1, "中野邦人",                       "=前提!$B$5", None,          None, None),
    (2, "合同会社WineBank P3（WBP3）",     "=前提!$B$6", None,          None, None),
]
R_NAKANO, R_WBP3 = 5, 6
R_SUB_KOU = 7
R_INVA, R_INVB, R_INVC, R_INVD = 8, 9, 10, 11
R_SUB_EXT = 12
R_SO = 13
R_TOTAL = 14

S(ws2, f"A{R_NAKANO}", border=True, align="center"); ws2[f"A{R_NAKANO}"] = 1
S(ws2, f"B{R_NAKANO}", border=True); ws2[f"B{R_NAKANO}"] = "中野邦人"
S(ws2, f"A{R_WBP3}", border=True, align="center"); ws2[f"A{R_WBP3}"] = 2
S(ws2, f"B{R_WBP3}", border=True); ws2[f"B{R_WBP3}"] = "合同会社WineBank P3（WBP3）"
S(ws2, f"B{R_SUB_KOU}", bold=True, border=True); ws2[f"B{R_SUB_KOU}"] = "小計　甲側（中野＋WBP3）"
inv_names = [(R_INVA, 3, "投資家A（AnyColorオーナー）", "$B$13"),
             (R_INVB, 4, "投資家B", "$B$14"),
             (R_INVC, 5, "投資家C", "$B$15"),
             (R_INVD, 6, "投資家D", "$B$16")]
for rr, no, nm, ref in inv_names:
    S(ws2, f"A{rr}", border=True, align="center"); ws2[f"A{rr}"] = no
    S(ws2, f"B{rr}", border=True); ws2[f"B{rr}"] = nm
S(ws2, f"B{R_SUB_EXT}", bold=True, border=True); ws2[f"B{R_SUB_EXT}"] = "小計　外部投資家"
S(ws2, f"B{R_SO}", border=True); ws2[f"B{R_SO}"] = "ストックオプション（未行使）"
S(ws2, f"B{R_TOTAL}", bold=True, border=True); ws2[f"B{R_TOTAL}"] = "合　計"

# --- 現在 (C,D) ---
ws2[f"C{R_NAKANO}"] = "=前提!$B$5"
ws2[f"C{R_WBP3}"] = "=前提!$B$6"
ws2[f"C{R_SUB_KOU}"] = f"=SUM(C{R_NAKANO}:C{R_WBP3})"
for rr, _, _, _ in inv_names:
    ws2[f"C{rr}"] = 0
ws2[f"C{R_SUB_EXT}"] = f"=SUM(C{R_INVA}:C{R_INVD})"
ws2[f"C{R_SO}"] = 0
ws2[f"C{R_TOTAL}"] = f"=C{R_SUB_KOU}+C{R_SUB_EXT}"
for rr in range(R_NAKANO, R_TOTAL + 1):
    bold = rr in (R_SUB_KOU, R_SUB_EXT, R_TOTAL)
    S(ws2, f"C{rr}", bold=bold, color=GREEN if rr in (R_NAKANO, R_WBP3) else BLACK,
      fmt=SH, border=True, fill=TOT if bold else None)
    S(ws2, f"D{rr}", bold=bold, fmt=PCT, border=True, align="center", fill=TOT if bold else None)
    ws2[f"D{rr}"] = f"=IF($C${R_TOTAL}=0,0,C{rr}/$C${R_TOTAL})"

# --- 第三者割当① (E,F,G,H) ---
ws2[f"E{R_NAKANO}"] = 0
ws2[f"E{R_WBP3}"] = 0
ws2[f"E{R_SUB_KOU}"] = f"=SUM(E{R_NAKANO}:E{R_WBP3})"
for rr, _, _, ref in inv_names:
    ws2[f"E{rr}"] = f"=前提!{ref}"
ws2[f"E{R_SUB_EXT}"] = f"=SUM(E{R_INVA}:E{R_INVD})"
ws2[f"E{R_SO}"] = 0
ws2[f"E{R_TOTAL}"] = f"=E{R_SUB_KOU}+E{R_SUB_EXT}"
for rr in range(R_NAKANO, R_TOTAL + 1):
    bold = rr in (R_SUB_KOU, R_SUB_EXT, R_TOTAL)
    S(ws2, f"E{rr}", bold=bold, color=GREEN if rr in [i[0] for i in inv_names] else BLACK,
      fmt=SH, border=True, fill=TOT if bold else None)
    S(ws2, f"F{rr}", bold=bold, fmt=JPY, border=True, fill=TOT if bold else None)
    ws2[f"F{rr}"] = f"=E{rr}*前提!$B$12"
    S(ws2, f"G{rr}", bold=bold, fmt=SH, border=True, fill=TOT if bold else None)
    ws2[f"G{rr}"] = f"=C{rr}+E{rr}"
    S(ws2, f"H{rr}", bold=bold, fmt=PCT, border=True, align="center",
      fill=KEY if rr == R_SUB_KOU else (TOT if bold else None))
    ws2[f"H{rr}"] = f"=IF($G${R_TOTAL}=0,0,G{rr}/$G${R_TOTAL})"

# --- SO付与 (I,J) 潜在株式考慮後 ---
for rr in range(R_NAKANO, R_TOTAL + 1):
    bold = rr in (R_SUB_KOU, R_SUB_EXT, R_TOTAL)
    S(ws2, f"I{rr}", bold=bold, fmt=SH, border=True, fill=TOT if bold else None)
    if rr == R_SO:
        ws2[f"I{rr}"] = "=前提!$B$22"
    elif rr == R_SUB_KOU:
        ws2[f"I{rr}"] = f"=SUM(I{R_NAKANO}:I{R_WBP3})"
    elif rr == R_SUB_EXT:
        ws2[f"I{rr}"] = f"=SUM(I{R_INVA}:I{R_INVD})"
    elif rr == R_TOTAL:
        ws2[f"I{rr}"] = f"=I{R_SUB_KOU}+I{R_SUB_EXT}+I{R_SO}"
    else:
        ws2[f"I{rr}"] = f"=G{rr}"
    S(ws2, f"J{rr}", bold=bold, fmt=PCT, border=True, align="center",
      fill=NG if rr == R_SUB_KOU else (TOT if bold else None))
    ws2[f"J{rr}"] = f"=IF($I${R_TOTAL}=0,0,I{rr}/$I${R_TOTAL})"

# --- 第三者割当② (K,L,M) ---
for rr in range(R_NAKANO, R_TOTAL + 1):
    bold = rr in (R_SUB_KOU, R_SUB_EXT, R_TOTAL)
    S(ws2, f"K{rr}", bold=bold, fmt=SH, border=True, fill=TOT if bold else None)
    if rr == R_SUB_EXT:
        ws2[f"K{rr}"] = "=前提!$B$25"
    elif rr in (R_SUB_KOU, R_SO):
        ws2[f"K{rr}"] = 0
    elif rr == R_TOTAL:
        ws2[f"K{rr}"] = f"=K{R_SUB_KOU}+K{R_SUB_EXT}"
    else:
        ws2[f"K{rr}"] = 0
    S(ws2, f"L{rr}", bold=bold, fmt=SH, border=True, fill=TOT if bold else None)
    ws2[f"L{rr}"] = f"=I{rr}+K{rr}"
    S(ws2, f"M{rr}", bold=bold, fmt=PCT, border=True, align="center", fill=TOT if bold else None)
    ws2[f"M{rr}"] = f"=IF($L${R_TOTAL}=0,0,L{rr}/$L${R_TOTAL})"

# --- 下部：価額・調達額・資本金 ---
r = R_TOTAL + 2
foot = [
    ("1株あたり発行価額（円）", {"E": "=前提!$B$12", "K": "=前提!$B$26"}, JPY),
    ("調達額（円）",            {"E": f"=F{R_TOTAL}", "K": "=K%d*前提!$B$26" % R_TOTAL}, JPY),
    ("Pre-money（円）",         {"E": f"=C{R_TOTAL}*前提!$B$12", "K": f"=I{R_TOTAL}*前提!$B$26"}, JPY),
    ("Post-money（円）",        {"E": f"=G{R_TOTAL}*前提!$B$12", "K": f"=L{R_TOTAL}*前提!$B$26"}, JPY),
]
for lab, cells, fmt in foot:
    S(ws2, f"B{r}", bold=True, border=True); ws2[f"B{r}"] = lab
    for col in ["C","D","E","F","G","H","I","J","K","L","M"]:
        S(ws2, f"{col}{r}", border=True)
    for col, f in cells.items():
        ws2.merge_cells(f"{col}{r}:{chr(ord(col)+1)}{r}")
        S(ws2, f"{col}{r}", bold=True, fmt=fmt, border=True, fill=TOT, align="center")
        ws2[f"{col}{r}"] = f
    r += 1
R_PRE, R_POST = r - 2, r - 1

S(ws2, f"B{r}", bold=True, border=True); ws2[f"B{r}"] = "資本金（円）"
S(ws2, f"C{r}", fmt=JPY, border=True); ws2[f"C{r}"] = "=前提!$B$7*1000"
ws2.merge_cells(f"E{r}:F{r}")
S(ws2, f"E{r}", bold=True, fmt=JPY, border=True, fill=TOT, align="center")
ws2[f"E{r}"] = f"=C{r}+F{R_TOTAL}*前提!$B$17"
r += 1
S(ws2, f"B{r}", bold=True, border=True); ws2[f"B{r}"] = "資本準備金（円）"
S(ws2, f"C{r}", fmt=JPY, border=True); ws2[f"C{r}"] = "=前提!$B$8*1000"
ws2.merge_cells(f"E{r}:F{r}")
S(ws2, f"E{r}", bold=True, fmt=JPY, border=True, fill=TOT, align="center")
ws2[f"E{r}"] = f"=C{r}+F{R_TOTAL}*(1-前提!$B$17)"
r += 2

S(ws2, f"B{r}", bold=True, size=10, color=RED)
ws2[f"B{r}"] = "▼ 旧版からの主な修正点"
notes = [
    "① 2028年以降で投資家A（134株）が「潜在株式考慮後」から消え、合計が3,000→2,866株に減っていた計算式の破損を修正",
    "② SO付与株数が0のままだったため、備考「潜在株10％以内」に沿って300株を計上（潜在込3,300株）",
    "③ 2026年12月末の列見出し「株式譲渡」を実態（第三者割当増資）に修正。払込期日は合意書に合わせ2026年10月末に統一",
    "④ 2027年の@50,000円（前回@150,000円に対するダウンラウンド）を削除。2029年は価額未確定のため空欄の想定枠に",
    "⑤ 備考欄のWineBank本体の記載（ワインファンド、375株中野移転、有限会社中村→WineBank商号変更）を削除",
    "⑥ 社名表記「Thieery」→「Thierry」に修正",
]
for i, n in enumerate(notes):
    S(ws2, f"B{r+1+i}", size=9, color="7F7F7F"); ws2[f"B{r+1+i}"] = n

# ============================================================
# 3. 合意書整合チェック
# ============================================================
ws3 = wb.create_sheet("合意書整合チェック")
ws3.sheet_view.showGridLines = False
ws3.column_dimensions["A"].width = 34
ws3.column_dimensions["B"].width = 18
ws3.column_dimensions["C"].width = 18
ws3.column_dimensions["D"].width = 14
ws3.column_dimensions["E"].width = 10
ws3.column_dimensions["F"].width = 4
ws3.column_dimensions["G"].width = 62

band(ws3, 1, "合意書との整合チェック", "G")
S(ws3, "A2", size=9, color="7F7F7F")
ws3["A2"] = "合意書（2026年9月）記載値と本資本政策の突合。判定は自動計算。"
for c, t in [("A","チェック項目"),("B","合意書の記載"),("C","本資本政策"),("D","差"),("E","判定"),("G","コメント")]:
    S(ws3, f"{c}3", bold=True, border=True, fill=TOT, align="center"); ws3[f"{c}3"] = t

# (項目, 合意書値, 本表値, 書式, 判定種別, コメント)  判定種別: "eq"=一致, "min"=下限クリア
CHK = [
    ("投資家Aの出資額", "=前提!$B$31", f"=資本政策!F{R_INVA}", JPY, "eq",
     "@150,000円では20,000,000円が割り切れない（133.33株）。133株なら19,950,000円、134株なら20,100,000円。合意書の金額か株数の調整が必要"),
    ("本件出資総額", "=前提!$B$33", f"=資本政策!F{R_TOTAL}", JPY, "eq",
     "133/267/400/200株の構成で総額はちょうど150,000,000円"),
    ("Pre-money", "=前提!$B$32", f"=資本政策!E{R_PRE}", JPY, "eq", "一致"),
    ("甲側持株比率（発行済ベース）", "=前提!$B$29", f"=資本政策!H{R_SUB_KOU}", PCT2, "min",
     "合意書 第2条3項3号。現行条文は「WBP3の持株比率」だがWBP3単独は23.3％。「中野邦人及びWBP3の合計」への修正が必要"),
    ("甲側持株比率（SO発行後・潜在込）", "=前提!$B$29", f"=資本政策!J{R_SUB_KOU}", PCT2, "min",
     "SO300株を発行すると60.6％となり66％を下回る。合意書に「発行済株式ベース」と明記しない限りSO発行が確約違反になりうる"),
    ("外部投資家 合計比率", "=前提!$B$30", f"=資本政策!H{R_SUB_EXT}", PCT2, "eq",
     "合意書の「本件34％投資家」は実際には33.3％。単独34％の投資家は存在せず（最大は投資家Cの13.3％）、定義規定も本文にない"),
]
r = 4
for lab, a, b, fmt, kind, cm in CHK:
    S(ws3, f"A{r}", border=True, size=9); ws3[f"A{r}"] = lab
    S(ws3, f"B{r}", color=GREEN, fmt=fmt, border=True, align="center"); ws3[f"B{r}"] = a
    S(ws3, f"C{r}", color=GREEN, fmt=fmt, border=True, align="center"); ws3[f"C{r}"] = b
    S(ws3, f"D{r}", fmt=fmt, border=True, align="center"); ws3[f"D{r}"] = f"=C{r}-B{r}"
    S(ws3, f"E{r}", bold=True, border=True, align="center")
    if kind == "min":
        ws3[f"E{r}"] = f'=IF(C{r}>=B{r}-0.0000001,"OK","要修正")'
    else:
        ws3[f"E{r}"] = f'=IF(ABS(D{r})<0.0000001,"OK","要修正")'
    S(ws3, f"G{r}", size=9, color="7F7F7F", wrap=True); ws3[f"G{r}"] = cm
    r += 1
for rr in range(4, r):
    ws3[f"E{rr}"].fill = TOT
R_CHK_LAST = r - 1

r += 1
sec(ws3, r, "■ 66％条項による追加エクイティ調達余地", "G")
r += 1
S(ws3, f"A{r}", size=9, color=RED, bold=True)
ws3[f"A{r}"] = "甲側は2,000株で固定（今回ラウンドで追加出資なし）。66％を維持できる発行済株式総数の上限から逆算。"
r += 1
cap = [
    ("甲側 保有株数（固定）", f"=資本政策!G{R_SUB_KOU}", SH, ""),
    ("66％維持時の発行済株式総数 上限", f"=資本政策!G{R_SUB_KOU}/前提!$B$29", SH, "＝甲側株数 ÷ 66％"),
    ("第三者割当①後の発行済株式総数", f"=資本政策!G{R_TOTAL}", SH, ""),
    ("追加発行できる株数", None, SH, "上限 － 現在。マイナスなら既に抵触"),
    ("@150,000円での追加調達可能額（円）", None, JPY, "追加発行株数 × @150,000円"),
]
base = r
for i, (lab, f, fmt, note) in enumerate(cap):
    rr = base + i
    S(ws3, f"A{rr}", border=True, size=9, bold=(i >= 3)); ws3[f"A{rr}"] = lab
    S(ws3, f"B{rr}", bold=(i >= 3), fmt=fmt, border=True, align="center",
      fill=NG if i >= 3 else None)
    if i == 3:
        ws3[f"B{rr}"] = f"=B{base+1}-B{base+2}"
    elif i == 4:
        ws3[f"B{rr}"] = f"=B{base+3}*前提!$B$12"
    else:
        ws3[f"B{rr}"] = f
    S(ws3, f"G{rr}", size=9, color="7F7F7F"); ws3[f"G{rr}"] = note
r = base + len(cap) + 1
S(ws3, f"A{r}", bold=True, size=10, color=RED)
ws3[f"A{r}"] = "→ 事業収支資料p11「調達金額2.0-3.0億円」は、合意書の66％条項の下では達成不能。"
S(ws3, f"A{r+1}", size=9, color="7F7F7F")
ws3[f"A{r+1}"] = "　 2億円を30株で調達するには@6,666,667円/株（現行の44倍）が必要。甲側の同時出資、66％の緩和、"
S(ws3, f"A{r+2}", size=9, color="7F7F7F")
ws3[f"A{r+2}"] = "　 または「本件ラウンド時点の比率」への限定のいずれかが必須。種類株式の発行禁止（第2条3項4号）も併せて要見直し。"

# ============================================================
# 4. 資金需要
# ============================================================
ws4 = wb.create_sheet("資金需要")
ws4.sheet_view.showGridLines = False
ws4.column_dimensions["A"].width = 34
for c in ["B","C","D","E","F","G"]: ws4.column_dimensions[c].width = 13
ws4.column_dimensions["H"].width = 4
ws4.column_dimensions["I"].width = 56

band(ws4, 1, "資金需要と調達の突合（単位：百万円）", "G")
S(ws4, "A2", size=9, color="7F7F7F")
ws4["A2"] = "出典：5か年事業計画FY2027-2031（六本木1号店＋9店舗連結）。青字＝同計画からの転記値"
YRS = ["FY2027","FY2028","FY2029","FY2030","FY2031"]
YC = ["B","C","D","E","F"]
S(ws4, "A3", bold=True, border=True, fill=TOT); ws4["A3"] = "項　目"
for c, t in zip(YC, YRS):
    S(ws4, f"{c}3", bold=True, border=True, fill=TOT, align="center"); ws4[f"{c}3"] = t
S(ws4, "G3", bold=True, border=True, fill=TOT, align="center"); ws4["G3"] = "5年累計"

DATA = [
    ("連結EBITDA",            [37.4, 103.5, 233.7, 377.9, 523.2], BLUE),
    ("投資キャッシュアウト",   [-123.7, -251.3, -272.5, -303.7, -334.9], BLUE),
    ("営業利益",              [16.6, 56.7, 145.9, 249.6, 354.2], BLUE),
]
r = 4
rowmap = {}
for lab, vals, col in DATA:
    S(ws4, f"A{r}", border=True); ws4[f"A{r}"] = lab
    for c, v in zip(YC, vals):
        S(ws4, f"{c}{r}", color=col, fmt=MM, border=True); ws4[f"{c}{r}"] = v
    S(ws4, f"G{r}", bold=True, fmt=MM, border=True, fill=TOT); ws4[f"G{r}"] = f"=SUM(B{r}:F{r})"
    rowmap[lab] = r
    r += 1

S(ws4, f"A{r}", border=True); ws4[f"A{r}"] = "（前提）法人税 実効税率　※B列セルが入力値"
S(ws4, "B%d" % (r), color=BLUE, fmt=PCT, border=True, align="center")
ws4[f"B{r}"] = 0.30
S(ws4, f"I{r}", size=9, color="7F7F7F"); ws4[f"I{r}"] = "簡易。繰越欠損金の控除は考慮せず"
R_TAXRATE = r
r += 1
S(ws4, f"A{r}", border=True); ws4[f"A{r}"] = "法人税等（推計）"
for c in YC:
    S(ws4, f"{c}{r}", fmt=MM, border=True)
    ws4[f"{c}{r}"] = f"=-MAX(0,{c}{rowmap['営業利益']})*$B${R_TAXRATE}"
S(ws4, f"G{r}", bold=True, fmt=MM, border=True, fill=TOT); ws4[f"G{r}"] = f"=SUM(B{r}:F{r})"
R_TAX = r
r += 1
S(ws4, f"A{r}", bold=True, border=True); ws4[f"A{r}"] = "税引後フリーキャッシュフロー"
for c in YC:
    S(ws4, f"{c}{r}", bold=True, fmt=MM, border=True, fill=TOT)
    ws4[f"{c}{r}"] = f"={c}{rowmap['連結EBITDA']}+{c}{rowmap['投資キャッシュアウト']}+{c}{R_TAX}"
S(ws4, f"G{r}", bold=True, fmt=MM, border=True, fill=TOT); ws4[f"G{r}"] = f"=SUM(B{r}:F{r})"
R_FCF = r
r += 1
S(ws4, f"A{r}", bold=True, border=True); ws4[f"A{r}"] = "累計フリーキャッシュフロー"
for i, c in enumerate(YC):
    S(ws4, f"{c}{r}", bold=True, fmt=MM, border=True, fill=KEY)
    ws4[f"{c}{r}"] = f"={c}{R_FCF}" if i == 0 else f"={YC[i-1]}{r}+{c}{R_FCF}"
R_CUM = r
r += 2

sec(ws4, r, "■ 必要調達額との比較", "G")
r += 1
comp = [
    ("最大資金需要（累計FCFの最小値）", f"=-MIN(B{R_CUM}:F{R_CUM})", MM, "税引後ベース。FY2029末が底"),
    ("既存調達額（エクイティ＋デット）", 200.0, MM, "事業収支資料p11「2億調達済み」。内訳未確認のため要確認"),
    ("第三者割当①（今回）", None, MM, "＝資本政策の本件出資総額"),
    ("調達額　計", None, MM, ""),
    ("差（プラスがバッファ）", None, MM, "デットの元本返済・分割弁済に係る金利相当は未考慮"),
]
base = r
for i, (lab, f, fmt, note) in enumerate(comp):
    rr = base + i
    S(ws4, f"A{rr}", border=True, bold=(i >= 3)); ws4[f"A{rr}"] = lab
    S(ws4, f"B{rr}", bold=(i >= 3), color=BLUE if i == 1 else BLACK, fmt=fmt, border=True,
      align="center", fill=NG if i == 4 else (TOT if i == 3 else None))
    if i == 2:
        ws4[f"B{rr}"] = "=資本政策!F%d/1000000" % R_TOTAL
    elif i == 3:
        ws4[f"B{rr}"] = f"=B{base+1}+B{base+2}"
    elif i == 4:
        ws4[f"B{rr}"] = f"=B{base+3}-B{base}"
    else:
        ws4[f"B{rr}"] = f
    S(ws4, f"I{rr}", size=9, color="7F7F7F"); ws4[f"I{rr}"] = note
r = base + len(comp) + 1
S(ws4, f"A{r}", bold=True, size=10, color=RED)
ws4[f"A{r}"] = "→ バッファは1億円未満。既存2億円にデットが含まれるため、その元本返済を入れると不足に転じる可能性がある。"
S(ws4, f"A{r+1}", size=9, color="7F7F7F")
ws4[f"A{r+1}"] = "※ 対象会社は六本木1号店の運営会社。上記は9店舗連結の計画値。2号店以降を対象会社で持つのか別会社にするのかを"
S(ws4, f"A{r+2}", size=9, color="7F7F7F")
ws4[f"A{r+2}"] = "　 明確にしないと、投資家が見るバリュエーション（Post4.5億）と出資対象の事業範囲が一致しない。"

if "Sheet" in wb.sheetnames:
    del wb["Sheet"]
SHEETS = ["前提", "資本政策", "合意書整合チェック", "資金需要"]
for sh in wb.worksheets:
    for row in sh.iter_rows():
        for c in row:
            if isinstance(c.value, str) and c.value.startswith("="):
                f = c.value
                for nm in SHEETS:
                    f = f.replace(f"'{nm}'!", f"{nm}!")
                for nm in SHEETS:
                    f = f.replace(f"{nm}!", f"'{nm}'!")
                c.value = f
wb.calculation.fullCalcOnLoad = True
wb.save(OUT)
print("saved:", OUT)
