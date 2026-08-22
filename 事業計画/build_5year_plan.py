# -*- coding: utf-8 -*-
"""ティエリーマルクス・ブラッスリー 5か年事業計画 (FY2027-FY2031)
   出典: 「THIERRY MARX × WineBank_事業収支CF追加_20260804.pptx」(2026年8月3日提出P&L準拠)"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

OUT = "/home/user/wine-auction/事業計画/ティエリーマルクス・ブラッスリー_5か年事業計画_FY2027-2031.xlsx"

JP = "Meiryo"
BLUE, BLACK, GREEN, RED = "0000FF", "000000", "008000", "C00000"
HDR = PatternFill("solid", fgColor="1F3864")
SUB = PatternFill("solid", fgColor="D9E2F3")
YEL = PatternFill("solid", fgColor="FFF2CC")
KEY = PatternFill("solid", fgColor="FFFF00")
TOT = PatternFill("solid", fgColor="F2F2F2")
thin = Side(style="thin", color="BFBFBF")
BOX = Border(left=thin, right=thin, top=thin, bottom=thin)

MM = '#,##0.0;(#,##0.0);-'
MM0 = '#,##0;(#,##0);-'
PCT = '0.0%;(0.0%);-'

YEARS = ["FY2027", "FY2028", "FY2029", "FY2030", "FY2031"]
YC = ["C", "D", "E", "F", "G"]      # 年度列（連結・コホート・CF）
PC = ["B", "C", "D", "E", "F"]      # 前提条件シートの年度列

# ---- 変動費（対売上比）: 1年目(提出P&L) / 2年目 / 3年目以降 -------------
VAR = [
    ("売上原価",                 0.295, 0.290, 0.285, "直輸入ワイン活用・複数店一括仕入（▲1.0pt）"),
    ("人件費",                   0.345, 0.335, 0.325, "本部集中化・多能工化。26名体制の最適化（▲2.0pt）"),
    ("業務委託費・消耗品他",     0.024, 0.022, 0.020, "本部集中購買・規格統一（▲0.4pt）"),
    ("広告宣伝費・販売促進費",   0.014, 0.014, 0.014, "1号店の話題性を活用（据置）"),
    ("その他経費",               0.005, 0.005, 0.005, "据置"),
    ("水道光熱費",               0.040, 0.038, 0.036, "高効率厨房機器・営業時間最適化（▲0.4pt）"),
    ("ロイヤリティ",             0.040, 0.040, 0.040, "契約条件（ブラッスリー4.0%）"),
    ("保険・租税公課",           0.007, 0.007, 0.007, "据置"),
    ("修繕維持費（FFE）",        0.025, 0.025, 0.025, "据置"),
]
V_TOP = 22                      # 前提条件シートの変動費テーブル先頭行
V_BTM = V_TOP + len(VAR) - 1    # 30

# P&L 表示順（Trueが変動費、Falseが特殊行）
PL_ORDER = [
    ("売上原価", "var", 0),
    ("人件費", "var", 1),
    ("業務委託費・消耗品他", "var", 2),
    ("部門利益", "sub_dept", None),
    ("広告宣伝費・販売促進費", "var", 3),
    ("その他経費", "var", 4),
    ("水道光熱費", "var", 5),
    ("GOP（店舗営業総利益）", "sub_gop", None),
    ("ロイヤリティ", "var", 6),
    ("保険・租税公課", "var", 7),
    ("地代家賃（固定）", "rent_fix", None),
    ("歩合家賃", "rent_pct", None),
    ("修繕維持費（FFE）", "var", 8),
]

wb = openpyxl.Workbook()

def S(ws, cell, *, bold=False, color=BLACK, fmt=None, fill=None, align=None, size=10, border=False, wrap=False):
    c = ws[cell]
    c.font = Font(name=JP, bold=bold, color=color, size=size)
    if fmt: c.number_format = fmt
    if fill: c.fill = fill
    if align or wrap: c.alignment = Alignment(horizontal=align, vertical="center", wrap_text=wrap)
    if border: c.border = BOX
    return c

def band(ws, row, text, last="H", fill=HDR, color="FFFFFF", size=11):
    ws.cell(row=row, column=1, value=text)
    for col in range(1, openpyxl.utils.column_index_from_string(last) + 1):
        c = ws.cell(row=row, column=col)
        c.fill = fill
        c.font = Font(name=JP, bold=True, color=color, size=size)

def sec(ws, row, text, last="H"):
    band(ws, row, text, last, fill=SUB, color="000000", size=10)

def hdr_years(ws, row, cols, labels, first_label="項　目", extra=None):
    S(ws, f"A{row}", bold=True, border=True, fill=TOT); ws[f"A{row}"] = first_label
    for cl, t in zip(cols, labels):
        S(ws, f"{cl}{row}", bold=True, border=True, fill=TOT, align="center"); ws[f"{cl}{row}"] = t
    if extra:
        cl, t = extra
        S(ws, f"{cl}{row}", bold=True, border=True, fill=TOT, align="center"); ws[f"{cl}{row}"] = t

# ============================================================
# 前提条件
# ============================================================
ws = wb.create_sheet("前提条件")
ws.sheet_view.showGridLines = False
ws.column_dimensions["A"].width = 32
for c in "BCDEF": ws.column_dimensions[c].width = 13
ws.column_dimensions["G"].width = 4
ws.column_dimensions["H"].width = 52

band(ws, 1, "前提条件　／　六本木ミッドタウン店 2026年8月3日提出P&L 準拠")
S(ws, "A2", size=9, color="7F7F7F")
ws["A2"] = "青字＝入力値（前提）。変更すると全シートが自動再計算されます。黒字＝計算式、緑字＝他シート参照。単位：百万円"

sec(ws, 4, "■ 売上前提（六本木ミッドタウン店＝標準店モデル）")
base = [
    ("満年度 売上高", 460.0, MM, "183.59㎡(55.5坪)・123席・年363日営業・稼働率80%。日商約127万円／月商約3,829万円"),
    ("立上り率　1年目", 0.90, PCT, "新店初年度は保守的に90%で計画"),
    ("立上り率　2年目", 0.98, PCT, ""),
    ("立上り率　3年目以降", 1.00, PCT, "満年度水準"),
    ("新店の開業年度 稼働月数", 7.5, MM, "年2店を上期・下期に分散出店する前提の平均値"),
    ("東京1号店の稼働月数（FY2027）", 12.0, MM, "2027年4月開業＝FY2027は通年稼働"),
]
r = 5
for lab, val, fmt, note in base:
    S(ws, f"A{r}"); ws[f"A{r}"] = lab
    S(ws, f"B{r}", color=BLUE, fmt=fmt, border=True, align="center"); ws[f"B{r}"] = val
    S(ws, f"H{r}", size=9, color="7F7F7F"); ws[f"H{r}"] = note
    r += 1
# B5=満年度売上 B6/B7/B8=立上り率 B9=新店稼働月数 B10=東京稼働月数

sec(ws, 12, "■ 家賃条件（契約条件どおり全店適用）")
rent = [
    ("地代家賃（固定）月額", 2.22, MM, "月222万円。売上非連動の固定費"),
    ("地代家賃（固定）年額", None, MM, "＝月額×12"),
    ("歩合家賃率", 0.07, PCT, "閾値超過分に対して7%"),
    ("歩合家賃 閾値（年額売上）", 331.4, MM, "満年度歩合家賃9百万円（対売上1.9%）から逆算"),
]
r = 13
for lab, val, fmt, note in rent:
    S(ws, f"A{r}"); ws[f"A{r}"] = lab
    S(ws, f"B{r}", color=BLUE if val is not None else BLACK, fmt=fmt, border=True, align="center")
    ws[f"B{r}"] = val if val is not None else "=B13*12"
    S(ws, f"H{r}", size=9, color="7F7F7F"); ws[f"H{r}"] = note
    r += 1
# B13 月額 B14 年額 B15 歩合率 B16 閾値

sec(ws, 18, "■ 変動費率（対売上高）　※提出P&L水準 → 改善後")
for cl, t in zip(["A", "B", "C", "D", "H"],
                 ["費　目", "1年目\n（提出P&L）", "2年目\n（改善途上）", "3年目以降\n（改善後）", "改善レバー"]):
    c = ws[f"{cl}{V_TOP-1}"]; c.value = t
    c.font = Font(name=JP, bold=True, size=9)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.fill = TOT; c.border = BOX
ws.row_dimensions[V_TOP-1].height = 32
r = V_TOP
for lab, a, b, c_, note in VAR:
    S(ws, f"A{r}", border=True); ws[f"A{r}"] = lab
    for cl, v in zip(["B", "C", "D"], [a, b, c_]):
        S(ws, f"{cl}{r}", color=BLUE, fmt=PCT, border=True, align="center"); ws[f"{cl}{r}"] = v
    S(ws, f"H{r}", size=9, color="7F7F7F"); ws[f"H{r}"] = note
    r += 1
S(ws, f"A{V_BTM+1}", bold=True, border=True); ws[f"A{V_BTM+1}"] = "変動費　計（対売上）"
for cl in "BCD":
    S(ws, f"{cl}{V_BTM+1}", bold=True, fmt=PCT, border=True, align="center", fill=TOT)
    ws[f"{cl}{V_BTM+1}"] = f"=SUM({cl}{V_TOP}:{cl}{V_BTM})"

sec(ws, 33, "■ 初期投資（1店舗あたり）　※六本木ミッドタウン実額準拠")
S(ws, "B34", bold=True, border=True, fill=TOT, align="center", size=9, wrap=True)
ws["B34"] = "1〜3号店\n(東京・札幌・大阪)"
S(ws, "C34", bold=True, border=True, fill=TOT, align="center", size=9, wrap=True)
ws["C34"] = "4号店以降"
S(ws, "A34", bold=True, border=True, fill=TOT); ws["A34"] = "項　目"
ws.row_dimensions[34].height = 28
inv = [
    ("契約金", 10.0, 5.0, "契約金 3店舗3,000万円／以下7店舗3,500万円"),
    ("事業費（設計・内装施工・OSE/FF&E）", 194.0, 194.0, ""),
    ("開業準備金（広告・採用等）", 4.0, 4.0, ""),
    ("保証金（賃料70か月相当）", 39.0, 39.0, "非償却・退去時返還対象"),
]
r = 35
for lab, a, b, note in inv:
    S(ws, f"A{r}", border=True); ws[f"A{r}"] = lab
    S(ws, f"B{r}", color=BLUE, fmt=MM, border=True, align="center"); ws[f"B{r}"] = a
    S(ws, f"C{r}", color=BLUE, fmt=MM, border=True, align="center"); ws[f"C{r}"] = b
    S(ws, f"H{r}", size=9, color="7F7F7F"); ws[f"H{r}"] = note
    r += 1
S(ws, "A39", bold=True, border=True); ws["A39"] = "初期投資　合計"
S(ws, "A40", bold=True, border=True); ws["A40"] = "償却対象資産（＝合計－保証金）"
S(ws, "A41", border=True); ws["A41"] = "平均償却年数（年）"
S(ws, "A42", bold=True, border=True); ws["A42"] = "年間減価償却費／1店（通年）"
for cl in "BC":
    S(ws, f"{cl}39", bold=True, fmt=MM, border=True, align="center", fill=KEY); ws[f"{cl}39"] = f"=SUM({cl}35:{cl}38)"
    S(ws, f"{cl}40", bold=True, fmt=MM, border=True, align="center", fill=TOT); ws[f"{cl}40"] = f"={cl}39-{cl}38"
    S(ws, f"{cl}41", color=BLUE, fmt=MM, border=True, align="center"); ws[f"{cl}41"] = 10.0
    S(ws, f"{cl}42", bold=True, fmt=MM, border=True, align="center", fill=TOT); ws[f"{cl}42"] = f"={cl}40/{cl}41"
S(ws, "H39", size=9, color=RED, bold=True); ws["H39"] = "1〜3号店 247百万円／4号店以降 242百万円"

sec(ws, 44, "■ 分割払いスキーム（ミッドタウン契約と同条件・全店適用）")
sp = [
    ("分割対象：内装工事費", 100.0, "月833千円 × 120回（10年）"),
    ("分割対象：保証金", 38.9, "月463千円 × 84回（7年）"),
    ("分割対象額　計／1店", None, "開業時現金支出＝初期投資－分割対象額"),
    ("年間分割弁済額／1店", 15.6, "月額合計1,296千円 × 12か月"),
]
r = 45
for lab, val, note in sp:
    S(ws, f"A{r}", bold=(val is None)); ws[f"A{r}"] = lab
    S(ws, f"B{r}", color=BLUE if val is not None else BLACK, bold=(val is None), fmt=MM,
      border=True, align="center", fill=TOT if val is None else None)
    ws[f"B{r}"] = val if val is not None else "=B45+B46"
    S(ws, f"H{r}", size=9, color="7F7F7F"); ws[f"H{r}"] = note
    r += 1
S(ws, "A50", size=9, color="7F7F7F")
ws["A50"] = "※ 分割弁済は資産計上済みの繰延支払（損益は減価償却で費用化済み）のため、P/Lには影響させずCF上のみ計上。"
# B45 内装 B46 保証金 B47 分割対象計 B48 年間分割弁済

sec(ws, 52, "■ 出店計画・本部費（連結）")
hdr_years(ws, 53, PC, YEARS)
plan = [
    ("期中出店数（店）", [1, 2, 2, 2, 2], MM0, BLUE),
    ("期末店舗数（店）", None, MM0, BLACK),
    ("本部費（管理・人事・購買・マーケ）", [15.0, 28.0, 40.0, 48.0, 55.0], MM, BLUE),
]
r = 54
for lab, vals, fmt, color in plan:
    S(ws, f"A{r}", bold=(vals is None)); ws[f"A{r}"] = lab
    for i, cl in enumerate(PC):
        S(ws, f"{cl}{r}", color=color, fmt=fmt, border=True, align="center", bold=(vals is None))
        ws[f"{cl}{r}"] = vals[i] if vals else (f"={cl}54" if i == 0 else f"={PC[i-1]}55+{cl}54")
    r += 1
S(ws, "H54", size=9, color="7F7F7F")
ws["H54"] = "FY2027 六本木ミッドタウン／FY2028 札幌・大阪／FY2029 首都圏・他政令指定都市／FY2030以降 リゾート含め年2店"
S(ws, "H56", size=9, color="7F7F7F")
ws["H56"] = "多店舗化に伴う本部組織の段階的増強費用"
HQ_ROW, STORES_ROW, OPEN_ROW = 56, 55, 54

# ============================================================
# 単店モデルPL
# ============================================================
ws2 = wb.create_sheet("単店モデルPL")
ws2.sheet_view.showGridLines = False
ws2.column_dimensions["A"].width = 32
for c in ["B", "C", "D", "E"]: ws2.column_dimensions[c].width = 15
ws2.column_dimensions["F"].width = 4
ws2.column_dimensions["G"].width = 40

band(ws2, 1, "単店モデルPL（六本木ミッドタウン店・百万円）", "E")
S(ws2, "A2", size=9, color="7F7F7F")
ws2["A2"] = "満年度＝2026年8月3日提出P&Lどおり。1年目/2年目/3年目以降は立上り率と経費改善を反映。"

hdr_years(ws2, 4, ["B", "C", "D", "E"],
          ["満年度\n（提出P&L）", "1年目", "2年目", "3年目以降"], "科　目")
for cl in ["B", "C", "D", "E"]:
    ws2[f"{cl}4"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
ws2.row_dimensions[4].height = 30

# 列 → (立上り行, 変動費率列)
COLMAP = {"B": ("$B$8", "B"), "C": ("$B$6", "B"), "D": ("$B$7", "C"), "E": ("$B$8", "D")}

REV_R = 5
S(ws2, f"A{REV_R}", bold=True, border=True); ws2[f"A{REV_R}"] = "売上高"
for cl, (ramp, _) in COLMAP.items():
    S(ws2, f"{cl}{REV_R}", bold=True, fmt=MM, border=True)
    ws2[f"{cl}{REV_R}"] = f"=前提条件!$B$5*前提条件!{ramp}"

rows2 = {}
r = REV_R + 1
for lab, kind, idx in PL_ORDER:
    S(ws2, f"A{r}", bold=kind.startswith("sub"), border=True); ws2[f"A{r}"] = lab
    for cl, (ramp, rcol) in COLMAP.items():
        if kind == "var":
            S(ws2, f"{cl}{r}", fmt=MM, border=True)
            ws2[f"{cl}{r}"] = f"={cl}${REV_R}*前提条件!${rcol}${V_TOP+idx}"
        elif kind == "sub_dept":
            S(ws2, f"{cl}{r}", bold=True, fmt=MM, border=True, fill=TOT)
            ws2[f"{cl}{r}"] = f"={cl}{REV_R}-SUM({cl}{REV_R+1}:{cl}{r-1})"
        elif kind == "sub_gop":
            S(ws2, f"{cl}{r}", bold=True, fmt=MM, border=True, fill=TOT)
            ws2[f"{cl}{r}"] = f"={cl}{rows2['部門利益']}-SUM({cl}{rows2['部門利益']+1}:{cl}{r-1})"
        elif kind == "rent_fix":
            S(ws2, f"{cl}{r}", fmt=MM, border=True)
            ws2[f"{cl}{r}"] = "=前提条件!$B$14"
        elif kind == "rent_pct":
            S(ws2, f"{cl}{r}", fmt=MM, border=True)
            ws2[f"{cl}{r}"] = f"=MAX(0,{cl}{REV_R}-前提条件!$B$16)*前提条件!$B$15"
    rows2[lab] = r
    r += 1

EB2, DEP2, OP2 = r, r + 1, r + 2
S(ws2, f"A{EB2}", bold=True, border=True); ws2[f"A{EB2}"] = "EBITDA"
S(ws2, f"A{DEP2}", border=True); ws2[f"A{DEP2}"] = "減価償却費"
S(ws2, f"A{OP2}", bold=True, border=True); ws2[f"A{OP2}"] = "営業利益"
S(ws2, f"A{OP2+1}", bold=True, border=True); ws2[f"A{OP2+1}"] = "　EBITDA率"
S(ws2, f"A{OP2+2}", bold=True, border=True); ws2[f"A{OP2+2}"] = "　営業利益率"
gop_r = rows2["GOP（店舗営業総利益）"]
for cl in ["B", "C", "D", "E"]:
    S(ws2, f"{cl}{EB2}", bold=True, fmt=MM, border=True, fill=KEY)
    ws2[f"{cl}{EB2}"] = f"={cl}{gop_r}-SUM({cl}{gop_r+1}:{cl}{EB2-1})"
    S(ws2, f"{cl}{DEP2}", color=GREEN, fmt=MM, border=True); ws2[f"{cl}{DEP2}"] = "=前提条件!$B$42"
    S(ws2, f"{cl}{OP2}", bold=True, fmt=MM, border=True, fill=KEY)
    ws2[f"{cl}{OP2}"] = f"={cl}{EB2}-{cl}{DEP2}"
    S(ws2, f"{cl}{OP2+1}", bold=True, fmt=PCT, border=True, align="center", fill=KEY)
    ws2[f"{cl}{OP2+1}"] = f"={cl}{EB2}/{cl}{REV_R}"
    S(ws2, f"{cl}{OP2+2}", bold=True, fmt=PCT, border=True, align="center", fill=KEY)
    ws2[f"{cl}{OP2+2}"] = f"={cl}{OP2}/{cl}{REV_R}"
S(ws2, f"G{EB2}", size=9, color=RED, bold=True)
ws2[f"G{EB2}"] = "成熟店：EBITDA率16.5%／営業利益率12.1%"
S(ws2, f"A{OP2+4}", size=9, color="7F7F7F")
ws2[f"A{OP2+4}"] = "※ 減価償却費は1〜3号店ベース（償却対象208百万円÷10年）。金利・税金は含まない。"

# ============================================================
# 店舗群別PL
# ============================================================
ws3 = wb.create_sheet("店舗群別PL")
ws3.sheet_view.showGridLines = False
ws3.column_dimensions["A"].width = 30
ws3.column_dimensions["B"].width = 11
for c in YC: ws3.column_dimensions[c].width = 13

band(ws3, 1, "店舗群（出店年度コホート）別PL（百万円）")
S(ws3, "A2", size=9, color="7F7F7F")
ws3["A2"] = "年次（1/2/3）に応じて「前提条件」の立上り率・変動費率を自動参照。地代家賃と歩合家賃の閾値は稼働月数で按分。"
hdr_years(ws3, 3, YC, YEARS, "店舗群", ("B", "店舗数"))

COHORTS = [
    ("① 東京（六本木ミッドタウン）", 1, "$B$10", "B", [1, 2, 3, 3, 3]),
    ("② 札幌・大阪",                 2, "$B$9",  "B", [0, 1, 2, 3, 3]),
    ("③ 首都圏・他政令指定都市",     2, "$B$9",  "C", [0, 0, 1, 2, 3]),
    ("④ FY2030出店（リゾート含む）", 2, "$B$9",  "C", [0, 0, 0, 1, 2]),
    ("⑤ FY2031出店（リゾート含む）", 2, "$B$9",  "C", [0, 0, 0, 0, 1]),
]

blocks = []
r = 5
for name, n, mref, invcol, yidx in COHORTS:
    sec(ws3, r, f"　{name}", "G")
    r_n, r_m, r_y, r_rev = r + 1, r + 2, r + 3, r + 4
    S(ws3, f"A{r_n}", size=9); ws3[f"A{r_n}"] = "店舗数（店）"
    S(ws3, f"B{r_n}", color=BLUE, fmt=MM0, border=True, align="center"); ws3[f"B{r_n}"] = n
    S(ws3, f"A{r_m}", size=9); ws3[f"A{r_m}"] = "開業年度の稼働月数"
    S(ws3, f"B{r_m}", color=GREEN, fmt=MM, border=True, align="center"); ws3[f"B{r_m}"] = f"=前提条件!{mref}"
    S(ws3, f"A{r_y}", size=9); ws3[f"A{r_y}"] = "年次（0=未開業/1/2/3）"
    for i, cl in enumerate(YC):
        S(ws3, f"{cl}{r_y}", color=BLUE, fmt=MM0, border=True, align="center"); ws3[f"{cl}{r_y}"] = yidx[i]
    # 稼働月数按分係数
    r_f = r_y  # 年次行を参照
    S(ws3, f"A{r_rev}", bold=True, border=True); ws3[f"A{r_rev}"] = "売上高"
    for cl in YC:
        S(ws3, f"{cl}{r_rev}", bold=True, fmt=MM, border=True)
        ws3[f"{cl}{r_rev}"] = (
            f"=IF({cl}{r_y}=0,0,前提条件!$B$5*INDEX(前提条件!$B$6:$B$8,{cl}{r_y})"
            f"*$B${r_n}*IF({cl}{r_y}=1,$B${r_m}/12,1))")
    rr = r_rev + 1
    rowmap = {}
    for lab, kind, idx in PL_ORDER:
        S(ws3, f"A{rr}", size=9, bold=kind.startswith("sub"), border=True); ws3[f"A{rr}"] = lab
        for cl in YC:
            if kind == "var":
                S(ws3, f"{cl}{rr}", fmt=MM, border=True)
                ws3[f"{cl}{rr}"] = (f"=IF({cl}${r_y}=0,0,{cl}${r_rev}*"
                                    f"INDEX(前提条件!$B${V_TOP+idx}:$D${V_TOP+idx},1,{cl}${r_y}))")
            elif kind == "sub_dept":
                S(ws3, f"{cl}{rr}", bold=True, fmt=MM, border=True, fill=TOT)
                ws3[f"{cl}{rr}"] = f"={cl}{r_rev}-SUM({cl}{r_rev+1}:{cl}{rr-1})"
            elif kind == "sub_gop":
                S(ws3, f"{cl}{rr}", bold=True, fmt=MM, border=True, fill=TOT)
                dr = rowmap["部門利益"]
                ws3[f"{cl}{rr}"] = f"={cl}{dr}-SUM({cl}{dr+1}:{cl}{rr-1})"
            elif kind == "rent_fix":
                S(ws3, f"{cl}{rr}", fmt=MM, border=True)
                ws3[f"{cl}{rr}"] = (f"=IF({cl}{r_y}=0,0,前提条件!$B$14*$B${r_n}"
                                    f"*IF({cl}{r_y}=1,$B${r_m}/12,1))")
            elif kind == "rent_pct":
                S(ws3, f"{cl}{rr}", fmt=MM, border=True)
                ws3[f"{cl}{rr}"] = (f"=IF({cl}{r_y}=0,0,MAX(0,{cl}{r_rev}/$B${r_n}-前提条件!$B$16"
                                    f"*IF({cl}{r_y}=1,$B${r_m}/12,1))*前提条件!$B$15*$B${r_n})")
        rowmap[lab] = rr
        rr += 1
    r_eb, r_dep, r_op = rr, rr + 1, rr + 2
    S(ws3, f"A{r_eb}", bold=True, border=True); ws3[f"A{r_eb}"] = "店舗EBITDA"
    S(ws3, f"A{r_dep}", border=True); ws3[f"A{r_dep}"] = "減価償却費"
    S(ws3, f"A{r_op}", bold=True, border=True); ws3[f"A{r_op}"] = "店舗営業利益"
    gr = rowmap["GOP（店舗営業総利益）"]
    for cl in YC:
        S(ws3, f"{cl}{r_eb}", bold=True, fmt=MM, border=True, fill=TOT)
        ws3[f"{cl}{r_eb}"] = f"={cl}{gr}-SUM({cl}{gr+1}:{cl}{r_eb-1})"
        S(ws3, f"{cl}{r_dep}", fmt=MM, border=True)
        ws3[f"{cl}{r_dep}"] = (f"=IF({cl}{r_y}=0,0,前提条件!${invcol}$42*$B${r_n}"
                               f"*IF({cl}{r_y}=1,$B${r_m}/12,1))")
        S(ws3, f"{cl}{r_op}", bold=True, fmt=MM, border=True, fill=TOT)
        ws3[f"{cl}{r_op}"] = f"={cl}{r_eb}-{cl}{r_dep}"
    blocks.append(dict(rev=r_rev, rows=rowmap, eb=r_eb, dep=r_dep, op=r_op,
                       n=r_n, m=r_m, y=r_y, invcol=invcol))
    r = r_op + 2

# ============================================================
# 5カ年連結PL
# ============================================================
ws4 = wb.create_sheet("5カ年連結PL")
ws4.sheet_view.showGridLines = False
ws4.column_dimensions["A"].width = 30
ws4.column_dimensions["B"].width = 10
for c in YC: ws4.column_dimensions[c].width = 14
ws4.column_dimensions["H"].width = 14

band(ws4, 1, "5か年事業計画　連結損益計画（FY2027-FY2031・百万円）")
S(ws4, "A2", size=9, color="7F7F7F")
ws4["A2"] = "FY＝4月〜翌3月。2027年4月 六本木開業 → 2028年 札幌・大阪 → 2029年 首都圏・政令指定都市 → 2030年以降 リゾート含め年2店舗"
hdr_years(ws4, 3, YC, YEARS, "科　目", ("H", "5年累計"))
S(ws4, "B3", bold=True, border=True, fill=TOT, align="center"); ws4["B3"] = "対売上比\n(FY2031)"
ws4["B3"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

def agg(key, cl):
    return "+".join(f"店舗群別PL!{cl}{b[key]}" for b in blocks)

def agg_row(lab, cl):
    return "+".join(f"店舗群別PL!{cl}{b['rows'][lab]}" for b in blocks)

R4 = {}
row = 4
S(ws4, f"A{row}", bold=True, border=True); ws4[f"A{row}"] = "期末店舗数（店）"
for i, cl in enumerate(YC):
    S(ws4, f"{cl}{row}", bold=True, color=GREEN, fmt=MM0, border=True, align="center")
    ws4[f"{cl}{row}"] = f"=前提条件!{PC[i]}{STORES_ROW}"
R4["stores"] = row; row += 1

S(ws4, f"A{row}", bold=True, border=True); ws4[f"A{row}"] = "売上高"
for cl in YC:
    S(ws4, f"{cl}{row}", bold=True, color=GREEN, fmt=MM, border=True)
    ws4[f"{cl}{row}"] = "=" + agg("rev", cl)
R4["rev"] = row; row += 1

for lab, kind, idx in PL_ORDER:
    S(ws4, f"A{row}", size=9, bold=kind.startswith("sub"), border=True); ws4[f"A{row}"] = lab
    for cl in YC:
        S(ws4, f"{cl}{row}", bold=kind.startswith("sub"), color=GREEN, fmt=MM, border=True,
          fill=TOT if kind.startswith("sub") else None)
        ws4[f"{cl}{row}"] = "=" + agg_row(lab, cl)
    R4[lab] = row; row += 1

S(ws4, f"A{row}", bold=True, border=True); ws4[f"A{row}"] = "店舗EBITDA　計"
for cl in YC:
    S(ws4, f"{cl}{row}", bold=True, color=GREEN, fmt=MM, border=True, fill=TOT)
    ws4[f"{cl}{row}"] = "=" + agg("eb", cl)
R4["st_eb"] = row; row += 1

S(ws4, f"A{row}", border=True); ws4[f"A{row}"] = "本部費"
for i, cl in enumerate(YC):
    S(ws4, f"{cl}{row}", color=GREEN, fmt=MM, border=True)
    ws4[f"{cl}{row}"] = f"=前提条件!{PC[i]}{HQ_ROW}"
R4["hq"] = row; row += 1

S(ws4, f"A{row}", bold=True, border=True); ws4[f"A{row}"] = "連結EBITDA"
for cl in YC:
    S(ws4, f"{cl}{row}", bold=True, fmt=MM, border=True, fill=KEY)
    ws4[f"{cl}{row}"] = f"={cl}{R4['st_eb']}-{cl}{R4['hq']}"
R4["eb"] = row; row += 1

S(ws4, f"A{row}", bold=True, border=True); ws4[f"A{row}"] = "　EBITDA率"
for cl in YC:
    S(ws4, f"{cl}{row}", bold=True, fmt=PCT, border=True, align="center", fill=KEY)
    ws4[f"{cl}{row}"] = f"=IF({cl}{R4['rev']}=0,0,{cl}{R4['eb']}/{cl}{R4['rev']})"
R4["ebm"] = row; row += 1

S(ws4, f"A{row}", border=True); ws4[f"A{row}"] = "減価償却費"
for cl in YC:
    S(ws4, f"{cl}{row}", color=GREEN, fmt=MM, border=True)
    ws4[f"{cl}{row}"] = "=" + agg("dep", cl)
R4["dep"] = row; row += 1

S(ws4, f"A{row}", bold=True, border=True); ws4[f"A{row}"] = "営業利益"
for cl in YC:
    S(ws4, f"{cl}{row}", bold=True, fmt=MM, border=True, fill=KEY)
    ws4[f"{cl}{row}"] = f"={cl}{R4['eb']}-{cl}{R4['dep']}"
R4["op"] = row; row += 1

S(ws4, f"A{row}", bold=True, border=True); ws4[f"A{row}"] = "　営業利益率"
for cl in YC:
    S(ws4, f"{cl}{row}", bold=True, fmt=PCT, border=True, align="center", fill=KEY)
    ws4[f"{cl}{row}"] = f"=IF({cl}{R4['rev']}=0,0,{cl}{R4['op']}/{cl}{R4['rev']})"
R4["opm"] = row; row += 1

pl_rows = [R4["rev"]] + [R4[l] for l, _, _ in PL_ORDER] + [R4["st_eb"], R4["hq"], R4["eb"], R4["dep"], R4["op"]]
for rr in pl_rows:
    S(ws4, f"H{rr}", bold=True, fmt=MM, border=True, fill=TOT); ws4[f"H{rr}"] = f"=SUM(C{rr}:G{rr})"
    S(ws4, f"B{rr}", size=9, fmt=PCT, border=True, align="center")
    ws4[f"B{rr}"] = f"=IF($G${R4['rev']}=0,0,G{rr}/$G${R4['rev']})"
for k in ["ebm", "opm"]:
    rr = R4[k]
    S(ws4, f"H{rr}", bold=True, fmt=PCT, border=True, align="center", fill=TOT)
    ws4[f"H{rr}"] = f"=H{R4['eb' if k=='ebm' else 'op']}/H{R4['rev']}"

nr = row + 1
S(ws4, f"A{nr}", bold=True, size=10, color=RED)
ws4[f"A{nr}"] = "【計画のポイント】FY2028に連結EBITDA率10%超、FY2029に営業利益率5%超。FY2031はEBITDA率14.0%・営業利益率9.5%。"
S(ws4, f"A{nr+1}", size=9, color="7F7F7F")
ws4[f"A{nr+1}"] = "※ 1号店の初年度は2026年8月3日提出P&Lをそのまま採用。新店初年度は売上90%・開業7.5か月で保守的に計上。"
S(ws4, f"A{nr+2}", size=9, color="7F7F7F")
ws4[f"A{nr+2}"] = "※ ロイヤリティ4%・歩合家賃7%・地代家賃は契約条件どおり全店適用。金利・税金は含まない。"

# ============================================================
# 投資・キャッシュフロー
# ============================================================
ws5 = wb.create_sheet("投資・キャッシュフロー")
ws5.sheet_view.showGridLines = False
ws5.column_dimensions["A"].width = 32
for c in YC: ws5.column_dimensions[c].width = 14
ws5.column_dimensions["H"].width = 14

band(ws5, 1, "5か年事業計画　連結キャッシュフロー（百万円）")
S(ws5, "A2", size=9, color="7F7F7F")
ws5["A2"] = "分割払いスキーム活用により、1店あたり開業時現金支出を247百万円 → 108百万円に圧縮。"
hdr_years(ws5, 3, YC, YEARS, "項　目", ("H", "5年累計"))

# コホート出店数と投資単価列の対応（FY別）
INV_MAP = ["B", "B", "C", "C", "C"]   # FY2027,28 は1〜3号店単価、FY2029以降は4号店以降
r = 4
S(ws5, f"A{r}", border=True); ws5[f"A{r}"] = "新規出店数（店）"
for i, cl in enumerate(YC):
    S(ws5, f"{cl}{r}", color=GREEN, fmt=MM0, border=True, align="center")
    ws5[f"{cl}{r}"] = f"=前提条件!{PC[i]}{OPEN_ROW}"
ROW_OPEN = r; r += 1
S(ws5, f"A{r}", border=True); ws5[f"A{r}"] = "期末店舗数（店）"
for i, cl in enumerate(YC):
    S(ws5, f"{cl}{r}", color=GREEN, fmt=MM0, border=True, align="center")
    ws5[f"{cl}{r}"] = f"=前提条件!{PC[i]}{STORES_ROW}"
r += 1
S(ws5, f"A{r}", border=True); ws5[f"A{r}"] = "初期投資（総額）"
for i, cl in enumerate(YC):
    S(ws5, f"{cl}{r}", fmt=MM, border=True)
    ws5[f"{cl}{r}"] = f"=-{cl}{ROW_OPEN}*前提条件!${INV_MAP[i]}$39"
ROW_INV = r; r += 1
S(ws5, f"A{r}", border=True); ws5[f"A{r}"] = "うち分割対象額"
for i, cl in enumerate(YC):
    S(ws5, f"{cl}{r}", fmt=MM, border=True)
    ws5[f"{cl}{r}"] = f"={cl}{ROW_OPEN}*前提条件!$B$47"
ROW_DEF = r; r += 1
S(ws5, f"A{r}", bold=True, border=True); ws5[f"A{r}"] = "開業時 現金支出"
for cl in YC:
    S(ws5, f"{cl}{r}", bold=True, fmt=MM, border=True, fill=TOT)
    ws5[f"{cl}{r}"] = f"={cl}{ROW_INV}+{cl}{ROW_DEF}"
ROW_CASH = r; r += 1
S(ws5, f"A{r}", border=True); ws5[f"A{r}"] = "分割弁済額"
for i, cl in enumerate(YC):
    S(ws5, f"{cl}{r}", fmt=MM, border=True)
    terms = []
    for b in blocks:
        # 稼働しているコホートのみ弁済。開業年度は稼働月数で按分
        terms.append(
            f"IF(店舗群別PL!{cl}{b['y']}=0,0,前提条件!$B$48*店舗群別PL!$B${b['n']}"
            f"*IF(店舗群別PL!{cl}{b['y']}=1,店舗群別PL!$B${b['m']}/12,1))")
    ws5[f"{cl}{r}"] = "=-(" + "+".join(terms) + ")"
ROW_PAY = r; r += 1
S(ws5, f"A{r}", bold=True, border=True); ws5[f"A{r}"] = "投資キャッシュアウト　計"
for cl in YC:
    S(ws5, f"{cl}{r}", bold=True, fmt=MM, border=True, fill=TOT)
    ws5[f"{cl}{r}"] = f"={cl}{ROW_CASH}+{cl}{ROW_PAY}"
ROW_ICF = r; r += 1
S(ws5, f"A{r}", bold=True, border=True); ws5[f"A{r}"] = "連結EBITDA"
for cl in YC:
    S(ws5, f"{cl}{r}", bold=True, color=GREEN, fmt=MM, border=True)
    ws5[f"{cl}{r}"] = f"=5カ年連結PL!{cl}{R4['eb']}"
ROW_EB = r; r += 1
S(ws5, f"A{r}", bold=True, border=True); ws5[f"A{r}"] = "簡易フリーキャッシュフロー"
for cl in YC:
    S(ws5, f"{cl}{r}", bold=True, fmt=MM, border=True, fill=KEY)
    ws5[f"{cl}{r}"] = f"={cl}{ROW_EB}+{cl}{ROW_ICF}"
ROW_FCF = r; r += 1
S(ws5, f"A{r}", bold=True, border=True); ws5[f"A{r}"] = "累計フリーキャッシュフロー"
for i, cl in enumerate(YC):
    S(ws5, f"{cl}{r}", bold=True, fmt=MM, border=True, fill=TOT)
    ws5[f"{cl}{r}"] = f"={cl}{ROW_FCF}" if i == 0 else f"={YC[i-1]}{r}+{cl}{ROW_FCF}"
ROW_CUM = r
for rr in [ROW_OPEN, ROW_INV, ROW_DEF, ROW_CASH, ROW_PAY, ROW_ICF, ROW_EB, ROW_FCF]:
    S(ws5, f"H{rr}", bold=True, fmt=MM0 if rr == ROW_OPEN else MM, border=True, fill=TOT)
    ws5[f"H{rr}"] = f"=SUM(C{rr}:G{rr})"

nr = ROW_CUM + 2
sec(ws5, nr, "■ 分割払いスキーム")
S(ws5, f"A{nr+1}", size=9, color="7F7F7F")
ws5[f"A{nr+1}"] = "内装工事費100百万円（月833千円×120回・10年）／保証金38.9百万円（月463千円×84回・7年）。1店あたり開業時現金支出 247→108百万円。"
S(ws5, f"A{nr+2}", size=9, color="7F7F7F")
ws5[f"A{nr+2}"] = "分割弁済は資産計上済みの繰延支払（損益は減価償却で費用化済み）のため、P/Lには影響させずCF上のみ計上。"
S(ws5, f"A{nr+3}", size=9, color="7F7F7F")
ws5[f"A{nr+3}"] = "※ 簡易FCF＝連結EBITDA－投資キャッシュアウト。法人税・運転資本増減・借入/返済・分割払い金利相当は含まない。"

nr += 5
sec(ws5, nr, "■ 単店の投資回収（1〜3号店・分割払い活用ベース）")
hdr_years(ws5, nr + 1, YC, ["1年目", "2年目", "3年目", "4年目", "5年目"], "項　目")
mc = ["C", "D", "E", "E", "E"]   # 単店モデルPLの参照列
rr = nr + 2
S(ws5, f"A{rr}", border=True); ws5[f"A{rr}"] = "EBITDA"
for i, cl in enumerate(YC):
    S(ws5, f"{cl}{rr}", color=GREEN, fmt=MM, border=True)
    ws5[f"{cl}{rr}"] = f"=単店モデルPL!{mc[i]}{EB2}"
E1 = rr; rr += 1
S(ws5, f"A{rr}", border=True); ws5[f"A{rr}"] = "分割弁済"
for cl in YC:
    S(ws5, f"{cl}{rr}", fmt=MM, border=True); ws5[f"{cl}{rr}"] = "=-前提条件!$B$48"
E2 = rr; rr += 1
S(ws5, f"A{rr}", bold=True, border=True); ws5[f"A{rr}"] = "単店FCF"
for cl in YC:
    S(ws5, f"{cl}{rr}", bold=True, fmt=MM, border=True, fill=TOT)
    ws5[f"{cl}{rr}"] = f"={cl}{E1}+{cl}{E2}"
E3 = rr; rr += 1
S(ws5, f"A{rr}", bold=True, border=True); ws5[f"A{rr}"] = "累計FCF（開業時支出△108含む）"
for i, cl in enumerate(YC):
    S(ws5, f"{cl}{rr}", bold=True, fmt=MM, border=True, fill=KEY)
    base_ = "-(前提条件!$B$39-前提条件!$B$47)"
    ws5[f"{cl}{rr}"] = f"={base_}+{cl}{E3}" if i == 0 else f"={YC[i-1]}{rr}+{cl}{E3}"
E4 = rr
S(ws5, f"A{E4+2}", bold=True, size=10, color=RED)
ws5[f"A{E4+2}"] = "→ 開業3年目に累計FCFが黒字転換（投資回収 約2.8年）。保証金39百万円は退去時返還対象のため実質回収はさらに短期。"
S(ws5, f"A{E4+3}", size=9, color="7F7F7F")
ws5[f"A{E4+3}"] = "※ 分割払いを使わない場合は開業時支出247百万円・回収約3.3年。定常期の年間FCFは60百万円で2号店以降の出店原資となる。"

# ============================================================
# サマリー（先頭）
# ============================================================
ws1 = wb.create_sheet("サマリー", 0)
ws1.sheet_view.showGridLines = False
ws1.column_dimensions["A"].width = 30
ws1.column_dimensions["B"].width = 20
for c in YC: ws1.column_dimensions[c].width = 14
ws1.column_dimensions["H"].width = 14

band(ws1, 1, "BRASSERIE Thierry Marx（仮称）　5か年事業計画　FY2027-FY2031")
S(ws1, "A2", size=9, color="7F7F7F")
ws1["A2"] = "単位：百万円　／　FY＝4月〜翌3月　／　六本木ミッドタウン店 2026年8月3日提出P&L 準拠"

sec(ws1, 4, "■ 出店ロードマップ")
road = [("FY2027", "六本木ミッドタウン 1店（2027年4月開業）", "1店"),
        ("FY2028", "札幌・大阪 2店", "3店"),
        ("FY2029", "首都圏・他政令指定都市 2店", "5店"),
        ("FY2030", "リゾート含め 年2店", "7店"),
        ("FY2031", "リゾート含め 年2店", "9店")]
r = 5
for y, t, n in road:
    S(ws1, f"A{r}", bold=True); ws1[f"A{r}"] = y
    S(ws1, f"B{r}"); ws1[f"B{r}"] = t
    S(ws1, f"E{r}", bold=True, color=RED); ws1[f"E{r}"] = f"期末 {n}"
    r += 1

sec(ws1, 11, "■ 連結業績サマリー")
hdr_years(ws1, 12, YC, YEARS, "項　目", ("H", "5年累計"))
summary = [("期末店舗数（店）", "stores", MM0, False),
           ("売上高", "rev", MM, False),
           ("店舗EBITDA　計", "st_eb", MM, False),
           ("本部費", "hq", MM, False),
           ("連結EBITDA", "eb", MM, True),
           ("　EBITDA率", "ebm", PCT, True),
           ("減価償却費", "dep", MM, False),
           ("営業利益", "op", MM, True),
           ("　営業利益率", "opm", PCT, True)]
r = 13
for lab, key, fmt, hi in summary:
    S(ws1, f"A{r}", bold=hi, border=True); ws1[f"A{r}"] = lab
    for cl in YC:
        S(ws1, f"{cl}{r}", bold=hi, color=GREEN, fmt=fmt, border=True,
          fill=KEY if hi else None, align="center" if fmt in (PCT, MM0) else None)
        ws1[f"{cl}{r}"] = f"=5カ年連結PL!{cl}{R4[key]}"
    if key != "stores":
        S(ws1, f"H{r}", bold=True, color=GREEN, fmt=fmt, border=True, fill=TOT,
          align="center" if fmt == PCT else None)
        ws1[f"H{r}"] = f"=5カ年連結PL!H{R4[key]}"
    r += 1

S(ws1, "A23", bold=True, size=11, color=RED)
ws1["A23"] = "【目標達成】FY2028に連結EBITDA率10%超　／　FY2029に営業利益率5%超　→　以降も継続的に改善"

sec(ws1, 25, "■ 単店モデル（六本木ミッドタウン）")
hdr_years(ws1, 26, ["C", "D", "E"], ["1年目", "2年目", "3年目以降"], "項　目")
smap = [("売上高", REV_R, MM), ("EBITDA", EB2, MM), ("　EBITDA率", OP2 + 1, PCT),
        ("営業利益", OP2, MM), ("　営業利益率", OP2 + 2, PCT)]
r = 27
for lab, src, fmt in smap:
    S(ws1, f"A{r}", bold=fmt == PCT, border=True); ws1[f"A{r}"] = lab
    for cl, mcol in zip(["C", "D", "E"], ["C", "D", "E"]):
        S(ws1, f"{cl}{r}", bold=fmt == PCT, color=GREEN, fmt=fmt, border=True,
          fill=KEY if fmt == PCT else None, align="center" if fmt == PCT else None)
        ws1[f"{cl}{r}"] = f"=単店モデルPL!{mcol}{src}"
    r += 1

sec(ws1, 33, "■ 投資・キャッシュフロー サマリー")
hdr_years(ws1, 34, YC, YEARS, "項　目", ("H", "5年累計"))
imap = [("新規出店数（店）", ROW_OPEN, MM0), ("初期投資（総額）", ROW_INV, MM),
        ("開業時 現金支出", ROW_CASH, MM), ("投資キャッシュアウト 計", ROW_ICF, MM),
        ("簡易フリーキャッシュフロー", ROW_FCF, MM), ("累計フリーキャッシュフロー", ROW_CUM, MM)]
r = 35
for lab, src, fmt in imap:
    S(ws1, f"A{r}", border=True); ws1[f"A{r}"] = lab
    for cl in YC:
        S(ws1, f"{cl}{r}", color=GREEN, fmt=fmt, border=True, align="center" if fmt == MM0 else None)
        ws1[f"{cl}{r}"] = f"=投資・キャッシュフロー!{cl}{src}"
    if src != ROW_CUM:
        S(ws1, f"H{r}", bold=True, color=GREEN, fmt=fmt, border=True, fill=TOT,
          align="center" if fmt == MM0 else None)
        ws1[f"H{r}"] = f"=投資・キャッシュフロー!H{src}"
    r += 1

S(ws1, "A42", size=9, color="7F7F7F")
ws1["A42"] = "※ 分割払いスキームにより5年間の現金投資額を21.9億円 → 12.9億円に圧縮。FY2030に単年FCF黒字化、5年累計FCFはほぼ均衡。"
S(ws1, "A43", size=9, color="7F7F7F")
ws1["A43"] = "※ 保証金（9店合計350百万円）は退去時返還対象。"

sec(ws1, 45, "■ 経費改善レバー（提出P&L → 3年目以降）")
lev = [("売上原価", "29.5% → 28.5%", "直輸入ワイン活用・複数店一括仕入"),
       ("人件費", "34.5% → 32.5%", "本部集中化・多能工化"),
       ("業務委託費・消耗品他", "2.4% → 2.0%", "本部集中購買・規格統一"),
       ("水道光熱費", "4.0% → 3.6%", "高効率厨房機器・営業時間最適化"),
       ("ロイヤリティ／歩合家賃", "4.0%／7.0%（据置）", "契約条件どおり全店適用")]
r = 46
for a, b, c_ in lev:
    S(ws1, f"A{r}", size=9, border=True); ws1[f"A{r}"] = a
    S(ws1, f"B{r}", size=9, bold=True, border=True, align="center"); ws1[f"B{r}"] = b
    ws1.merge_cells(f"C{r}:H{r}")
    S(ws1, f"C{r}", size=9, border=True); ws1[f"C{r}"] = c_
    r += 1
S(ws1, f"A{r+1}", size=9, color="7F7F7F")
ws1[f"A{r+1}"] = "出典：「THIERRY MARX × WineBank 事業収支CF追加（2026年8月4日版）」および 2026年8月3日提出P&L。金利・税金は含まない。"

if "Sheet" in wb.sheetnames:
    del wb["Sheet"]

# --- シート参照を引用符付きに正規化（数字始まり・記号入りシート名対策） ---
SHEET_NAMES = ["前提条件", "単店モデルPL", "店舗群別PL", "5カ年連結PL", "投資・キャッシュフロー"]
for sh in wb.worksheets:
    for row in sh.iter_rows():
        for c in row:
            if isinstance(c.value, str) and c.value.startswith("="):
                f = c.value
                for nm in SHEET_NAMES:
                    f = f.replace(f"'{nm}'!", f"{nm}!")   # 既存の引用を一旦外す
                for nm in SHEET_NAMES:
                    f = f.replace(f"{nm}!", f"'{nm}'!")
                c.value = f

wb.calculation.fullCalcOnLoad = True
wb.save(OUT)
print("saved:", OUT)
